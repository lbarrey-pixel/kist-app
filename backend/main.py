import os, csv, io, re, time, base64 as _b64
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Depends, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
from typing import Optional, List
import anthropic
from supabase import create_client
from datetime import date
import extract_msg
from google.oauth2 import id_token
from google.auth.transport import requests as google_requests
import requests

# Motor de busca de preços na internet (Frente A) — módulo separado de propósito:
# o /gerar-csv e o fmt_preco são invariantes intocáveis, e um subsistema grande no
# meio do monólito é onde eles morrem. Import protegido: se o arquivo faltar no
# deploy, o app sobe normal e só o /ficha-internet avisa que o motor não está lá.
try:
    from motor_precos import (
        resolver_ficha as _resolver_ficha_precos,
        aprender_no_internet as _mp_aprender_no_internet,
        aprender_interpretacao as _mp_aprender_interpretacao,
        tem_no_internet as _mp_tem_no_internet,
    )
    _MOTOR_PRECOS_OK = True
except Exception:
    _resolver_ficha_precos = None
    _mp_aprender_no_internet = None
    _mp_aprender_interpretacao = None
    _mp_tem_no_internet = None
    _MOTOR_PRECOS_OK = False

# ── Ingestão canônica (v3.25) ────────────────────────────────────────────────
# Protegido igual ao motor: se o módulo faltar, o boot NÃO quebra e o /extrair
# cai no caminho antigo. Produção não pode ficar de pé só se um arquivo novo
# subiu junto.
try:
    from ingestao import (
        ler_msg as _ing_ler_msg,
        ler_email as _ing_ler_email,
        documento_de_texto as _ing_doc_texto,
        documento_de_arquivo as _ing_doc_arquivo,
        montar_payload as _ing_montar_payload,
        consolidar_itens as _ing_consolidar,
    )
    _INGESTAO_OK = True
except Exception:
    _ing_ler_msg = _ing_ler_email = _ing_doc_texto = None
    _ing_doc_arquivo = _ing_montar_payload = _ing_consolidar = None
    _INGESTAO_OK = False

# Versão do backend. O núcleo do Analista guarda a versão que ele descreve; se as
# duas divergirem, o agente é avisado de que o conhecimento dele está atrasado.
# Conhecimento velho não avisa que é velho — ele responde com a mesma confiança
# e erra. Este número é a única coisa que impede isso.
VERSAO_BACKEND = "3.24"

app = FastAPI(title="Kist Cotações API", version=VERSAO_BACKEND)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])


@app.exception_handler(Exception)
async def _erro_nao_tratado(request: Request, exc: Exception):
    """Erro não tratado precisa voltar COM cabeçalho CORS.

    Sem isso, o 500 sai sem os headers do CORSMiddleware, o navegador bloqueia a
    resposta e o operador vê só "Failed to fetch" — escondendo a mensagem real.
    O traceback continua indo pro log do Render normalmente.
    """
    return JSONResponse(
        status_code=500,
        content={"detail": f"{type(exc).__name__}: {exc}"},
        headers={"Access-Control-Allow-Origin": "*"},
    )

# ── Auth ──────────────────────────────────────────────────────────────────────
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID", "822792475898-4l9ctl5jc1urpi2tvbuaut2tpelgevfo.apps.googleusercontent.com")
USUARIOS_PERMITIDOS = set(os.environ.get("USUARIOS_PERMITIDOS", "leonardobarrey@gmail.com,thiagokist@gmail.com,fabiokist@gmail.com").split(","))
security = HTTPBearer()
_token_cache: dict = {}

def _verificar_token_str(token: str) -> str:
    if token in _token_cache:
        return _token_cache[token]
    try:
        info = id_token.verify_oauth2_token(token, google_requests.Request(), GOOGLE_CLIENT_ID, clock_skew_in_seconds=30)
        email = info.get("email", "").lower()
        if email not in USUARIOS_PERMITIDOS:
            raise HTTPException(status_code=403, detail=f"Acesso negado para {email}")
        _token_cache[token] = email
        if len(_token_cache) > 200:
            _token_cache.clear()
        return email
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=401, detail=f"Token inválido: {str(e)}")

def verificar_token(credentials: HTTPAuthorizationCredentials = Depends(security)):
    return _verificar_token_str(credentials.credentials)

# ── Clientes singleton ────────────────────────────────────────────────────────
ANTHROPIC_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
SUPABASE_URL  = os.environ.get("SUPABASE_URL", "https://owpmcoithvzdlhmfkvbe.supabase.co")
SUPABASE_KEY  = os.environ.get("SUPABASE_KEY", "")
_claude_client = None
_supabase_client = None

def get_claude():
    global _claude_client
    if _claude_client is None:
        _claude_client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
    return _claude_client

def get_supabase():
    global _supabase_client
    if _supabase_client is None:
        _supabase_client = create_client(SUPABASE_URL, SUPABASE_KEY)
    return _supabase_client

# ── Prompts ───────────────────────────────────────────────────────────────────
SYSTEM_EXTRACAO = """Você é o assistente comercial da Kist Soluções em Telecom e Energia.
Extraia itens de cotação de e-mails, textos, planilhas ou imagens e retorne JSON.

RETORNE APENAS JSON VÁLIDO. Sem markdown, sem ```json, sem ```. Só o objeto JSON puro.

Formato — SEMPRE retorne um array de propostas, mesmo que seja apenas uma:
{
  "propostas": [
    {
      "titulo": "identificador curto da demanda (ex: 'SC 18712 Rack', 'RC 45321', 'Planilha 1')",
      "cliente": "NOME DO CLIENTE",
      "cnpj": "XX.XXX.XXX/XXXX-XX ou null",
      "rc_neg": "RC XXXXX ou NEG-XXXXXXX ou null",
      "itens": [
        {
          "descricao": "descrição comercial curta — máx 120 chars",
          "descricao_original": "texto exato do cliente, preservado integralmente",
          "codigo_cliente": "código do item no sistema do cliente, ou null",
          "specs_complementares": "specs técnicas ou PN se presentes, senão null",
          "quantidade": 1,
          "unidade": "UN",
          "sugerir_pn": false
        }
      ]
    }
  ]
}

QUANDO CRIAR MÚLTIPLAS PROPOSTAS:
- LOCAL DE ENTREGA DIFERENTE = PROPOSTA DIFERENTE. Esta é a regra mais forte e ela
  vem ANTES de todas as outras. Cada endereço/base/filial de entrega vira uma proposta
  própria, porque cada uma vira um frete e um faturamento separados.
  • Duas requisições diferentes para o MESMO destino = UMA proposta só.
  • Uma requisição com DOIS destinos = DUAS propostas.
  • O que manda é o destino, nunca o número da requisição.
- O conteúdo traz seções claramente separadas por arquivo/planilha/aba → uma proposta por seção
- O e-mail menciona explicitamente múltiplos projetos, RCs ou solicitações distintas → uma por demanda
- Lista de itens é única, com destino único (mesmo que longa) → uma única proposta
- Quando quebrar por destino: "cliente", "cnpj" e "rc_neg" são os MESMOS em todas as
  propostas (é um pedido só, do mesmo cliente). O "titulo" é que muda — use o destino
  como título (ex.: "Presidente Prudente", "Base João Dias — SP"), porque é ele que o
  operador vê na aba.
- O MESMO item pode aparecer em propostas diferentes com quantidades diferentes. Isso é
  normal e correto: são entregas distintas. NUNCA junte, NUNCA descarte como duplicata.

REGRAS DE DESCRIÇÃO:
- "descricao": sempre curta e comercial. Formato: [Categoria] [Marca/Modelo] [Spec principal]
- "descricao_original": preservar EXATAMENTE como veio, sem alterar nada
- "specs_complementares": preencher quando original for longa ou tabela; incluir PN/código se presente
- "sugerir_pn": true SÓ para itens de alto valor (notebook, servidor, switch gerenciável, UPS,
  câmera IP, storage) sem modelo específico definido. Commodities → sempre false

REGRAS DE CNPJ (o campo mais negligenciado — leia com atenção):
- O CNPJ do CLIENTE é OBRIGATÓRIO sempre que existir no material. Sem ele a proposta não pode
  ser exportada nem alimentar o banco de preços. Procure ATIVAMENTE, nesta ordem:
  1. Corpo do e-mail e ASSINATURA do remetente (é o lugar mais comum: bloco no rodapé com
     razão social, CNPJ, IE, endereço e telefone)
  2. Cabeçalho "De:" — o domínio do remetente identifica a empresa (ex.: @convergint.com).
     O domínio NÃO é o CNPJ, mas confirma de quem é a demanda quando houver mais de uma empresa
     no material.
  3. RODAPÉ e CABEÇALHO de PDFs anexos (pedido de compra, RC, cotação): quase sempre trazem
     o CNPJ do emissor
  4. Planilhas: procure células com "CNPJ", "C.N.P.J.", "Cadastro Nacional"
  5. Imagens e prints: leia rodapés e timbres
- Formato: 14 dígitos. Devolva como "XX.XXX.XXX/XXXX-XX". Aceite variações na origem
  (só números, com pontos, com espaços) e normalize.
- CUIDADO — há SEMPRE dois CNPJs em jogo. O da KIST é 10.573.732/0003-96 (ou qualquer CNPJ
  com raiz 10573732): esse é o NOSSO, é o fornecedor, e NUNCA deve ir no campo "cnpj".
  O que você quer é o CNPJ de QUEM PEDIU a cotação.
- Se houver mais de um CNPJ de cliente (matriz e filial), escolha o que emitiu a demanda —
  normalmente o da assinatura de quem enviou, ou o do cabeçalho do pedido.
- NUNCA invente, complete ou "conserte" um CNPJ. Se não encontrar, devolva null. CNPJ errado
  é pior que CNPJ ausente: ele amarra o preço à empresa errada e ninguém percebe.

REGRAS GERAIS:
- Extraia TODOS os itens, inclusive de imagens/prints
- IMAGENS EMBUTIDAS NO E-MAIL: muitos clientes colam a tabela de itens como IMAGEM no
  corpo, e não como texto. Leia cada imagem. Junto delas vêm logo, banner e faixa de
  assinatura, que não têm item nenhum — ignore essas sem comentar. Nunca invente item a
  partir de logo, telefone, slogan ou selo de rodapé.
- Quando uma imagem vier precedida de "[imagem de: <assunto>]", ela pertence AQUELE
  e-mail. Com vários e-mails na mesma leitura, use esse rótulo para saber de quem é
  cada tabela — o assunto costuma trazer o cliente e o CNPJ.
- quantidade = número, nunca string
- Em tabelas com coluna Qtd/Quantidade/QTDE: leia a célula exata da coluna. A quantidade NÃO
  é o número da linha nem o código do item. Verifique cada linha antes de confirmar.
- Em tabelas com coluna PN/Código/Nº do item/SKU: inclua em specs_complementares como "PN: XXXXX"

CÓDIGO DO CLIENTE (campo "codigo_cliente"):
- É o código do item no ERP/sistema DO CLIENTE — a coluna "Item", "Código", "Material",
  "Nº do item". Exemplos reais: "UC.107572", "AF.101766", "ES.102241", "MAT-4471".
- Copie EXATAMENTE como está, sem reformatar, sem completar, sem corrigir.
- NÃO confunda com o número da LINHA (1, 2, 3...) nem com o número da requisição.
- NÃO é o PN do fabricante nem o nosso SKU. Se o material só traz o PN do fabricante e
  nenhum código do cliente, devolva null e deixe o PN em specs_complementares.
- Não existe código do cliente na maioria das cotações. Ausente → null. NUNCA invente:
  este campo é usado para copiar preço entre propostas, e um código errado copia o
  preço do item errado.

LOCAL DE ENTREGA:
- Quando o material informar destino por item, escreva-o SEMPRE no FIM de
  specs_complementares, no formato "Entrega: <texto do destino como veio>".
- O destino nunca entra na "descricao" nem na "descricao_original" — ele não é o produto.
"""

SYSTEM_MATCHING = """Você é especialista em materiais elétricos, telecom, infraestrutura e TI.
Sua tarefa é decidir, para cada item solicitado, se algum produto do banco de preços é O MESMO ITEM
— e explicar por quê, como um colega experiente explicaria pra outro antes de fechar a venda.

O item solicitado tem DUAS partes e as duas valem:
  - DESCRIÇÃO: o texto principal do cliente
  - SPECS: as especificações complementares dele (bitola, dimensão, norma, modelo, acabamento…)
É MUITO comum a descrição bater e a SPEC divergir. Quando isso acontece NÃO é o mesmo item —
esse erro faz a Kist vender errado, comprar errado, entregar e o cliente devolver.

RETORNE APENAS JSON VÁLIDO. Sem markdown. Só o objeto JSON puro.

Formato:
{
  "matches": [
    {
      "indice": 0,
      "banco_descricao": "descrição exata do item do banco que corresponde, ou null se nenhum",
      "banco_preco": 0.00,
      "banco_proposta": "número da proposta de referência ou null",
      "confianca": "alta/media/baixa/nenhuma",
      "veredito": "mesmo | diferente | inconclusivo",
      "motivo": "uma frase: por que é ou por que não é",
      "diferencas": ["atributo que diverge: o cliente pede X, o banco tem Y"],
      "falta": "qual especificação está faltando pra decidir, ou null"
    }
  ]
}

COMO PREENCHER veredito / diferencas / falta:
- "mesmo": é o mesmo item. diferencas = [], falta = null.
- "diferente": encontrei um candidato PARECIDO mas NÃO é o mesmo. Liste em "diferencas" cada
  atributo que diverge, sempre no formato "o cliente pede X, o banco tem Y". Seja concreto:
  "cliente pede 2,5mm², o banco tem 4mm²" — não "bitola diferente".
  Devolva o candidato mesmo assim (o operador quer ver de onde veio a dúvida), com confianca baixa.
- "inconclusivo": falta informação pra decidir. Diga em "falta" EXATAMENTE qual especificação
  resolveria — ex.: "o cliente não informou a bitola" ou "o banco não diz o comprimento".
  Isso é acionável: o operador vai perguntar. Não chute.
- Se não houver candidato nenhum que preste: banco_descricao = null, confianca = "nenhuma",
  veredito = "diferente", e explique em "motivo" o que o banco tem de mais próximo.

REGRAS DE MATCHING — seja rigoroso:
- "alta": mesmo produto, mesma especificação, mesmo fabricante (quando mencionado)
- "media": mesmo produto e spec, fabricante não mencionado no pedido (qualquer fabricante atende)
- "baixa": produto similar mas com dúvida — retornar mas sinalizar
- "nenhuma": produto diferente ou fabricante diferente do solicitado → banco_descricao = null, banco_preco = 0

ATENÇÃO especial:
- Fabricantes DIFERENTES = nenhuma. Ex: pedido "Wetzel" + banco "Tramontina" → nenhuma
- Categoria diferente = nenhuma. Ex: pedido "canaleta" + banco "tampa para canaleta" → nenhuma  
- Tamanho/bitola/dimensão diferente = nenhuma. Ex: pedido "6mm²" + banco "4mm²" → nenhuma
- Cor nunca diferencia preço em cabos elétricos. Ex: "cabo 6mm amarelo" = "cabo 6mm azul" → alta
- Quando pedido não menciona fabricante e banco tem fabricante → pode ser media se spec bater
- Prefira retornar nenhuma a retornar match errado — falso negativo é melhor que falso positivo
"""

# Mapa de atributos excludentes por categoria — legível em RUNTIME (config_kist),
# para o operador ajustar uma regra (ex.: "polos são excludentes em disjuntor")
# sem redeploy. O fallback embutido garante que o matching nunca fica sem as
# regras críticas se a config sumir.
_EXCLUDENTES_FALLBACK = """REGRAS DE ATRIBUTOS EXCLUDENTES por categoria (o que NÃO pode divergir):
- DISJUNTOR: polos (1P/monopolar/unipolar · 2P/bipolar · 3P · 4P), amperagem, curva (B/C/D), tensão. Polos diferentes = NÃO é o mesmo.
- CABO DE REDE: tipo (UTP/FTP/SFTP), categoria (Cat5e/Cat6/Cat6a). Cor não importa.
- CABO ELÉTRICO: bitola (mm²), tensão (450/750V). Cor não importa.
- ÓPTICO: conector (SC/LC), polimento (APC/UPC), modo (SM/MM), nº de fibras.
- Qualquer item: dimensão, tensão, amperagem/potência, fabricante+modelo. MPN igual confirma; MPN diferente descarta."""


def _excludentes_matching(sb) -> str:
    """Regras de excludente por categoria, lidas da config (fallback embutido)."""
    try:
        r = (sb.table("config_kist").select("valor")
             .eq("chave", "excludentes_matching").limit(1).execute())
        if r.data and (r.data[0].get("valor") or "").strip():
            return "\n\n" + r.data[0]["valor"]
    except Exception:
        pass
    return "\n\n" + _EXCLUDENTES_FALLBACK


# ── Helpers ───────────────────────────────────────────────────────────────────
def _pede_atencao(descricao: str, specs: str) -> bool:
    """Sinaliza item que provavelmente precisa de consulta técnica: sem PN/código
    aparente em nenhum dos lados. É só um DESTAQUE — o botão Conferir existe em
    todo item, porque quem sabe se tem dúvida é o operador, não a heurística.

    Substitui o _validar_sugerir_pn, que VETAVA o botão por duas regras erradas:
      1. lista de commodities por substring — 'SUPORTE' matava o botão no item que
         voltou em RMA; 'CABO' matava em cabo de fabricante específico;
      2. premissa de que item barato não merece consulta — nasceu de um veto de
         custo que hoje é $1,37/mês no sistema inteiro.
    """
    txt = f"{descricao or ''} {specs or ''}".upper()
    pn_patterns = [r'\b[A-Z]{2,}\s*\d{3,}\b', r'\b\d{3,}[A-Z]{1,3}\b', r'\b[A-Z]+-\d+[A-Z]*\b',
                   r'\b[A-Z]{4,}\d{2,}\b', r'\b\d{4,}\b']
    return not any(re.search(pat, txt) for pat in pn_patterns)


def _alerta_do_candidato(banco_desc, todos_candidatos):
    """Retorna dict do alerta do produto que casou, ou None."""
    if not banco_desc:
        return None
    for cand in todos_candidatos:
        if cand.get("descricao") == banco_desc:
            raw = cand.get("alerta")
            if not raw:
                return None
            try:
                import json as _jal
                return _jal.loads(raw)
            except Exception:
                return {"texto": str(raw)}
    return None


#  _prefiltro_candidatos foi REMOVIDO na v3.14.
#  Ele filtrava por sobreposição de palavras DENTRO da janela de 500 linhas mais
#  recentes carregada pelo _fazer_matching — e essa janela enxergava 0,8% do banco.
#  Medição contra 636 pares reais (recall@40, banco inteiro):
#      palavras  85,8%   ·   trigrama  95,1%   ·   híbrido dos dois  93,3%
#  O trigrama contém as palavras: dos casos que ele acha, as palavras perdiam 14;
#  o contrário só 1. Menos código e melhor resultado — hoje é candidatos_trgm_lote().

# ── Endpoints ─────────────────────────────────────────────────────────────────
@app.get("/health")
def health():
    return {"status": "ok"}

@app.get("/ping")
def ping():
    return {"pong": True}

@app.get("/proxima-proposta")
def proxima_proposta(usuario: str = Depends(verificar_token)):
    """Retorna o próximo número de proposta disponível.
    Fonte primária: propostas.numero_proposta (mais confiável).
    Fallback: produtos.proposta_tiny (banco de preços).
    """
    sb = get_supabase()
    def _max_numeros(registros, campo):
        nums = []
        for r in registros:
            try:
                v = r.get(campo, "")
                if v:
                    nums.append(int(str(v).strip()))
            except Exception:
                pass
        return max(nums) if nums else None

    try:
        # Fonte 1: tabela de propostas
        res = sb.table("propostas").select("numero_proposta") \
            .not_.is_("numero_proposta", "null") \
            .order("numero_proposta", desc=True).limit(20).execute()
        maximo = _max_numeros(res.data or [], "numero_proposta")
        if maximo:
            return {"proximo": str(maximo + 1)}
    except Exception:
        pass

    try:
        # Fallback: banco de preços
        res = sb.table("produtos").select("proposta_tiny") \
            .not_.is_("proposta_tiny", "null") \
            .order("proposta_tiny", desc=True).limit(50).execute()
        maximo = _max_numeros(res.data or [], "proposta_tiny")
        if maximo:
            return {"proximo": str(maximo + 1)}
    except Exception:
        pass

    return {"proximo": ""}

@app.get("/banco/stats")
def banco_stats(usuario: str = Depends(verificar_token)):
    sb = get_supabase()
    try:
        total = sb.table('produtos').select('id', count='exact').execute()
        from datetime import datetime, timedelta
        data_limite = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
        desatual = sb.table('produtos').select('id', count='exact')\
            .lt('data_ref', data_limite).gt('preco_un', 0).execute()
        return {"total_produtos": total.count, "desatualizados_90d": desatual.count}
    except Exception as e:
        return {"erro": str(e)}

def _media_type_img(b: bytes) -> str:
    """Detecta o tipo REAL da imagem pelos bytes (magic numbers).

    O Outlook nomeia toda imagem inline de 'image.png' independente do formato
    real, e a API da Anthropic rejeita (400) quando o media_type declarado não
    bate com o conteúdo. Nunca confiar na extensão do arquivo.
    """
    b = b or b""
    if b[:8] == b"\x89PNG\r\n\x1a\n":            return "image/png"
    if b[:3] == b"\xff\xd8\xff":                 return "image/jpeg"
    if b[:6] in (b"GIF87a", b"GIF89a"):          return "image/gif"
    if b[:4] == b"RIFF" and b[8:12] == b"WEBP":  return "image/webp"
    return "image/png"


def _img_dim(b: bytes) -> tuple:
    """(largura, altura) lendo SÓ o cabeçalho — sem Pillow, sem decodificar.

    Serve para orçar o custo em tokens e descartar espaçador/pixel de rastreio.
    Não conhece o formato → (0, 0), e quem chama trata como "não sei" (mantém a
    imagem). Preferir errar mandando a imagem a errar descartando a cotação.
    """
    b = b or b""
    try:
        if b[:8] == b"\x89PNG\r\n\x1a\n" and b[12:16] == b"IHDR":
            return (int.from_bytes(b[16:20], "big"), int.from_bytes(b[20:24], "big"))
        if b[:6] in (b"GIF87a", b"GIF89a"):
            return (int.from_bytes(b[6:8], "little"), int.from_bytes(b[8:10], "little"))
        if b[:4] == b"RIFF" and b[8:12] == b"WEBP":
            if b[12:16] == b"VP8X":
                return (int.from_bytes(b[24:27], "little") + 1,
                        int.from_bytes(b[27:30], "little") + 1)
            if b[12:16] == b"VP8 ":
                return (int.from_bytes(b[26:28], "little") & 0x3FFF,
                        int.from_bytes(b[28:30], "little") & 0x3FFF)
            if b[12:16] == b"VP8L":
                n = int.from_bytes(b[21:25], "little")
                return ((n & 0x3FFF) + 1, ((n >> 14) & 0x3FFF) + 1)
        if b[:3] == b"\xff\xd8\xff":
            i, n = 2, len(b)
            while i + 9 < n:
                if b[i] != 0xFF:
                    i += 1
                    continue
                marc = b[i + 1]
                if marc in (0xC0, 0xC1, 0xC2, 0xC3, 0xC5, 0xC6, 0xC7,
                            0xC9, 0xCA, 0xCB, 0xCD, 0xCE, 0xCF):
                    return (int.from_bytes(b[i + 7:i + 9], "big"),
                            int.from_bytes(b[i + 5:i + 7], "big"))
                if marc in (0xD8, 0xD9) or 0xD0 <= marc <= 0xD7:
                    i += 2
                    continue
                i += 2 + int.from_bytes(b[i + 2:i + 4], "big")
    except Exception:
        pass
    return (0, 0)


# Orçamento de imagem na extração. O teto ANTIGO era "4 imagens, maior primeiro"
# — heurística que se inverte: no e-mail do Grupo Cesari a assinatura pesa 27 KB e
# a tabela de itens pesa 1,4 KB. Agora o teto é por TOKEN (o que de fato custa) e a
# ORDEM vem da posição no HTML do e-mail (corpo antes da assinatura).
_IMG_MAX_N       = 16     # teto duro de imagens por chamada
_IMG_MAX_TOKENS  = 6000   # ~US$ 0,018 de entrada no Sonnet, por extração
_IMG_MIN_ALTURA  = 12     # px — abaixo disso é faixa/espaçador, nunca tabela
_IMG_MIN_LARGURA = 40     # px — abaixo disso é bullet/ícone
_IMG_MIN_AREA    = 1500   # px² — pixel de rastreio


def _img_tokens(b: bytes) -> int:
    """Estimativa de tokens de uma imagem (~ área/750, com o teto de 1568px da API).

    Dimensão desconhecida → estimativa conservadora pelo tamanho em bytes, para o
    orçamento não ser burlado por um formato que o _img_dim não lê.
    """
    w, h = _img_dim(b)
    if not w or not h:
        return max(300, min(1600, len(b or b"") // 40))
    maior = max(w, h)
    if maior > 1568:                      # a API redimensiona antes de tokenizar
        fator = 1568 / maior
        w, h = w * fator, h * fator
    return max(50, int((w * h) / 750))


def _img_descartavel(b: bytes) -> bool:
    """True só para o que NÃO PODE ser conteúdo: espaçador, filete, pixel.

    Deliberadamente estreito. Não tenta separar banner de tabela por geometria —
    no Cesari o banner da assinatura é 385x33 e a tabela de itens é 388x38. Não há
    corte geométrico entre os dois, e chutar aqui é o que apagava a cotação.
    """
    w, h = _img_dim(b)
    if not w or not h:
        return False
    return h < _IMG_MIN_ALTURA or w < _IMG_MIN_LARGURA or (w * h) < _IMG_MIN_AREA


_PROD_COLS = ("id,descricao,preco_un,proposta_tiny,data_ref,alerta,"
              "preco_custo,link_fornecedor,fornecedor,fornecedor_canal,"
              "fornecedor_contato,sku_fornecedor,cliente,cnpj,"
              "usuario_nome,usuario_email,criado_em")


def _norm_entrada(s: str) -> str:
    """Normaliza o texto do cliente para a chave da memória de matching.

    ATENÇÃO: esta regra tem que ser IDÊNTICA à função norm_entrada() do Postgres.
    Se mudar aqui, mude lá — senão a memória grava com uma chave e busca com outra,
    e o aprendizado some sem dar erro.
    """
    import unicodedata as _ud
    t = _ud.normalize("NFD", (s or ""))
    t = "".join(c for c in t if _ud.category(c) != "Mn")
    t = re.sub(r"[^a-z0-9]+", " ", t.lower())
    return re.sub(r"\s+", " ", t).strip()


# ═══════════════════════════════════════════════════════════════════════════
# IDENTIDADE DE CONTEÚDO — o anexo é a cotação ou só repete o corpo?
#
# Nasceu do e-mail da Universal (22/07/2026, NEG-0040613): o corpo trazia 45
# linhas de item e o CNPJ; junto vinha um PDF "Ferramentas Padrão Eletricista",
# que é ANEXO DE REFERÊNCIA — 18 dos 19 códigos dele já estavam no corpo.
# O /extrair promoveu o PDF a "CONTEÚDO PARA COTAÇÃO" e rebaixou o corpo a
# "contexto" cortado em 3.000 chars (36 das 46 linhas e o CNPJ, que estava a 88%
# do corpo, foram jogados fora). A proposta chegou vazia na tela.
#
# A decisão de descartar é DETERMINÍSTICA e acontece ANTES de qualquer chamada
# de IA — comparar dois textos não precisa de modelo.
# ═══════════════════════════════════════════════════════════════════════════

# Código de item no ERP do cliente: UC.107572, AF.101766, ES.102241, MAT-4471.
_COD_CLIENTE_RE = re.compile(r'\b[A-Z]{2,4}[.\-]\d{4,8}\b')


def _tokens_identidade(txt: str) -> set:
    """Tokens que identificam os ITENS de um texto, para comparar duas fontes.

    Prefere o CÓDIGO DO CLIENTE: é um identificador que o cliente deu, não um que
    a gente inferiu — a chave mais confiável que existe aqui. Mas ele é raro: em
    36 e-mails reais medidos, só 8 traziam código. Quando não há código suficiente,
    cai para linhas de descrição normalizadas, que é o que sobra de determinístico.

    As duas fontes precisam usar o MESMO tipo de token para a comparação valer —
    quem garante isso é o _anexo_redundante.
    """
    if not txt:
        return set()
    cods = {c.upper() for c in _COD_CLIENTE_RE.findall(txt)}
    if len(cods) >= 3:
        return cods
    linhas = set()
    for l in re.split(r'[\r\n]+', txt or ""):
        l = re.sub(r'\s+', ' ', l).strip()
        # 25 chars + ao menos uma letra: descarta número solto, unidade e célula vazia
        if len(l) >= 25 and re.search(r'[A-Za-zÀ-ÿ]', l):
            linhas.add(_norm_entrada(l))
    return linhas


def _anexo_redundante(txt_anexo: str, txt_corpo: str):
    """O anexo só repete o que o corpo já traz? -> (bool, lista de exclusivos).

    Conservador de propósito. Exige evidência dos DOIS lados (>= 3 tokens cada) e
    sobreposição alta (>= 80%). Na dúvida NÃO descarta: processar o mesmo item duas
    vezes é um incômodo que o operador vê e corrige; descartar o anexo que ERA a
    cotação é uma proposta vazia que ele não tem como diagnosticar.

    Os `exclusivos` voltam para a tela. Anexo descartado em silêncio esconderia do
    operador que alguma coisa ficou de fora.
    """
    ta = _tokens_identidade(txt_anexo)
    tc = _tokens_identidade(txt_corpo)
    if len(ta) < 3 or len(tc) < 3:
        return False, []
    comuns = ta & tc
    if len(comuns) / len(ta) < 0.8:
        return False, []
    return True, sorted(ta - comuns)


def _recorte_contexto(txt: str, teto: int = 12000) -> str:
    """Recorta preservando as DUAS pontas do e-mail.

    O CNPJ e a assinatura moram no RODAPÉ. Cortar só a cauda (`txt[:teto]`) é o jeito
    mais eficiente de perder exatamente o campo que o SYSTEM_EXTRACAO chama de "o mais
    negligenciado" — no e-mail da Universal o CNPJ estava no caractere 11.872 de 13.715.
    """
    txt = txt or ""
    if len(txt) <= teto:
        return txt
    cab = int(teto * 0.6)
    rod = teto - cab
    return txt[:cab] + "\n\n[... trecho do meio omitido ...]\n\n" + txt[-rod:]


_TITULOS_FALHA = {
    "candidatos_trgm_lote": "Busca no banco de preços falhou ao gerar proposta",
    "memoria_match":        "Memória de matching indisponível",
    "produtos_da_memoria":  "Memória reconheceu o item mas o produto não foi lido",
    "ia_matching":          "IA de matching não respondeu",
    "geral":                "Matching falhou por completo ao gerar proposta",
}


def _abrir_chamado_automatico(sb, usuario: str, assinatura: str, titulo: str,
                              solicitacao: str, dor: str, esperado: str,
                              detalhe: str, area: str = "proposta"):
    """O sistema abre chamado contra si mesmo quando FALHA.

    Falhar != não achar. Produto que não existe no banco é resultado legítimo e
    não gera chamado nenhum. Isto aqui é só para quando a busca não pôde acontecer.

    Dedup por assinatura: enquanto houver chamado aberto/em desenvolvimento com a
    mesma assinatura, novas falhas só incrementam `ocorrencias`. Sem isso, uma
    indisponibilidade de 10 min no Render viraria 30 chamados idênticos no kanban.
    Depois de resolvido, uma falha nova abre chamado novo — o problema voltou.

    NUNCA levanta exceção: se o Supabase está fora, gravar o chamado também falha,
    e isso não pode derrubar a proposta do operador junto.
    """
    if not assinatura:
        return None
    try:
        r = sb.table("chamados").select("id,numero,ocorrencias")\
            .eq("assinatura", assinatura).in_("status", ["aberto", "em_desenvolvimento"])\
            .limit(1).execute()
        if r.data:
            sb.table("chamados").update({
                "ocorrencias": int(r.data[0].get("ocorrencias") or 1) + 1,
                "ultima_ocorrencia": _now_iso(),
                "atualizado_em": _now_iso(),
            }).eq("id", r.data[0]["id"]).execute()
            return r.data[0].get("numero")

        novo = sb.table("chamados").insert({
            "operador_email": usuario,
            "operador_nome": APELIDOS.get(usuario, ""),
            "origem": "sistema",
            "assinatura": assinatura,
            "tipo": "bug", "area": area, "prioridade": "alta", "status": "aberto",
            "titulo": titulo,
            "solicitacao": solicitacao,
            "dor": dor,
            "comportamento_esperado": esperado,
            "parecer_analista": detalhe,
            "descricao_operador": "(aberto automaticamente pelo sistema ao falhar)",
            "ocorrencias": 1, "ultima_ocorrencia": _now_iso(),
        }).execute()
        return ((novo.data or [{}])[0] or {}).get("numero")
    except Exception as e:
        # Corrida: dois operadores falharam junto e o índice único barrou o 2º.
        if "23505" in str(e) or "duplicate key" in str(e).lower():
            try:
                r2 = sb.table("chamados").select("numero")\
                    .eq("assinatura", assinatura).in_("status", ["aberto", "em_desenvolvimento"])\
                    .limit(1).execute()
                if r2.data:
                    return r2.data[0].get("numero")
            except Exception:
                pass
        return None


def _itens_sem_match(itens_raw: list) -> list:
    """Itens no formato COMPLETO que o frontend espera: descrição do cliente
    preservada, tudo o mais em branco, confiança 'nenhuma'.

    Usado quando o matching não pôde rodar. Antes, o /extrair devolvia
    `itens_brutos` cru nesse caso — sem `descricao_final` — e a tela mostrava
    item VAZIO em vez de item sem preço. Foi o sintoma do chamado #4 (um
    NameError engolido pelo except). A causa foi corrigida; esta função desarma
    a armadilha para qualquer exceção futura.

    Item sem match aparece com o texto do cliente e preço em branco. Nunca some.
    """
    out = []
    for item in (itens_raw or []):
        # O modelo às vezes devolve um item como STRING solta em vez de objeto —
        # acontece quando a "descrição" é um parágrafo de specs sem tabela (foi o que
        # derrubou a cotação de suporte de TV: AttributeError 'str' has no 'get').
        # Formato imperfeito do modelo não pode derrubar a extração: normaliza e segue.
        if isinstance(item, str):
            item = {"descricao": item, "descricao_original": item}
        elif not isinstance(item, dict):
            continue
        desc = item.get("descricao", "") or ""
        out.append({
            "descricao_original":   item.get("descricao_original", desc) or desc,
            "descricao_final":      desc,
            "codigo_cliente":       (item.get("codigo_cliente") or "").strip(),
            "specs_complementares": item.get("specs_complementares") or "",
            "quantidade":           item.get("quantidade", 1),
            "unidade":              item.get("unidade", "UN"),
            "preco_un":             0.0,
            "preco_custo":          0.0,
            "link_fornecedor":      "",
            "fornecedor":           "",
            "sku_fornecedor":       "",
            "obs":                  "SEM PREÇO",
            "confianca_match":      "nenhuma",
            "banco_candidato":      None,
            "banco_candidato_preco": None,
            "motivo_incerto":       None,
            "tem_preco":            False,
            "pede_atencao":         _pede_atencao(desc, item.get("specs_complementares")),
            "alerta_produto":       None,
            "banco":                None,
        })
    return out


def _rastreavel(row: dict) -> bool:
    """O produto tem lastro? Origem (link OU nome do fornecedor) E custo E venda.

    Regra do Fábio: preço que não dá pra rastrear custa o tempo dele e não vira nada.
    O gargalo medido é o CUSTO (5,5% do banco tem) — não a origem (11%).
    Origem aceita texto livre: "DIGITALSAT", "IngramMicro e-mail" são rastreáveis
    tanto quanto um link. Saber a quem perguntar já é lastro.
    """
    if not row:
        return False
    # Rastrear = saber QUEM vendeu e POR ONDE falar com ele. Um link resolve os
    # dois de uma vez. Um nome sozinho ("DIGITALSAT") diz quem, mas não como —
    # e o Fábio pediu rastreio de um clique, não um nome pra lembrar depois.
    # Aceita nome + contato como equivalente ao link: WhatsApp é rastro tanto
    # quanto URL, desde que o contato esteja lá.
    _link = (row.get("link_fornecedor") or "").strip()
    _nome = (row.get("fornecedor") or "").strip()
    _cont = (row.get("fornecedor_contato") or "").strip()
    tem_origem = bool(_link or (_nome and _cont))
    try:
        tem_custo = float(row.get("preco_custo") or 0) > 0
    except (TypeError, ValueError):
        tem_custo = False
    try:
        tem_venda = float(row.get("preco_un") or 0) > 0
    except (TypeError, ValueError):
        tem_venda = False
    return tem_origem and tem_custo and tem_venda


def _falta_lastro(row: dict) -> str:
    """O que exatamente falta pro produto ser rastreável — pro operador saber o que preencher."""
    if not row:
        return "produto sem registro"
    faltas = []
    _link = (row.get("link_fornecedor") or "").strip()
    _nome = (row.get("fornecedor") or "").strip()
    _cont = (row.get("fornecedor_contato") or "").strip()
    if not _link and not (_nome and _cont):
        # Diz o que falta de verdade — "origem" genérico não ajuda a preencher.
        faltas.append("contato do " + _nome if _nome else "origem")
    try:
        if not float(row.get("preco_custo") or 0) > 0:
            faltas.append("custo")
    except (TypeError, ValueError):
        faltas.append("custo")
    try:
        if not float(row.get("preco_un") or 0) > 0:
            faltas.append("preço de venda")
    except (TypeError, ValueError):
        faltas.append("preço de venda")
    return " e ".join(faltas) if faltas else ""


def _fazer_matching(itens_raw: list, claude, sb, cliente: str = "",
                    avisos: list = None, so_rastreavel: bool = False) -> list:
    """Matching em 3 camadas (v3.14):

      [1] MEMÓRIA  — entrada já virou proposta antes? Cliente primeiro, depois
                     global. Sem ambiguidade => confiança alta, ZERO IA.
      [2] TRIGRAMA — candidatos do banco INTEIRO via índice GiST (recall@40 de
                     95,1% medido contra 636 pares reais). Substituiu a janela
                     de 500 linhas mais recentes, que enxergava 0,8% do banco.
      [3] HAIKU    — escolhe entre os 40 candidatos (temperature=0).

    O aprendizado ([4]) acontece no /upsert-precos, quando a proposta vira real.
    """
    import json as _jm

    avisos = avisos if avisos is not None else []

    def _falhou(etapa, e, msg):
        """Registra que a busca NÃO PÔDE acontecer — distinto de 'não achei'.

        A `assinatura` é a chave de dedup do chamado automático: mesma etapa +
        mesmo tipo de erro = mesmo chamado, com contador de ocorrências.
        """
        avisos.append({
            "tipo": "busca_falhou", "etapa": etapa,
            "assinatura": f"matching:{etapa}:{type(e).__name__}",
            "mensagem": msg,
            "detalhe": f"{type(e).__name__}: {e}"[:400],
        })

    if not itens_raw:
        return []

    descricoes = [(item.get("descricao") or "") for item in itens_raw]

    # ── [1] MEMÓRIA ──────────────────────────────────────────────────────
    # Cada falha vira AVISO. Sem isso, "quebrei" e "não achei" chegam iguais na
    # tela e o operador precifica na mão achando que o banco está pobre.
    mem = {}
    try:
        rm = sb.rpc("memoria_match", {"entradas": descricoes, "cli": cliente or ""}).execute()
        for m in (rm.data or []):
            mem[m["entrada_norm"]] = m
    except Exception as e:
        mem = {}
        _falhou("memoria_match", e,
                "A memória de matching não respondeu: itens já validados antes "
                "não foram reconhecidos e voltaram para a IA.")

    prod_mem = {}
    ids_mem = sorted({m["produto_id"] for m in mem.values()})
    if ids_mem:
        try:
            rp = sb.table("produtos").select(_PROD_COLS).in_("id", ids_mem).execute()
            prod_mem = {p["id"]: p for p in (rp.data or [])}
        except Exception as e:
            prod_mem = {}
            _falhou("produtos_da_memoria", e,
                    "A memória reconheceu os itens mas o banco não devolveu os produtos.")

    resolvido = {}   # idx -> (linha_produto, origem)
    for i, d in enumerate(descricoes):
        m = mem.get(_norm_entrada(d))
        if m and m.get("produto_id") in prod_mem:
            resolvido[i] = (prod_mem[m["produto_id"]], m.get("origem", "global"))

    # ── [2] TRIGRAMA (só para o que a memória não resolveu) ───────────────
    pendentes = [i for i in range(len(itens_raw)) if i not in resolvido]
    candidatos_por_item = {i: [] for i in range(len(itens_raw))}
    if pendentes:
        try:
            rc = sb.rpc("candidatos_trgm_lote",
                        {"termos": [descricoes[i] for i in pendentes], "n": 40}).execute()
            for row in (rc.data or []):
                pos = pendentes[int(row["idx"]) - 1]      # idx do WITH ORDINALITY é 1-based
                candidatos_por_item[pos].append(row)
        except Exception as e:
            # A pior falha silenciosa: sem candidatos, NADA vai pro Haiku e a
            # proposta inteira sai "sem match" como se o banco estivesse vazio.
            _falhou("candidatos_trgm_lote", e,
                    "Não consegui consultar o banco de preços. Os itens estão sem "
                    "preço por FALHA DO SISTEMA — não porque o banco não tenha o "
                    "produto. Confira antes de precificar na mão.")

    # Lookup descrição -> linha completa (candidatos + resolvidos pela memória)
    _banco_por_desc = {}
    for lista in list(candidatos_por_item.values()) + [[r for r, _ in resolvido.values()]]:
        for c in lista:
            k = (c.get("descricao") or "").strip().lower()
            if k and k not in _banco_por_desc:
                _banco_por_desc[k] = c
    todos_candidatos = list(_banco_por_desc.values())

    # ── [3] HAIKU ────────────────────────────────────────────────────────
    matches = {}
    itens_ia = [i for i in pendentes if candidatos_por_item[i]]
    if itens_ia:
        # As specs_complementares vão JUNTO. Elas estavam sendo descartadas — e é
        # nelas que a divergência mora (69% dos itens têm specs preenchidas). O
        # matcher casava "suporte 60cm" com "suporte 60cm" e dava alta, sem ver que
        # o cliente pedia haste 600x50x50 e o nosso é base parede/teto. Vendeu,
        # comprou, entregou, voltou em RMA.
        itens_txt = ""
        for i in itens_ia:
            itens_txt += f"\nItem {i}\n  DESCRIÇÃO: {descricoes[i]}"
            _sp = (itens_raw[i].get("specs_complementares") or "").strip()
            itens_txt += f"\n  SPECS: {_sp}\n" if _sp else "\n  SPECS: (o cliente não informou)\n"

        candidatos_txt = ""
        for i in itens_ia:
            candidatos_txt += f"\n\n--- Candidatos para Item {i} ({descricoes[i][:60]}) ---\n"
            for j, c in enumerate(candidatos_por_item[i][:40]):
                candidatos_txt += f"  [{j}] {c.get('descricao','')} | R$ {c.get('preco_un',0)} | ref {c.get('proposta_tiny','')}\n"

        prompt_matching = f"""Itens solicitados:{itens_txt}

Candidatos do banco de preços:{candidatos_txt}

Para cada item, decida se algum candidato é O MESMO ITEM — comparando a DESCRIÇÃO **e** as SPECS.
Preencha veredito, motivo, diferencas e falta. Lembre: fabricante diferente = null; categoria
diferente = null; spec divergente = não é o mesmo item, mesmo que a descrição bata."""

        try:
            resp_match = claude.messages.create(
                model="claude-haiku-4-5-20251001", max_tokens=6000,
                system=SYSTEM_MATCHING + _excludentes_matching(sb),
                messages=[{"role": "user", "content": prompt_matching}],
                temperature=0.0, timeout=45.0
            )
            raw_match = resp_match.content[0].text.strip()
            raw_match = re.sub(r'^```(?:json)?\s*', '', raw_match)
            raw_match = re.sub(r'\s*```$', '', raw_match.strip())
            matches = {m["indice"]: m for m in _jm.loads(raw_match).get("matches", [])}
        except Exception as e:
            matches = {}
            _falhou("ia_matching", e,
                    "A IA de matching não respondeu. Os candidatos do banco foram "
                    "encontrados, mas ninguém escolheu entre eles.")
            avisos.append(f"IA de matching falhou ({type(e).__name__}): os itens "
                          f"vieram sem sugestão do banco. Os candidatos existiam.")

    # A memória vence a IA: entra depois e sobrescreve.
    for i, (row, origem) in resolvido.items():
        matches[i] = {
            "indice": i,
            "confianca": "alta",
            "banco_descricao": row.get("descricao") or "",
            "banco_preco": row.get("preco_un") or 0,
            "banco_proposta": row.get("proposta_tiny") or "",
            "motivo": f"memória · {origem}",
            "veredito": "mesmo",     # o operador já validou este par antes
            "diferencas": [],
            "falta": "",
            "_memoria": origem,
        }

    itens_com_preco = []
    for i, item in enumerate(itens_raw):
        desc = item.get("descricao", "")
        desc_original = item.get("descricao_original", desc) or desc
        specs_comp = item.get("specs_complementares") or ""
        match = matches.get(i, {})
        confianca = match.get("confianca", "nenhuma")
        banco_desc = (match.get("banco_descricao") or "").strip()
        tem_match = bool(banco_desc) and confianca in ("alta", "media", "baixa")

        # ── REGRA (v3.17): a descrição do cliente NUNCA é substituída ────────
        # Antes, em alta/media o banco sobrescrevia o texto do cliente e o original
        # sumia da tela; em baixa só aparecia como texto solto no campo obs; e no
        # modo "preservar" o banco era jogado fora. Nos três casos o operador ficava
        # sem os dois lados na mesma tela — e é a comparação que pega o item errado
        # ANTES de vender (o caso do suporte de refletor que voltou em RMA).
        # Agora: cliente sempre no item, banco sempre na ficha, operador compara.
        desc_final = (desc_original or "").strip() or desc
        _row = _banco_por_desc.get(banco_desc.lower()) if banco_desc else None

        # ── FILTRO DE RASTREABILIDADE (v3.19) ────────────────────────────────
        # Opção (B): o match CONTINUA aparecendo na ficha — o operador vê o que
        # existe, de onde veio, quando e por quem — mas NENHUM valor é importado.
        # Ele recota e, ao preencher, o produto ganha o lastro que faltava.
        # Esconder o match (opção A) seria pior: o operador não validaria nada e a
        # memória registraria erro contra um par que ele nunca viu.
        # ── REGRA (Leonardo, jul/2026): valor SÓ entra automático em MATCH IDÊNTICO
        # COM lastro. Precisão acima de cobertura — menos item entra sozinho, mais o
        # operador confirma.
        #   • "Idêntico" = mesma chave normalizada (_norm_entrada ignora acento,
        #     pontuação e caixa) OU veio da MEMÓRIA (desfecho que o operador já mandou
        #     pro Tiny — identidade confirmada por ele). Haiku dizendo "alta" em TEXTO
        #     DIFERENTE é semântico: NÃO é idêntico, então não auto-preenche.
        #   • "Trazer do banco = trazer origem": nada entra sem lastro. Venda, custo e
        #     origem entram JUNTOS, e só quando idêntico E a linha tem lastro
        #     (origem + custo + venda, via _rastreavel).
        #   • Qualquer outra coisa (semântico, media, baixa, ou idêntico-sem-lastro):
        #     o candidato aparece na FICHA como referência, mas nenhum valor entra.
        _veio_memoria = bool(match.get("_memoria"))
        _desc_igual = bool(_row) and _norm_entrada(desc_original) == _norm_entrada((_row or {}).get("descricao") or "")
        _identico = (confianca == "alta") and (_veio_memoria or _desc_igual)
        _carrega = _identico and _rastreavel(_row)
        # idêntico mas o banco não tem lastro => mostra e pede recotar (não preenche)
        _sem_lastro = _identico and not _rastreavel(_row)

        preco_un = 0.0
        preco_custo, link_fornecedor, fornecedor, sku_fornecedor = 0.0, "", "", ""
        fornecedor_canal, fornecedor_contato = "", ""
        if _carrega:
            preco_un = float(match.get("banco_preco") or 0)
            try:
                preco_custo = float(_row.get("preco_custo") or 0)
            except (TypeError, ValueError):
                preco_custo = 0.0
            link_fornecedor = _row.get("link_fornecedor") or ""
            fornecedor = _row.get("fornecedor") or ""
            fornecedor_canal = _row.get("fornecedor_canal") or ""
            fornecedor_contato = _row.get("fornecedor_contato") or ""
            sku_fornecedor = _row.get("sku_fornecedor") or ""

        # IDENTIDADE (A) pro save: o item casou com ESTA linha do banco (idêntico) =>
        # ao salvar, atualiza ELA, sem criar gêmeo cru. Só quando idêntico (a linha
        # certa); match semântico/incerto não amarra em linha nenhuma.
        _banco_id = (_row or {}).get("id") if _identico else None

        if not tem_match:
            obs_item = "SEM PREÇO"
        elif _sem_lastro:
            obs_item = "⚠ SEM LASTRO — recotar"
        elif _carrega:
            _ref = match.get("banco_proposta", "")
            obs_item = f"✓ ref {_ref}" if _ref else "✓"
        else:
            obs_item = "⚠ CONFIRA"   # candidato (semântico/incerto): referência, não preencheu

        # ── FICHA DE PROCEDÊNCIA ─────────────────────────────────────────────
        # Tudo que o operador precisa pra bater o que o cliente pediu contra o que
        # o banco propõe, sem sair da tela.
        ficha = None
        if tem_match:
            ficha = {
                "descricao":      banco_desc,
                # Id do candidato do banco. Serve ao selo de datasheet (que e'
                # dado do produto) sem precisar recriar a RPC candidatos_trgm_lote,
                # que devolve TABLE fixo e e' o coracao do matching.
                "produto_id":     (_row or {}).get("id"),
                "preco_un":       float(match.get("banco_preco") or 0),
                "preco_custo":    float((_row or {}).get("preco_custo") or 0),
                "fornecedor":     (_row or {}).get("fornecedor") or "",
                "link_fornecedor": (_row or {}).get("link_fornecedor") or "",
                # Quem, por onde, e o contato em si — as três coisas que o campo
                # antigo tentava guardar sozinho ("volt - wpp", "WPP DATALINK 115848").
                "fornecedor_canal":   (_row or {}).get("fornecedor_canal") or "",
                "fornecedor_contato": (_row or {}).get("fornecedor_contato") or "",
                "sku_fornecedor": (_row or {}).get("sku_fornecedor") or "",
                "cliente":        (_row or {}).get("cliente") or "",
                "cnpj":           (_row or {}).get("cnpj") or "",
                "data_ref":       str((_row or {}).get("data_ref") or ""),
                "proposta_tiny":  match.get("banco_proposta") or "",
                "confianca":      confianca,
                "defesa":         _defesa_do_match(match, _row),
                "herdou_custo":   _carrega,
                "identico":       _identico,
                # Sem lastro: a ficha mostra tudo, mas nenhum valor entrou no item.
                "sem_lastro":     _sem_lastro,
                "falta_lastro":   _falta_lastro(_row) if _sem_lastro else "",
                # É o mesmo item, e por quê — o que o operador hoje faz na mão,
                # colando as duas descrições numa aba de chat e perguntando.
                "veredito":       match.get("veredito") or "",
                "diferencas":     [d for d in (match.get("diferencas") or []) if str(d).strip()],
                "falta":          (match.get("falta") or "") or "",
                # Procedência do cadastro: quando, por quem, pra quem.
                "criado_em":      str((_row or {}).get("criado_em") or ""),
                "usuario_nome":   (_row or {}).get("usuario_nome") or "",
                "usuario_email":  (_row or {}).get("usuario_email") or "",
            }

        itens_com_preco.append({
            "descricao_original": desc_original,
            "descricao_final": desc_final,
            # Código do item no ERP do cliente. Não participa do matching com o banco
            # (o banco é nosso, o código é dele) — é a chave de propagação de preço
            # entre propostas do MESMO pedido, quando o cliente quebra por destino.
            "codigo_cliente": (item.get("codigo_cliente") or "").strip(),
            "specs_complementares": specs_comp,
            "quantidade": item.get("quantidade", 1),
            "unidade": item.get("unidade", "UN"),
            "preco_un": preco_un,
            "preco_custo": preco_custo,
            "link_fornecedor": link_fornecedor,
            "fornecedor": fornecedor,
            "fornecedor_canal": fornecedor_canal,
            "fornecedor_contato": fornecedor_contato,
            "sku_fornecedor": sku_fornecedor,
            "obs": obs_item,
            "confianca_match": confianca,
            "identico": _identico,
            "banco_id": _banco_id,
            "banco": ficha,
            # Mantidos pra não quebrar tela/consumidor antigo enquanto o front migra
            "banco_candidato": banco_desc if confianca == "baixa" else None,
            "banco_candidato_preco": float(match.get("banco_preco") or 0) if confianca == "baixa" else None,
            "motivo_incerto": match.get("motivo", "") if confianca == "baixa" else None,
            "tem_preco": preco_un > 0,
            "pede_atencao": _pede_atencao(desc, item.get("specs_complementares")),
            "alerta_produto": _alerta_do_candidato(banco_desc, todos_candidatos),
        })

    _ds_marcar_itens(sb, itens_com_preco)
    _ds_marcar_aprovados(sb, itens_com_preco)
    return itens_com_preco


def _defesa_do_match(match: dict, row: dict) -> str:
    """Por que o sistema casou este item — em português, pro operador julgar.

    Serve a dois fins: o operador entende o raciocínio e decide mais rápido; e ele
    aprende como o sistema lê, o que melhora a forma dele lançar descrição.
    """
    origem = match.get("_memoria")
    if origem == "cliente":
        return ("Você já validou esta mesma descrição para este cliente em uma proposta "
                "anterior — e ela virou este produto.")
    if origem == "global":
        return ("Esta mesma descrição já foi validada antes (por outro cliente) e virou "
                "este produto.")
    motivo = (match.get("motivo") or "").strip()
    conf = match.get("confianca")
    base = {"alta":  "A IA considerou o mesmo produto.",
            "media": "A IA achou parecido, mas não idêntico.",
            "baixa": "A IA ficou em dúvida — confira antes de usar."}.get(conf, "")
    return (base + (" " + motivo if motivo else "")).strip()


def _parsear_excel_estruturado(data: bytes):
    """Parser determinístico para Excel com tabela de itens estruturada.
    Detecta colunas pelo header e extrai dados sem chamar a IA.
    Retorna lista de propostas no formato padrão, ou None se não detectar estrutura.
    """
    try:
        import openpyxl, io as _io2
        HQTD2  = re.compile(r'\bqtd|\bquant', re.I)
        HDESC2 = re.compile(r'descri|material|servi[çc]|equipamento', re.I)
        HMETA2 = re.compile(r'cnpj|empresa|cliente|faturamento|rfq|referência|nº\s*rc|pedido', re.I)
        CNPJ_PAT = re.compile(r'(\d{2}\.?\d{3}\.?\d{3}[\/]?\d{4}-?\d{2})')

        wb = openpyxl.load_workbook(_io2.BytesIO(data), read_only=True, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            metadados_raw, item_rows_raw, header_row, header_idx = [], [], None, None

            for i, row in enumerate(rows):
                vals = [str(c).strip() if c is not None else "" for c in row]
                nao_v = [v for v in vals if v and v != "None"]
                if not nao_v:
                    continue
                linha = " | ".join(nao_v)
                first = nao_v[0]
                if header_idx is None:
                    if len(nao_v) >= 4 and HQTD2.search(linha) and HDESC2.search(linha):
                        header_idx = i
                        header_row = vals
                        continue
                    if HMETA2.search(linha):
                        metadados_raw.append(linha)
                    continue
                try:
                    int(first)
                    item_rows_raw.append(vals)
                except ValueError:
                    pass

            if not header_row or not item_rows_raw:
                continue

            def _col(pattern, headers):
                for i, h in enumerate(headers):
                    if h and pattern.search(str(h)):
                        return i
                return None

            i_desc = _col(HDESC2, header_row)
            i_qtd  = _col(HQTD2, header_row)
            i_und  = _col(re.compile(r'\bund|\bun\b|\bunid', re.I), header_row)

            if i_desc is None or i_qtd is None:
                continue

            # Extrair metadados: CNPJ, cliente, RC
            cnpj, cliente, rc_neg = None, None, None
            for meta in metadados_raw:
                if not cnpj:
                    cm = CNPJ_PAT.search(meta)
                    if cm:
                        raw_cnpj = cm.group(1)
                        # Verificar que tem 14 dígitos
                        if len(re.sub(r'\D', '', raw_cnpj)) == 14:
                            cnpj = raw_cnpj
                if not cliente:
                    m2 = re.search(r'(?:empresa|cliente)[^\|]*\|\s*([^|]+)', meta, re.I)
                    if m2:
                        cliente = m2.group(1).strip()
                if not rc_neg:
                    m3 = re.search(r'(?:rfq|nº\s*rc|n[uú]mero\s*rc|pedido)[^\|]*\|\s*(\S+)', meta, re.I)
                    if m3:
                        rc_neg = m3.group(1).strip()

            # Extrair itens
            itens = []
            for vals in item_rows_raw:
                def _cell(idx):
                    if idx is None or idx >= len(vals): return ""
                    v = vals[idx]
                    return str(v).strip() if v is not None else ""
                desc = _cell(i_desc)
                if not desc or desc == "None":
                    continue
                try:
                    qtd = float(_cell(i_qtd).replace(",", "."))
                    if qtd <= 0: qtd = 1
                except Exception:
                    qtd = 1
                unid = (_cell(i_und) or "UN").upper()
                if not unid or unid == "NONE": unid = "UN"
                itens.append({
                    "descricao":          desc[:120],
                    "descricao_original": desc[:400],
                    "specs_complementares": None,
                    "quantidade":  int(qtd) if qtd == int(qtd) else qtd,
                    "unidade":     unid,
                    "sugerir_pn":  False,
                })

            if not itens:
                continue

            return [{"titulo": sname, "cliente": cliente or "",
                     "cnpj": cnpj, "rc_neg": rc_neg, "itens": itens}]
    except Exception:
        pass
    return None


@app.post("/extrair")
async def extrair_email(
    texto: str = Form(None),
    arquivos: list[UploadFile] = File(default=[]),   # múltiplos arquivos (email + Excels + PDFs)
    imagens: list[UploadFile] = File(default=[]),
    numero_proposta: str = Form(...),
    so_rastreavel: str = Form("0"),
    token_form: str = Form(None),
    request: Request = None,
    usuario: str = Depends(verificar_token)
):
    """Extrai itens do e-mail/prints/planilhas e faz matching com o banco.
    Aceita múltiplos arquivos simultaneamente; retorna uma ou mais propostas."""

    import json as _json_ext

    imgs_msg: list = []        # imagens embutidas nos .msg
    contexto_email = ""        # body/assunto do email (contexto de cliente/CNPJ)
    conteudo_files: list = []  # [(nome, texto, tipo)] por arquivo de conteúdo (Excel, PDF)
    todas_imgs_len = 0

    _IMG_SKIP_EXT = re.compile(r'logo|logotipo|assinatura|signature|bullet|icon', re.I)

    def _extrair_excel_bytes(data, fname):
        """Extrai Excel priorizando linhas de item.
        Detecta o header da tabela e extrai só metadados-chave + linhas de item.
        Fallback para extração bruta se header não for detectável.
        """
        try:
            import openpyxl, io as _io
            wb = openpyxl.load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
            partes = []
            HQTD = re.compile(r'\bqtd|\bquant', re.I)
            HDESC = re.compile(r'descri|material|servi[çc]|equipamento', re.I)
            HMETA = re.compile(r'cnpj|empresa|cliente|faturamento|rfq|referência|nº\s*rc|pedido', re.I)
            for sname in wb.sheetnames:
                ws = wb[sname]
                rows = list(ws.iter_rows(values_only=True))
                metadados, item_rows, header_idx = [], [], None
                for i, row in enumerate(rows):
                    vals = [str(c).strip() if c is not None else "" for c in row]
                    nao_v = [v for v in vals if v and v != "None"]
                    if not nao_v:
                        continue
                    linha = " | ".join(nao_v)
                    first = nao_v[0]
                    if header_idx is None:
                        if len(nao_v) >= 4 and HQTD.search(linha) and HDESC.search(linha):
                            header_idx = i
                            item_rows.append(f"COLUNAS: {linha}")
                            continue
                        if HMETA.search(linha):
                            metadados.append(linha)
                        continue
                    try:
                        int(first)
                        item_rows.append(linha)
                    except ValueError:
                        pass
                if metadados:
                    partes.append("[META]\n" + "\n".join(metadados[:12]))
                if item_rows:
                    partes.append(f"[ITENS - {sname}]\n" + "\n".join(item_rows))
            resultado = "\n\n".join(partes)
            if resultado.strip():
                return resultado
            # Fallback: extração bruta para planilhas simples sem header detectável
            partes2 = []
            wb2 = openpyxl.load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
            for sname in wb2.sheetnames:
                ws = wb2[sname]
                linhas = []
                for row in ws.iter_rows(values_only=True):
                    vals = [str(c).strip() if c is not None else "" for c in row]
                    if any(v and v != "None" for v in vals):
                        linhas.append(" | ".join(vals))
                if linhas:
                    partes2.append(f"[ABA: {sname}]\n" + "\n".join(linhas[:300]))
            return "\n\n".join(partes2)
        except Exception:
            return ""

    propostas_raw: list = []   # acumulador — declarado aqui para estar disponível
    #                              durante o loop de arquivos (parser det. insere direto)
    avisos_ingestao: list = [] # falhas ao abrir arquivo; migram para avisos_extracao

    # ── Processar cada arquivo enviado ──────────────────────────────────────
    # Os bytes crus ficam guardados: a ingestão canônica (v3.25) monta o
    # DOCUMENTO a partir deles. O laço abaixo segue existindo pelo parser
    # determinístico de Excel, que funciona e não se mexe.
    _brutos: list = []
    for arq in (arquivos or []):
        if not (arq and arq.filename):
            continue
        fname = arq.filename
        flo = fname.lower()
        dados = await arq.read()
        _brutos.append((fname, dados))

        # ── Com a ingestão ativa, este laço só existe para o parser
        # determinístico de Excel. Rodar o corpo antigo junto abria cada .msg
        # DUAS vezes e convertia cada PDF DUAS vezes — custo e latência
        # dobrados — e um arquivo corrompido estourava aqui, antes do try da
        # ingestão, derrubando a requisição inteira.
        if _INGESTAO_OK:
            try:
                _planilhas = []
                if flo.endswith(".msg"):
                    _tmp_x = f"/tmp/xls_{abs(hash(fname)) % 10**9}_{len(dados)}.msg"
                    _mx = None
                    # O temporário é criado ANTES do try de propósito não: se o
                    # openMsg falhava fora do try/finally, o arquivo ficava em
                    # /tmp para sempre. Num serviço de longa duração isso vaza
                    # disco a cada .msg corrompido.
                    try:
                        with open(_tmp_x, "wb") as _f:
                            _f.write(dados)
                        _mx = extract_msg.openMsg(_tmp_x)
                        for _att in (_mx.attachments or []):
                            _an = (getattr(_att, "longFilename", None)
                                   or getattr(_att, "shortFilename", None) or "").lower()
                            if _an.endswith((".xlsx", ".xls", ".xlsm")) and _att.data:
                                _planilhas.append(_att.data)
                    finally:
                        try:
                            if _mx is not None:
                                _mx.close()
                        except Exception:
                            pass
                        try:
                            os.remove(_tmp_x)
                        except Exception:
                            pass
                elif flo.endswith((".xlsx", ".xls", ".xlsm")):
                    _planilhas.append(dados)
                for _pl in _planilhas:
                    _det = _parsear_excel_estruturado(_pl)
                    if _det:
                        propostas_raw.extend(_det)
            except Exception:
                # Falha aqui NÃO é fatal: a planilha ainda vai ao modelo pelo
                # bloco de anexo do Documento. O determinístico é atalho, não
                # única via.
                pass
            continue

        if flo.endswith(".msg"):
            # Email: extrai corpo (contexto) + anexos do email
            # INVARIANTE: os bytes vão pra /tmp antes do extract_msg (o mount de
            # upload é read-only e dá I/O error em acesso aleatório). Caminho ÚNICO
            # por arquivo: com dois .msg na mesma extração, o caminho fixo era
            # truncado pelo segundo enquanto o handle OLE do primeiro seguia aberto.
            _tmp_msg = f"/tmp/ext_upload_{abs(hash(fname)) % 10**9}_{len(dados)}.msg"
            with open(_tmp_msg, "wb") as _f:
                _f.write(dados)
            _msg = extract_msg.openMsg(_tmp_msg)
            corpo = (_msg.body or "").strip()
            # O HTML é a única fonte que diz ONDE cada imagem estava no e-mail.
            # É por ele que a tabela de itens (corpo) vem antes da assinatura.
            try:
                _html = _msg.htmlBody
                if isinstance(_html, (bytes, bytearray)):
                    _html = _html.decode("utf-8", "ignore")
            except Exception:
                _html = ""
            _html = _html or ""
            _imgs_deste: list = []
            # O remetente identifica de quem é a demanda (o domínio é o sinal mais
            # confiável quando há várias empresas no material) e a assinatura dele é
            # onde o CNPJ costuma estar. O /casar-po já mandava isso; aqui era
            # descartado — e era uma das causas de proposta sem CNPJ.
            _de = (getattr(_msg, "sender", "") or "").strip()
            _para = (getattr(_msg, "to", "") or "").strip()
            contexto_email += (f"Assunto: {_msg.subject}\n"
                               f"De: {_de}\n"
                               + (f"Para: {_para}\n" if _para else "")
                               + f"\nCorpo:\n{corpo}\n\n")

            for att in _msg.attachments:
                afn = (att.longFilename or att.shortFilename or "").lower()
                if not att.data:
                    continue
                if afn.endswith((".xlsx", ".xls", ".xlsm")):
                    # Tentar parser determinístico primeiro (Excel estruturado com header)
                    _props_det = _parsear_excel_estruturado(att.data)
                    if _props_det:
                        # Enriquecer com contexto do email (CNPJ pode estar no body)
                        for _p in _props_det:
                            if not _p.get("cnpj"):
                                # Pegar o PRIMEIRO CNPJ do texto era errado: o corpo quase
                                # sempre traz dois (o do cliente e o nosso, da assinatura ou
                                # do encadeamento). Percorre todos e fica com o primeiro que
                                # seja válido E não seja da Kist.
                                for _cand in re.findall(r'(\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2})', contexto_email):
                                    _ok = _cnpj_do_cliente(_cand)
                                    if _ok:
                                        _p["cnpj"] = _ok
                                        break
                            if not _p.get("cliente") and contexto_email:
                                _mc = re.search(r'(?:empresa|cliente|razão social)[:\s]+([^\n\|]{3,60})', contexto_email, re.I)
                                if _mc:
                                    _p["cliente"] = _mc.group(1).strip()
                        propostas_raw.extend(_props_det)
                    else:
                        # Fallback: extração por texto (IA vai processar)
                        xl = _extrair_excel_bytes(att.data, afn)
                        if xl.strip():
                            conteudo_files.append((att.longFilename or att.shortFilename, xl, "excel"))
                elif afn.endswith(".pdf") and not _PDF_SKIP.search(afn) and len(att.data) <= _PDF_MAX_BYTES:
                    pt = _pdf_po_texto(att.data)
                    if pt.strip():
                        conteudo_files.append((att.longFilename or att.shortFilename, pt, "pdf"))
                elif afn.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif")):
                    # O piso de 5.000 bytes que existia aqui era o bug: uma tabela
                    # de 1 linha (388x38) comprime pra 1,4 KB e era descartada como
                    # se fosse enfeite. Tamanho em bytes não distingue conteúdo de
                    # decoração. Agora só cai fora o que não pode ser tabela.
                    if _IMG_SKIP_EXT.search(afn) or _img_descartavel(att.data):
                        continue
                    _cid = (getattr(att, "cid", "") or "").strip("<> ")
                    _imgs_deste.append((att.longFilename or att.shortFilename or afn,
                                        att.data, _cid))

            # ── Ordem = posição no HTML do e-mail ────────────────────────────
            # A tabela de itens fica no CORPO (topo); logo e banners ficam na
            # ASSINATURA (fim). Ordenar por posição faz a cotação sempre entrar
            # antes do enfeite no orçamento de tokens. Sem HTML → cai no critério
            # antigo (maior primeiro), que é o que resolvia o caso da Universal.
            def _pos_no_html(t):
                _n, _b, _c = t
                for _alvo in (_c, _n):
                    if not _alvo:
                        continue
                    _p = _html.find(f"cid:{_alvo}")
                    if _p < 0:
                        _p = _html.find(_alvo)
                    if _p >= 0:
                        return _p
                return 10 ** 9          # não referenciada no corpo → por último
            if _html:
                _imgs_deste.sort(key=_pos_no_html)
            else:
                _imgs_deste.sort(key=lambda t: len(t[1]), reverse=True)
            # O 4º campo é a POSIÇÃO da imagem dentro do PRÓPRIO e-mail. Com dois
            # e-mails na mesma extração, ordenar a lista achatada faria a assinatura
            # do primeiro (5 banners) empurrar a tabela do segundo pra fora do teto.
            # Ordenando por posição relativa, o corpo de TODOS vem antes de qualquer
            # rodapé. Foi exatamente o caso CEINSPEC + CESLOG.
            _origem_msg = (_msg.subject or fname or "").strip()[:120]
            for _i, (_n, _b, _c) in enumerate(_imgs_deste):
                imgs_msg.append((_n, _b, _origem_msg, _i))

        elif flo.endswith((".xlsx", ".xls", ".xlsm")):
            _props_det = _parsear_excel_estruturado(dados)
            if _props_det:
                for _p in _props_det:
                    if not _p.get("cnpj") and contexto_email:
                        # Mesmo bug do caminho do .msg: pegar o PRIMEIRO CNPJ do corpo
                        # pega o da Kist quando ele vem antes (assinatura, encadeamento).
                        for _cand in re.findall(r'(\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2})', contexto_email):
                            _ok = _cnpj_do_cliente(_cand)
                            if _ok:
                                _p["cnpj"] = _ok
                                break
                propostas_raw.extend(_props_det)
            else:
                xl = _extrair_excel_bytes(dados, fname)
                if xl.strip():
                    conteudo_files.append((fname, xl, "excel"))

        elif flo.endswith(".pdf") and not _PDF_SKIP.search(flo) and len(dados) <= _PDF_MAX_BYTES:
            pt = _pdf_po_texto(dados)
            if pt.strip():
                conteudo_files.append((fname, pt, "pdf"))

        elif flo.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif")):
            if not _IMG_SKIP_EXT.search(flo) and not _img_descartavel(dados):
                # Posição 0: imagem que o operador anexou de propósito é conteúdo,
                # nunca rodapé — disputa o teto em pé de igualdade com a tabela.
                imgs_msg.append((fname, dados, fname, 0))

        else:
            try:
                contexto_email += dados.decode("utf-8", "ignore")[:4000] + "\n\n"
            except Exception:
                pass

    if texto:
        contexto_email += texto

    # ── DOCUMENTO CANÔNICO (v3.25) ──────────────────────────────────────────
    # O norte: se o operador joga o e-mail num chatbot e pergunta "quais itens
    # devo cotar", ele responde — porque não há máquina entre o arquivo e o
    # modelo. Aqui é a mesma coisa: o e-mail vira uma lista ORDENADA de blocos
    # (texto, tabela, imagem, anexo) e vai inteiro, numa chamada só.
    #
    # O que morre com isto, tudo medido no corpus real de 264 cotações:
    #   · piso de 5 KB por imagem     → tabela de 1,4 KB era descartada
    #   · teto de 4 imagens por byte  → assinatura de 27 KB tomava a vaga
    #   · "anexo vence o corpo"       → PDF cadastral sequestrava o e-mail
    #   · corpo cortado em 3.000      → 36 de 46 itens perdidos (Universal)
    #   · 1 chamada por anexo         → e-mail com 6 PDFs virava 6 propostas
    documentos: list = []
    if _INGESTAO_OK:
        _conv = {"pdf": lambda b: _pdf_po_texto(b) if len(b) <= _PDF_MAX_BYTES else "",
                 "planilha": lambda b: _extrair_excel_bytes(b, "anexo"),
                 "word": lambda b: _ler_docx(b) if "_ler_docx" in globals() else ""}
        for _fn, _dd in _brutos:
            _fl = _fn.lower()
            try:
                if _fl.endswith(".msg"):
                    documentos.append(_ing_ler_msg(_dd, _conv))
                elif _fl.endswith(".eml"):
                    documentos.append(_ing_ler_email(_dd, _conv))
                else:
                    documentos.append(_ing_doc_arquivo(_fn, _dd, _conv))
            except Exception as _e:
                avisos_ingestao.append({
                    "tipo": "busca_falhou", "etapa": "ingestao",
                    "assinatura": f"ingestao:{type(_e).__name__}",
                    "mensagem": (f"Não consegui abrir “{_fn}”. Os itens não foram "
                                 f"perdidos no e-mail — foi falha do sistema."),
                    "detalhe": f"{type(_e).__name__}: {_e}"[:400],
                })
        for _img in (imagens or []):
            if _img and _img.filename:
                try:
                    documentos.append(_ing_doc_arquivo(
                        _img.filename, await _img.read(), _conv))
                except Exception:
                    pass

    # ── Avisos e notas (declarados aqui: o filtro abaixo já escreve neles) ───
    # avisos_extracao = FALHA do sistema (abre chamado automático).
    # notas_extracao  = decisão que o operador precisa saber, mas não é falha.
    # Misturar os dois faria "descartei um anexo repetido" abrir chamado como se
    # fosse quebra — e chamado que não é problema treina todo mundo a ignorar chamado.
    avisos_extracao: list = []
    notas_extracao: list = []

    # ── Anexo que só repete o corpo NÃO é a cotação ──────────────────────────
    # Sem isto, o anexo de referência vira "CONTEÚDO PARA COTAÇÃO" e o corpo — onde
    # estão os itens e o CNPJ — é rebaixado a contexto recortado. Foi exatamente o
    # que esvaziou a proposta da Universal (NEG-0040613, 22/07).
    # Só vale para PDF: Excel estruturado nem chega aqui (o parser determinístico
    # insere direto em propostas_raw), e quando um Excel chega junto ele É a cotação.
    descartados_files: list = []   # [(nome, texto)] — não foram à IA, mas ficam na fonte
    if conteudo_files and len(contexto_email.strip()) >= 400:
        _mantidos = []
        for _nome, _cont, _tipo in conteudo_files:
            if _tipo != "pdf":
                _mantidos.append((_nome, _cont, _tipo))
                continue
            _red, _excl = _anexo_redundante(_cont, contexto_email)
            if not _red:
                _mantidos.append((_nome, _cont, _tipo))
                continue
            descartados_files.append((_nome, _cont))
            _msg_nota = (f"O anexo “{_nome}” repete os itens que já estão no corpo do "
                         f"e-mail, então usei o corpo como fonte da cotação.")
            if _excl:
                _msg_nota += (" Estes códigos aparecem só no anexo e NÃO viraram item: "
                              + ", ".join(_excl[:12])
                              + (f" (e mais {len(_excl) - 12})" if len(_excl) > 12 else "")
                              + ". Confira se algum deles deveria estar na proposta.")
            notas_extracao.append({
                "tipo": "anexo_redundante",
                "arquivo": _nome,
                "mensagem": _msg_nota,
                "exclusivos": _excl[:40],
            })
        conteudo_files = _mantidos

    # ── Montar chamadas de extração ──────────────────────────────────────────
    # Regra: um arquivo de conteúdo (Excel/PDF) = uma proposta candidata
    # Sem arquivos de conteúdo = tudo junto em uma chamada (body + imagens)
    imgs_validas = [img for img in (imagens or []) if img and img.filename]
    todas_imgs_len = len(imgs_validas) + len(imgs_msg)
    if _INGESTAO_OK and documentos:
        # 91% das cotações reais trazem imagem embutida — quem manda no modelo é
        # o DOCUMENTO, não a contagem antiga de anexos soltos.
        todas_imgs_len = sum(len(d.imagens()) for d in documentos)
    modelo_extracao = "claude-sonnet-4-6" if todas_imgs_len > 0 else "claude-haiku-4-5-20251001"
    claude = get_claude()

    async def _chamar_extracao(payload_txt, imgs_inline=None, imgs_upload=None):
        """Monta o payload e chama o Claude para extração."""
        msg_content = []
        if payload_txt.strip():
            msg_content.append({"type": "text", "text": payload_txt[:20000]})
        if imgs_inline:
            # O teto ANTIGO era "4 imagens, maior primeiro". A ordem por tamanho
            # nasceu do e-mail da Universal (banners leves, prints pesados) e se
            # INVERTE no Grupo Cesari: assinatura de 27 KB, tabela de itens de
            # 1,4 KB — as 4 vagas iam todas pro rodapé e a cotação chegava vazia.
            # Agora a ordem vem de onde a imagem estava no e-mail (corpo antes da
            # assinatura, resolvido na leitura do .msg) e o teto é o que custa de
            # verdade: TOKEN. Imagem de e-mail é pequena; 12 delas cabem no mesmo
            # orçamento que 1 print de tela.
            # Sort ESTÁVEL pela posição dentro do e-mail de origem: corpo de todos
            # os e-mails primeiro, rodapé de todos depois. Empate mantém a ordem de
            # chegada (e-mail 1 antes do e-mail 2).
            imgs_inline = sorted(imgs_inline, key=lambda t: (t[3] if len(t) > 3 else 0))
            sel, gasto, cortadas = [], 0, 0
            for _t in imgs_inline:
                _nome, _b = _t[0], _t[1]
                _orig = _t[2] if len(_t) > 2 else ""
                _tk = _img_tokens(_b)
                if sel and (gasto + _tk > _IMG_MAX_TOKENS or len(sel) >= _IMG_MAX_N):
                    cortadas += 1
                    continue
                sel.append((_nome, _b, _orig))
                gasto += _tk
            msg_content.append({"type": "text", "text": (
                f"Analise também {len(sel)} imagem(ns) embutida(s) no(s) e-mail(s). "
                "Algumas são assinatura/logo/banner e não têm item nenhum — ignore "
                "essas em silêncio. As tabelas de itens são as que importam.")})
            for _nome, _b, _orig in sel:
                # O rótulo amarra a imagem ao e-mail de onde veio. Sem ele, com dois
                # e-mails de clientes diferentes na mesma extração, o modelo não tem
                # como saber qual tabela é de qual — e a quebra por destino erra.
                if _orig:
                    msg_content.append({"type": "text",
                                        "text": f"[imagem de: {_orig}]"})
                msg_content.append({"type": "image", "source": {"type": "base64",
                    "media_type": _media_type_img(_b), "data": _b64.standard_b64encode(_b).decode()}})
            if cortadas:
                # NOTA, não aviso: as que sobram são as do fim do e-mail (rodapé), e
                # abrir chamado automático por banner de assinatura descartado treina
                # todo mundo a ignorar chamado. O operador só precisa saber.
                notas_extracao.append({
                    "tipo": "imagens_cortadas",
                    "arquivo": "",
                    "mensagem": (f"O(s) e-mail(s) traziam {len(sel) + cortadas} imagens e li as "
                                 f"{len(sel)} primeiras (as do corpo vêm antes das de "
                                 f"assinatura). Se faltar item, confira as imagens à mão."),
                    "exclusivos": [],
                })
        if imgs_upload:
            _gasto_up = 0
            for img in (imgs_upload or [])[:_IMG_MAX_N]:
                ib = await img.read()
                if _img_descartavel(ib):
                    continue
                _tk = _img_tokens(ib)
                if _gasto_up and _gasto_up + _tk > _IMG_MAX_TOKENS:
                    break
                _gasto_up += _tk
                msg_content.append({"type": "image", "source": {"type": "base64", "media_type": _media_type_img(ib),
                    "data": _b64.standard_b64encode(ib).decode()}})
        if not msg_content:
            return []
        # 4000 tokens não cabiam uma cotação grande. O e-mail da Universal
        # (NEG-0040613) precisava de ~4.400 no piso e ~7.000 com os endereços de
        # entrega por item: a resposta cortava no meio, o JSON ficava inválido, o
        # except devolvia [] e a proposta chegava VAZIA na tela — sem nenhum sinal
        # de que tinha havido corte. Teto alto + o aviso abaixo: se bater, o
        # operador fica sabendo em vez de receber uma lista silenciosamente curta.
        resp = claude.messages.create(
            model=modelo_extracao, max_tokens=16000,
            system=SYSTEM_EXTRACAO,
            messages=[{"role": "user", "content": msg_content}],
        )
        if getattr(resp, "stop_reason", "") == "max_tokens":
            avisos_extracao.append({
                "tipo": "busca_falhou", "etapa": "extracao_truncada",
                "assinatura": "extracao:max_tokens",
                "mensagem": ("A lista de itens era grande demais e a extração foi cortada "
                             "no meio. A proposta pode ter vindo INCOMPLETA — confira "
                             "contra o e-mail antes de gerar o CSV."),
                "detalhe": f"stop_reason=max_tokens · modelo={modelo_extracao}",
            })
        raw = "".join(b.text for b in resp.content if getattr(b, "type", "") == "text").strip()
        # Extração robusta: remove backticks/preamble e extrai pelo primeiro { → último }
        raw = re.sub(r'^```(?:json)?\s*', '', raw.strip())
        raw = re.sub(r'\s*```$', '', raw.strip())
        # Se ainda tiver texto antes do JSON (ex: preamble "Aqui está:" ou "json" no início)
        start = raw.find('{')
        if start > 0:
            raw = raw[start:]
        end = raw.rfind('}')
        if end != -1 and end < len(raw) - 1:
            raw = raw[:end + 1]
        try:
            parsed = _json_ext.loads(raw)
            # Normalizar: se retornar formato antigo (sem propostas[]), encapsular
            if "itens" in parsed and "propostas" not in parsed:
                parsed = {"propostas": [parsed]}
            return parsed.get("propostas", [])
        except Exception as e:
            # Antes este except devolvia [] calado e a tela mostrava proposta vazia
            # como se o e-mail não tivesse itens. "Não veio nada" e "não consegui ler"
            # são coisas diferentes e o operador precisa saber qual das duas é.
            avisos_extracao.append({
                "tipo": "busca_falhou", "etapa": "extracao_json",
                "assinatura": f"extracao:json:{type(e).__name__}",
                "mensagem": ("Li o e-mail mas não consegui interpretar a resposta da "
                             "extração. Os itens NÃO foram perdidos no e-mail — foi "
                             "falha do sistema. Tente de novo; se repetir, me chame."),
                "detalhe": f"{type(e).__name__}: {e}"[:400],
            })
            return []

    async def _chamar_com_content(msg_content):
        """Mesma chamada, recebendo o content já montado pela ingestão."""
        if not msg_content:
            return []
        resp = claude.messages.create(
            model=modelo_extracao, max_tokens=16000,
            system=SYSTEM_EXTRACAO,
            messages=[{"role": "user", "content": msg_content}],
        )
        if getattr(resp, "stop_reason", "") == "max_tokens":
            avisos_extracao.append({
                "tipo": "busca_falhou", "etapa": "extracao_truncada",
                "assinatura": "extracao:max_tokens",
                "mensagem": ("A lista de itens era grande demais e a extração foi "
                             "cortada no meio. A proposta pode ter vindo INCOMPLETA "
                             "— confira contra o e-mail antes de gerar o CSV."),
                "detalhe": f"stop_reason=max_tokens · modelo={modelo_extracao}",
            })
        raw = "".join(b.text for b in resp.content
                      if getattr(b, "type", "") == "text").strip()
        raw = re.sub(r'^```(?:json)?\s*', '', raw.strip())
        raw = re.sub(r'\s*```$', '', raw.strip())
        s0 = raw.find('{')
        if s0 > 0:
            raw = raw[s0:]
        e0 = raw.rfind('}')
        if e0 != -1 and e0 < len(raw) - 1:
            raw = raw[:e0 + 1]
        try:
            parsed = _json_ext.loads(raw)
            if "itens" in parsed and "propostas" not in parsed:
                parsed = {"propostas": [parsed]}
            return parsed.get("propostas", [])
        except Exception as e:
            avisos_extracao.append({
                "tipo": "busca_falhou", "etapa": "extracao_json",
                "assinatura": f"extracao:json:{type(e).__name__}",
                "mensagem": ("Li o e-mail mas não consegui interpretar a resposta da "
                             "extração. Os itens NÃO foram perdidos no e-mail — foi "
                             "falha do sistema. Tente de novo; se repetir, me chame."),
                "detalhe": f"{type(e).__name__}: {e}"[:400],
            })
            return []

    avisos_extracao.extend(avisos_ingestao)
    _rel_ing = {}

    if _INGESTAO_OK and documentos and not propostas_raw:
        # ── UMA chamada, com o documento inteiro e em ordem ─────────────────
        # Uma chamada, não uma por anexo: um e-mail da Universal com 6 PDFs é UMA
        # demanda do cliente, não seis propostas. Quem quebra em abas é o MODELO,
        # pela regra de DESTINO — endereço/CNPJ diferente = proposta diferente —
        # e para decidir isso ele precisa ver o pedido inteiro de uma vez.
        _content, _rel_ing = _ing_montar_payload(documentos, texto_extra=(texto or ""))
        propostas_raw.extend(await _chamar_com_content(_content))
        if _rel_ing.get("imagens_cortadas"):
            notas_extracao.append({
                "tipo": "imagens_cortadas", "arquivo": "",
                "mensagem": (f"O material trazia mais imagens do que coube na leitura "
                             f"({_rel_ing['imagens_cortadas']} ficaram de fora). Se "
                             f"faltar item, confira as imagens à mão."),
                "exclusivos": [],
            })
    elif conteudo_files:
        # Caminho antigo — só roda se a ingestão não estiver disponível.
        for _ordem, (nome_arq, conteudo_arq, _tipo_arq) in enumerate(conteudo_files):
            ctx = f"CONTEXTO (cliente/CNPJ/referência do e-mail):\n{_recorte_contexto(contexto_email)}\n\n" \
                  f"CONTEÚDO PARA COTAÇÃO — arquivo: {nome_arq}\n{conteudo_arq[:15000]}"
            props = await _chamar_extracao(
                ctx,
                imgs_inline=(imgs_msg if _ordem == 0 else None),
                imgs_upload=(imgs_validas if _ordem == 0 else None),
            )
            for p in props:
                if isinstance(p, dict):
                    p.setdefault("titulo", nome_arq)
            propostas_raw.extend(props)
    else:
        # Sem arquivo de conteúdo: corpo + imagens numa chamada só. Quando há vários
        # e-mails, seus corpos já estão concatenados em contexto_email — e é o MODELO
        # que decide a quebra em propostas, pela regra de DESTINO do SYSTEM_EXTRACAO
        # (endereço/CNPJ diferente = proposta diferente). A quebra é por CONTEÚDO, não
        # por número de arquivos: 2 e-mails pro mesmo destino = 1 aba; 1 e-mail com 2
        # destinos = 2 abas.
        props = await _chamar_extracao(contexto_email, imgs_inline=imgs_msg, imgs_upload=imgs_validas)
        propostas_raw.extend(props)

    if not propostas_raw:
        propostas_raw = [{"titulo": "", "cliente": "", "cnpj": None,
                          "rc_neg": None, "itens": []}]

    # ── Matching com banco + atribuição de números de proposta ──────────────
    try:
        base_num = int(numero_proposta)
    except Exception:
        base_num = None

    t0 = time.time()
    resultado_propostas = []
    # ATENÇÃO: `avisos_extracao` NÃO é zerado aqui. Ele é declarado antes do filtro
    # de anexos e já pode conter avisos da própria extração (corte por max_tokens,
    # JSON ilegível). Zerar neste ponto apagava justamente os avisos que dizem por
    # que a proposta veio curta.

    # Tudo que a IA viu: corpo do e-mail, anexos convertidos e o texto colado.
    # O anexo descartado entra MARCADO: a fonte é o registro do que chegou, e sumir
    # com ele esconderia a decisão de quem for auditar a proposta depois.
    # A FONTE é o documento renderizado: exatamente o que a IA leu, na ordem em
    # que leu, com a imagem marcada no lugar dela. Antes era a concatenação solta
    # de corpo + anexos, que já não correspondia ao que foi enviado.
    _fonte_ing = ""
    if _INGESTAO_OK and documentos:
        try:
            _fonte_ing = "\n\n".join(d.render_texto() for d in documentos).strip()
        except Exception:
            _fonte_ing = ""
    _fonte = _fonte_ing or "\n\n".join(x for x in [
        contexto_email.strip(),
        "\n\n".join(f"--- {n} ---\n{c}" for n, c, _t in conteudo_files),
        "\n\n".join(f"--- {n} (anexo redundante — não enviado à IA) ---\n{c}"
                    for n, c in descartados_files),
        (texto or "").strip(),
    ] if x).strip()
    # ── Normalizar o que o modelo devolveu ──────────────────────────────────
    # O modelo é instruído a devolver propostas[] de objetos, cada uma com itens[]
    # de objetos. Quando falha nisso — proposta como string, itens como lista de
    # strings, itens ausente — o código antigo estourava (AttributeError/TypeError)
    # e derrubava a extração INTEIRA. Uma proposta malformada não pode custar as
    # outras quatro. Normaliza aqui, uma vez, antes de qualquer acesso.
    _props_norm = []
    for _pr in propostas_raw:
        if isinstance(_pr, str):
            _pr = {"itens": [{"descricao": _pr, "descricao_original": _pr}]}
        elif not isinstance(_pr, dict):
            continue
        _its = _pr.get("itens")
        if isinstance(_its, dict):            # um item solto fora de lista
            _its = [_its]
        elif not isinstance(_its, list):
            _its = []
        _its_ok = []
        for _it in _its:
            if isinstance(_it, str):
                _its_ok.append({"descricao": _it, "descricao_original": _it})
            elif isinstance(_it, dict):
                _its_ok.append(_it)
            # qualquer outra coisa (número, null) é descartada em silêncio
        _pr["itens"] = _its_ok
        # ── Linha idêntica repetida DENTRO da mesma proposta SOMA ───────────
        # Regra do operador, confirmada nos dados: na Syntegon RC 24202087 o
        # cliente repetiu as mesmas 5 linhas na MESMA tabela e a proposta 1050722
        # saiu 2/2/4/2/6 — somado, 5 itens, não 10. Hoje isso acontece por SORTE:
        # nada manda somar, e a regra escrita diz "NUNCA junte" (ela existe para
        # propostas DIFERENTES). Se o modelo aplicá-la aqui, o valor da proposta
        # dobra e ninguém percebe, porque 10 linhas parecem válidas.
        # A quebra por DESTINO vem antes e não é tocada: isto roda DENTRO de uma
        # proposta, ou seja, dentro de um destino.
        if _INGESTAO_OK and _ing_consolidar:
            try:
                _antes = len(_pr["itens"])
                _pr["itens"] = _ing_consolidar(_pr["itens"])
                if len(_pr["itens"]) < _antes:
                    notas_extracao.append({
                        "tipo": "itens_somados", "arquivo": "",
                        "mensagem": (f"O cliente repetiu linhas idênticas na mesma "
                                     f"tabela: juntei {_antes} em {len(_pr['itens'])} "
                                     f"itens, somando as quantidades. Confira se era "
                                     f"isso mesmo."),
                        "exclusivos": [],
                    })
            except Exception:
                pass
        _props_norm.append(_pr)
    propostas_raw = _props_norm

    if not propostas_raw:
        propostas_raw = [{"titulo": "", "cliente": "", "cnpj": None,
                          "rc_neg": None, "itens": []}]

    for idx_p, prop_raw in enumerate(propostas_raw):
        num_prop = str(base_num + idx_p) if base_num is not None else (
            numero_proposta if idx_p == 0 else f"{numero_proposta}-{idx_p + 1}")
        prop_raw["proposta"] = num_prop
        # A fonte que a IA leu viaja junto: é a spec ORIGINAL do cliente, e é contra
        # ela que o /conferir compara. Sem isso, comparamos contra o nosso resumo.
        prop_raw["fonte_texto"] = _fonte[:60000]
        # Conferir o CNPJ ANTES de ele viajar. Inválido ou nosso vira None em vez de
        # seguir até o CSV e até o banco de preços.
        prop_raw["cnpj"] = _cnpj_do_cliente(prop_raw.get("cnpj"))

        itens_brutos = prop_raw.get("itens", []) or []
        if not itens_brutos:
            resultado_propostas.append({**prop_raw, "itens": []})
            continue

        # Matching com o banco de preços
        sb = get_supabase()
        _so_rastreavel = str(so_rastreavel).strip().lower() in ("1", "true", "on", "yes", "sim")
        try:
            itens_enriquecidos = _fazer_matching(itens_brutos, claude, sb,
                                                 cliente=prop_raw.get("cliente") or "",
                                                 avisos=avisos_extracao,
                                                 so_rastreavel=_so_rastreavel)
        except Exception as e:
            # Fallback com a FORMA correta: descrição do cliente, tudo em branco.
            # (Devolver itens_brutos aqui produzia item vazio na tela.)
            itens_enriquecidos = _itens_sem_match(itens_brutos)
            avisos_extracao.append({
                "tipo": "busca_falhou", "etapa": "geral",
                "assinatura": f"matching:geral:{type(e).__name__}",
                "mensagem": "O matching falhou por completo. Os itens vieram com a "
                            "descrição do cliente e sem preço, por FALHA DO SISTEMA.",
                "detalhe": f"{type(e).__name__}: {e}"[:400],
            })

        resultado_propostas.append({**prop_raw, "itens": itens_enriquecidos})

    elapsed = time.time() - t0

    # ── Falha != ausência ────────────────────────────────────────────────
    # Produto que o banco não tem é resultado legítimo: sai sem match, sem alarme.
    # Busca que NÃO PÔDE acontecer abre chamado (deduplicado por assinatura) e
    # volta pro operador com o número — ligado a ele, que é quem viu o problema.
    avisos_saida, vistas = [], set()
    if avisos_extracao:
        _sb = get_supabase()
        for a in avisos_extracao:
            assin = a.get("assinatura") or ""
            if not assin or assin in vistas:
                continue
            vistas.add(assin)
            num = _abrir_chamado_automatico(
                _sb, usuario, assin,
                titulo=_TITULOS_FALHA.get(a.get("etapa"), "Falha no matching da proposta"),
                solicitacao=("Ao gerar uma proposta, o sistema não conseguiu consultar o "
                             "banco de preços. Os itens saíram com a descrição original do "
                             "cliente e sem preço."),
                dor=("O operador não distingue 'o banco não tem esse produto' de 'o sistema "
                     "quebrou'. Ele precifica na mão itens que o banco já tinha — e ninguém "
                     "descobre que houve falha."),
                esperado=("A busca deve funcionar. Enquanto não funciona, o operador precisa "
                          "ver na tela que é falha do sistema, não ausência no banco."),
                detalhe=f"Etapa: {a.get('etapa')} · {a.get('detalhe')}",
                area="proposta",
            )
            avisos_saida.append({
                "tipo": a.get("tipo"), "etapa": a.get("etapa"),
                "mensagem": a.get("mensagem"), "detalhe": a.get("detalhe"),
                "chamado": num,
            })

    # `notas` é chave NOVA e separada de `avisos` de propósito: o frontend antigo
    # ignora o que não conhece, então o backend pode subir sozinho sem a tela
    # renderizar uma nota informativa dentro do banner vermelho de falha.
    return {"propostas": resultado_propostas, "elapsed": round(elapsed, 2),
            "avisos": avisos_saida, "notas": notas_extracao}


def _ilike_literal(s: str) -> str:
    """Neutraliza os curingas do LIKE para casar a descrição LITERALMENTE.

    Sem isso, '%' e '_' viram curinga: 'LUVA VAQUETA 100% PETROLEIRA VT07' casa com
    qualquer coisa entre '100' e ' PETROLEIRA VT07', e o update pode cair no produto
    ERRADO. Há 43 descrições com % ou _ no banco.
    """
    return (s or "").replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


def _aprender_memoria(sb, pares: list, cliente: str) -> dict:
    """[4] Grava o desfecho da proposta na memória de matching.

    Premissa (definida pelo Leonardo): o OPERADOR é a hierarquia superior. O que
    ele mandou pro Tiny é verdade — o sistema aprende com ele, não o audita.

    - ACERTO: o par (entrada do cliente -> produto que a proposta virou) é FATO.
      Grava sempre, inclusive quando o matching não sugeriu nada: item que hoje
      sai com confianca='nenhuma' e o operador digita na mão vira treino.
    - ERRO:   só quando a própria memória diria OUTRO produto para a mesma entrada.
      Aí a sugestão dela foi contrariada pelo desfecho e ela é rebaixada.
      Na dúvida não julga nada — mas o fato fica registrado de qualquer forma.
    """
    cli = _norm_entrada(cliente or "")
    acertos, erros = 0, 0

    # O que a memória diria HOJE para estas entradas (antes de aprender)
    antes = {}
    try:
        r = sb.rpc("memoria_match", {"entradas": [p[0] for p in pares], "cli": cliente or ""}).execute()
        for m in (r.data or []):
            antes[m["entrada_norm"]] = m.get("produto_id")
    except Exception:
        antes = {}

    for entrada, pid in pares:
        # ACERTO — o desfecho
        try:
            ex = sb.table("match_memoria").select("id,acertos")\
                .eq("entrada_norm", entrada).eq("cliente_norm", cli).eq("produto_id", pid)\
                .limit(1).execute()
            if ex.data:
                sb.table("match_memoria").update({
                    "acertos": int(ex.data[0].get("acertos") or 0) + 1,
                    "ultima_vez": _now_iso(),
                }).eq("id", ex.data[0]["id"]).execute()
            else:
                sb.table("match_memoria").insert({
                    "entrada_norm": entrada, "cliente_norm": cli,
                    "produto_id": pid, "acertos": 1,
                }).execute()
            acertos += 1
        except Exception:
            continue

        # ERRO — a memória apontava para outro produto
        sugerido = antes.get(entrada)
        if sugerido and sugerido != pid:
            try:
                ex2 = sb.table("match_memoria").select("id,erros")\
                    .eq("entrada_norm", entrada).eq("produto_id", sugerido).limit(1).execute()
                if ex2.data:
                    sb.table("match_memoria").update({
                        "erros": int(ex2.data[0].get("erros") or 0) + 1,
                        "ultima_vez": _now_iso(),
                    }).eq("id", ex2.data[0]["id"]).execute()
                    erros += 1
            except Exception:
                pass

    return {"aprendidos": acertos, "rebaixados": erros}


@app.post("/upsert-precos")
async def upsert_precos(payload: dict, usuario: str = Depends(verificar_token)):
    sb = get_supabase()
    proposta = payload.get("proposta", "")
    cliente  = payload.get("cliente", "")
    # O CNPJ já vinha no payload (o front manda a proposta inteira) e ninguém lia.
    # É ele que diz PRA QUEM o preço foi oferecido — matriz e filial têm o mesmo
    # nome. Grava a cada CSV, como o custo: a coluna enche sozinha com o uso.
    cnpj     = _cnpj_do_cliente(payload.get("cnpj"))
    # Quem gravou este preço. A ficha mostra "por quem foi inserido" — sem isto,
    # o operador vê um preço órfão e não sabe a quem perguntar.
    quem     = {"usuario_email": usuario,
                "usuario_nome": APELIDOS.get(usuario, "") or (usuario or "").split("@")[0]}
    itens    = _itens_payload(payload)
    hoje     = date.today().isoformat()
    atualizados, inseridos, ignorados = 0, 0, 0
    aprender = []   # [(entrada_norm, produto_id)] — desfechos pra memória de matching

    for item in itens:
        preco = item.get("preco_un", 0)
        desc  = item.get("descricao_final", "").strip()
        if not desc or not preco or float(preco) <= 0:
            ignorados += 1
            continue

        # Campos de origem — só entram no gravação se o operador realmente preencheu.
        # Regra: "preencheu = atualiza; veio vazio = mantém o que o banco já sabia."
        origem = {}
        try:
            _c = item.get("preco_custo")
            if _c is not None and float(_c) > 0:
                origem["preco_custo"] = float(_c)
        except (TypeError, ValueError):
            pass
        if (item.get("link_fornecedor") or "").strip():
            origem["link_fornecedor"] = item["link_fornecedor"].strip()
        if (item.get("fornecedor") or "").strip():
            origem["fornecedor"] = item["fornecedor"].strip()
        if (item.get("fornecedor_canal") or "").strip():
            origem["fornecedor_canal"] = item["fornecedor_canal"].strip()
        if (item.get("fornecedor_contato") or "").strip():
            origem["fornecedor_contato"] = item["fornecedor_contato"].strip()
        if (item.get("sku_fornecedor") or "").strip():
            origem["sku_fornecedor"] = item["sku_fornecedor"].strip()
        # O datasheet aprovado e' dado do produto, igual a origem: viaja pro banco
        # junto com o preco. Sem isto, o documento fica preso na proposta e o
        # mesmo item precisa ser regerado na proxima cotacao.
        for _campo in ("datasheet_id", "apresentacao_id"):
            try:
                _v = item.get(_campo)
                if _v not in (None, "", 0, "0", False):
                    origem[_campo] = int(_v)
            except (TypeError, ValueError):
                pass

        # IDENTIDADE (A) — quando o item CASOU (idêntico) com uma linha do banco na
        # geração, o front devolve o `banco_id`. Atualiza ELA — sem procurar por texto
        # e sem criar gêmeo cru ("Switch 24 portas" ao lado de "Switch de Rede Gigabit
        # com 24 Portas"). A descrição da linha canônica NÃO é tocada (fica a rica).
        _bid = item.get("banco_id")
        try:
            _bid = int(_bid) if _bid not in (None, "", 0, "0", False) else None
        except (TypeError, ValueError):
            _bid = None

        _pid = None
        if _bid:
            try:
                sb.table("produtos").update({
                    "preco_un": float(preco), "data_ref": hoje,
                    "proposta_tiny": proposta, "cliente": cliente,
                    **origem, **({"cnpj": cnpj} if cnpj else {}), **quem,
                }).eq("id", _bid).execute()
                atualizados += 1
                _pid = _bid
            except Exception:
                _pid = None          # cai no caminho por texto abaixo
        if _pid is None:
          try:
            # REGRA (v3.13): o banco tem UMA linha por descrição.
            #   achou   -> atualiza (o preço mais fresco sempre vence)
            #   não achou -> insere (produto novo)
            #
            # O flag `_alterado` foi REMOVIDO. Ele partia da premissa de que "operador
            # editou = item diferente", e isso é falso: editar a descrição não cria
            # produto. Atributo técnico cria (cor, categoria, bitola, embalagem/unidade)
            # — e quando o atributo muda, a descrição muda junto, o lookup não acha e o
            # insert acontece sozinho, sem precisar de flag. Ruído de cliente
            # ("conforme RC 60938") é a mesma mercadoria e não pode virar produto novo.
            #
            # Efeito colateral do flag: item com descrição IDÊNTICA e preço novo pulava
            # o lookup e inseria linha gêmea a cada chamada — 15 a 21x por proposta.
            res = sb.table("produtos").select("id,preco_un")\
                .ilike("descricao", _ilike_literal(desc)).limit(1).execute()
            if res and res.data:
                sb.table("produtos").update({
                    "preco_un": float(preco), "data_ref": hoje,
                    "proposta_tiny": proposta, "cliente": cliente,
                    **origem,
                    **({"cnpj": cnpj} if cnpj else {}), **quem,
                }).eq("id", res.data[0]["id"]).execute()
                atualizados += 1
                _pid = res.data[0]["id"]
            else:
                try:
                    _ins = sb.table("produtos").insert({
                        "descricao": desc, "un": item.get("unidade", "UN"),
                        "preco_un": float(preco), "data_ref": hoje,
                        "proposta_tiny": proposta, "cliente": cliente,
                        "obs": "inserido automaticamente via app",
                        **origem,
                        **({"cnpj": cnpj} if cnpj else {}), **quem,
                    }).execute()
                    inseridos += 1
                    _pid = ((_ins.data or [{}])[0] or {}).get("id")
                except Exception as e_ins:
                    # Corrida entre operadores (ou índice único): alguém inseriu a mesma
                    # descrição entre o SELECT e o INSERT. Não perde o preço: vira update.
                    if "23505" in str(e_ins) or "duplicate key" in str(e_ins).lower():
                        r2 = sb.table("produtos").select("id")\
                            .ilike("descricao", _ilike_literal(desc)).limit(1).execute()
                        if r2.data:
                            sb.table("produtos").update({
                                "preco_un": float(preco), "data_ref": hoje,
                                "proposta_tiny": proposta, "cliente": cliente,
                                **origem,
                                **({"cnpj": cnpj} if cnpj else {}), **quem,
                            }).eq("id", r2.data[0]["id"]).execute()
                            atualizados += 1
                            _pid = r2.data[0]["id"]
                        else:
                            ignorados += 1
                    else:
                        raise
          except Exception:
            ignorados += 1

        # [4] APRENDIZADO: o par (texto do cliente -> produto que a proposta virou)
        # é FATO, não opinião. Grava sempre que houver os dois lados.
        _entrada = _norm_entrada(item.get("descricao_original") or "")
        if _pid and _entrada:
            aprender.append((_entrada, _pid))

        # [4b] NÓ DA INTERNET: se o operador ESCOLHEU a opção da internet para este
        # item, aprende "(CNPJ + input) -> internet, esta origem" — próxima vez que
        # este CNPJ pedir este input, a busca dispara sozinha. Sem CNPJ, não grava
        # (fica sem âncora — o front já avisou em vermelho). Mesmo gatilho do banco.
        if (_MOTOR_PRECOS_OK and _entrada and cnpj
                and (item.get("origem_escolha") == "internet")):
            try:
                _mp_aprender_no_internet(sb, _entrada, cnpj, item.get("origem_internet") or {})
                _interp = item.get("interpretacao")
                if _interp:
                    _mp_aprender_interpretacao(sb, _entrada, _interp)
            except Exception:
                pass   # aprender não pode derrubar o salvamento da proposta

    memoria = _aprender_memoria(sb, aprender, cliente) if aprender else {}
    return {"atualizados": atualizados, "inseridos": inseridos,
            "ignorados": ignorados, "memoria": memoria}


KIST_CNPJ_RAIZ = "10573732"   # raiz do CNPJ da Kist — pega qualquer filial nossa


def _cnpj_do_cliente(bruto) -> Optional[str]:
    """Devolve o CNPJ do CLIENTE formatado, ou None. Nunca propaga lixo.

    Descarta dois casos:
      1. Dígito verificador inválido — a IA alucinou ou leu errado. CNPJ errado é PIOR que
         CNPJ ausente: ele amarra o preço à empresa errada com cara de dado bom.
      2. Qualquer CNPJ da própria Kist (raiz 10573732). O material do cliente quase sempre
         traz os dois (pedido de compra tem emissor e fornecedor); o nosso é o fornecedor.

    Antes da v3.16 o /extrair aceitava o que a IA devolvesse, sem conferir. A validação só
    existia lá no /gerar-csv, no criar_oc e no atualizar_oc.
    """
    dig = re.sub(r"\D", "", str(bruto or ""))
    if not dig or not _cnpj_valido(dig):
        return None
    if dig.startswith(KIST_CNPJ_RAIZ):
        return None
    return _cnpj_formatado(dig)


def _cnpj_valido(cnpj: str) -> bool:
    """Valida CNPJ pelos dígitos verificadores (independe de formatação)."""
    n = re.sub(r"\D", "", cnpj or "")
    if len(n) != 14 or n == n[0] * 14:
        return False
    def dv(base, pesos):
        s = sum(int(d) * p for d, p in zip(base, pesos)); r = s % 11
        return "0" if r < 2 else str(11 - r)
    p1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]; p2 = [6] + p1
    return n[12] == dv(n[:12], p1) and n[13] == dv(n[:13], p2)


def _cnpj_formatado(cnpj: str) -> str:
    """Reformata os 14 dígitos no padrão 00.000.000/0000-00."""
    n = re.sub(r"\D", "", cnpj or "")
    if len(n) != 14:
        return cnpj or ""
    return f"{n[:2]}.{n[2:5]}.{n[5:8]}/{n[8:12]}-{n[12:]}"


def _cep_formatado(cep: str) -> str:
    n = re.sub(r"\D", "", str(cep or ""))
    return f"{n[:5]}-{n[5:]}" if len(n) == 8 else (str(cep or ""))


def _consulta_receita(cnpj_digitos: str) -> dict:
    """Busca dados cadastrais na BrasilAPI (base da Receita Federal).
    Retorna {} em QUALQUER falha (timeout, fora do ar, CNPJ não achado) —
    a geração da proposta nunca é bloqueada por causa disto."""
    try:
        import urllib.request, json as _json
        url = f"https://brasilapi.com.br/api/cnpj/v1/{cnpj_digitos}"
        req = urllib.request.Request(url, headers={"User-Agent": "kist-cabine/1.0"})
        with urllib.request.urlopen(req, timeout=6) as resp:
            return _json.load(resp) or {}
    except Exception:
        return {}


@app.post("/gerar-csv")
async def gerar_csv(payload: dict, usuario: str = Depends(verificar_token)):
    COLUNAS = [
        'ID','Número da proposta','Data','Data proximo contato','ID contato',
        'Nome do contato','Aos cuidados de','Lista de Preço','Tipo de Pessoa','CPF/CNPJ',
        'RG/IE','CEP','Município','UF','Endereço','Endereço Nro','Complemento','Bairro',
        'Fone','Celular','E-mail','Desconto','Frete','Observações','Validade',
        'Prazo de Entrega','Situação','Introdução','ID produto','Descrição','Quantidade',
        'Valor unitário','Descrição complementar','Vendedor','Destinatário',
        'CPF/CNPJ entrega','CEP entrega','Município entrega','UF entrega',
        'Endereço entrega','Endereço Nro entrega','Complemento entrega','Bairro entrega',
        'Fone entrega','Inscrição Estadual entrega'
    ]

    def fmt_preco(v):
        try:
            f = float(v)
            inteiro = int(f)
            dec = round((f - inteiro) * 1000)
            return f"{inteiro:,}".replace(',', '.') + f",{dec:03d}"
        except:
            return "0,000"

    hoje = date.today().strftime('%d/%m/%Y')

    # Tratamento do CNPJ: válido entra formatado; inválido fica vazio (pra não
    # travar a importação) e o valor captado é sinalizado nas Observações.
    cnpj_raw = (payload.get("cnpj", "") or "").strip()
    cnpj_ok = _cnpj_valido(cnpj_raw)
    cnpj_saida = _cnpj_formatado(cnpj_raw) if cnpj_ok else ""
    aviso_cnpj = "" if (cnpj_ok or not cnpj_raw) else f"CNPJ captado (INVÁLIDO, conferir): {cnpj_raw}"

    # Para CNPJ válido, busca dados cadastrais na Receita (BrasilAPI).
    # Regra: a cotação manda — só preenchemos campos que vierem VAZIOS.
    # Se a consulta falhar, segue com o que tem (rec = {}).
    rec = _consulta_receita(re.sub(r"\D", "", cnpj_raw)) if cnpj_ok else {}
    def _vazio_ou(atual, receita):
        a = (atual or "").strip()
        return a if a else (str(receita).strip() if receita else "")

    nome_contato = _vazio_ou(payload.get("cliente", ""), rec.get("razao_social"))
    cep_val      = _vazio_ou("", _cep_formatado(rec.get("cep")) if rec.get("cep") else "")
    municipio_val= _vazio_ou("", rec.get("municipio"))
    uf_val       = _vazio_ou("", rec.get("uf"))
    endereco_val = _vazio_ou("", rec.get("logradouro"))
    numero_val   = _vazio_ou("", rec.get("numero"))
    bairro_val   = _vazio_ou("", rec.get("bairro"))

    # Campos preenchidos na tela de conferência
    prazo_entrega = (str(payload.get("prazo_entrega", "") or "")).strip()
    def _frete_fmt(v):
        s = str(v or "").strip()
        if not s:
            return "0,00"
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        try:
            return f"{float(s):.2f}".replace(".", ",")
        except Exception:
            return "0,00"
    frete_val = _frete_fmt(payload.get("frete"))

    rows = []
    for item in payload.get("itens", []):
        r = {c: '' for c in COLUNAS}
        r['Número da proposta'] = payload.get("proposta", "")
        r['Data'] = hoje
        r['Nome do contato'] = nome_contato
        r['Tipo de Pessoa'] = 'J'
        r['CPF/CNPJ'] = cnpj_saida
        r['CEP'] = cep_val
        r['Município'] = municipio_val
        r['UF'] = uf_val
        r['Endereço'] = endereco_val
        r['Endereço Nro'] = numero_val
        r['Bairro'] = bairro_val
        r['Desconto'] = '0,00'
        r['Frete'] = frete_val
        obs_base = payload.get("rc_neg", "") or ""
        r['Observações'] = f"{obs_base} | {aviso_cnpj}".strip(" |") if aviso_cnpj else obs_base
        r['Validade'] = '5'
        r['Prazo de Entrega'] = prazo_entrega
        r['Situação'] = 'Rascunho'
        r['ID produto'] = '0'
        r['Descrição'] = item.get("descricao_final", "")
        r['Quantidade'] = f"{int(item.get('quantidade', 1))},00"
        r['Valor unitário'] = fmt_preco(item.get("preco_un", 0))
        # Specs complementares vão para Descrição complementar se existirem
        specs = item.get("specs_complementares", "") or ""
        unidade = item.get("unidade", "UN") or "UN"
        r['Descrição complementar'] = f"{unidade} | {specs[:200]}" if specs else unidade
        rows.append(r)

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=COLUNAS, delimiter=',', quoting=csv.QUOTE_MINIMAL)
    writer.writeheader()
    writer.writerows(rows)
    csv_bytes = '\ufeff' + output.getvalue()
    nome = f"proposta_{payload.get('proposta', 'kist')}.csv"
    return StreamingResponse(io.BytesIO(csv_bytes.encode('utf-8')),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={nome}"})


SYSTEM_CONFERIR = """Você é o consultor técnico da Kist Soluções em Telecom e Energia — materiais
elétricos, telecom, infraestrutura, áudio/vídeo e TI.

Um operador está montando uma proposta comercial e travou num item: o código não resolve, a
descrição do cliente é vaga, ou o que ele achou no fornecedor não parece bater com o pedido.
Ele te chama exatamente como chamaria um colega que conhece part numbers de cor.

Tipos de pergunta que ele faz (são reais):
- "SHURE SB900A e SB900B é a mesma coisa?" → equivalência entre PNs
- "os PNs Neutrik NA3FM e NC3FXX são os mesmos?" → equivalência entre famílias
- "qual a diferença entre SDSQXAA-128G-AN6MA e SDSQXAA-128G-GN6MA?" → o sufixo muda o quê
- "tem diferença entre SWATEMSCN2/2ME1/4K e /1ME1/4K ou é erro de digitação?" → variante ou typo
- "que item é esse: 88037BNM?" / "estilete 481863?" → identificação por código
- "essa placa é de qual item Samsung? BN94-18201Q" → peça → produto pai
- "qual a descrição comercial do PN 02.07.056 da Resideo?" → PN → nome comercial
- "p10 Santo Ângelo ST ninja, qual o nome comercial?" → apelido de mercado → nome real
- "qual PN do switch mais em conta de 48p giga e 4 SFP?" → spec → PN sugerido

COMO RESPONDER:
- Responda a pergunta primeiro, em uma ou duas frases. O operador está no meio do trabalho.
- Depois o porquê, curto. Ele decide, você informa.
- Quando comparar dois itens, seja CONCRETO no que difere: "o AN6MA é A2 (app performance),
  o GN6MA é A1" — não "são versões diferentes".
- Quando for erro de digitação, diga: ele quer saber se pede 1 ou 2 do item.

BUSQUE NA WEB quando o código for específico (PN de fabricante, SKU, código de peça) ou quando
você não tiver certeza. Datasheet do fabricante e distribuidor oficial valem mais que
marketplace. É melhor buscar e confirmar do que responder de memória e errar um PN.

QUANDO NÃO DER PRA RESPONDER:
- Descrição genérica demais pra identificar item ("cabo 2x1,5mm" existe de 50 fabricantes):
  diga isso, e diga EXATAMENTE qual informação resolveria. Ele vai perguntar ao cliente.
- Não achou o código em lugar nenhum: diga que não achou. NÃO invente PN, fabricante ou spec.
  PN errado faz a Kist comprar errado, entregar errado e o cliente devolver — o silêncio custa
  um e-mail, o chute custa um RMA.
- Nunca preencha lacuna com o que "provavelmente é".

Português do Brasil, direto, sem preâmbulo. Nada de "ótima pergunta" ou "claro!". Vai direto.
"""


@app.post("/conferir")
async def conferir(payload: dict, usuario: str = Depends(verificar_token)):
    """Consulta técnica sobre UM item, com o contexto já carregado.

    Substitui o /sugerir-pn, que nasceu com a intenção certa (usar IA pra resolver
    item que não está claro) e morreu por dois motivos: um veto de custo que hoje
    não existe mais (~$1,37/mês o sistema inteiro), e uma lista de commodities que
    vetava por substring — 'SUPORTE' cegava o botão justamente no item que voltou
    em RMA.

    O operador já faz isto hoje, numa aba de chat, colando os textos na mão:
      "SHURE SB900A e SB900B é a mesma coisa?"
      "qual a diferença entre SDSQXAA-128G-AN6MA e GN6MA?"
      "que item é esse: 88037BNM?"
      "qual a descrição comercial do PN 02.07.056 da Resideo?"
      "qual PN do switch mais em conta de 48p giga e 4 SFP?"
    A diferença é que aqui o item já vem carregado, a resposta volta clicável, e a
    busca web é do sistema — não da aba dele.

    Fica FORA do caminho crítico de propósito: o matching precisa ser determinístico
    (temperature=0) e rápido; busca web é nem uma coisa nem outra. Aqui o operador
    pediu e está esperando — 3s é o que ele já gasta hoje.
    """
    pergunta = (payload.get("pergunta") or "").strip()
    if not pergunta:
        raise HTTPException(400, "Pergunta obrigatória")

    item     = payload.get("item") or {}
    fonte    = (payload.get("fonte_texto") or "").strip()
    historico_msgs = payload.get("historico") or []

    # ── Contexto: tudo que o sistema sabe deste item ────────────────────────
    ctx = ["### O QUE O CLIENTE PEDIU"]
    ctx.append(f"Descrição: {item.get('descricao_original') or '(vazio)'}")
    _sp = (item.get("specs_complementares") or "").strip()
    ctx.append(f"Specs complementares: {_sp}" if _sp else "Specs complementares: (o cliente não informou)")
    if item.get("quantidade"):
        ctx.append(f"Quantidade: {item.get('quantidade')} {item.get('unidade') or 'UN'}")

    ctx.append("\n### O QUE A KIST ESTÁ OFERTANDO")
    _df = (item.get("descricao_final") or "").strip()
    ctx.append(f"Descrição preenchida: {_df or '(ainda não preenchida)'}")
    for rot, campo in (("Fornecedor", "fornecedor"), ("Link/origem", "link_fornecedor"),
                       ("SKU do fornecedor", "sku_fornecedor")):
        if (item.get(campo) or "").strip():
            ctx.append(f"{rot}: {item[campo]}")

    banco = item.get("banco") or {}
    if banco.get("descricao"):
        ctx.append("\n### O QUE O BANCO DE PREÇOS PROPÔS")
        ctx.append(f"Produto: {banco['descricao']}")
        if banco.get("fornecedor"):
            ctx.append(f"Fornecedor: {banco['fornecedor']}")

    if fonte:
        # A spec completa costuma estar no e-mail/PDF, não nos campos destilados.
        ctx.append("\n### E-MAIL / ANEXOS ORIGINAIS DO CLIENTE (fonte da cotação)")
        ctx.append(fonte[:14000])

    contexto = "\n".join(ctx)

    msgs = []
    for m in historico_msgs[-6:]:
        papel = "assistant" if m.get("role") == "assistant" else "user"
        if (m.get("content") or "").strip():
            msgs.append({"role": papel, "content": m["content"][:4000]})
    msgs.append({"role": "user", "content": f"{contexto}\n\n### PERGUNTA\n{pergunta}"})

    claude = get_claude()
    try:
        resp = claude.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1500,
            system=SYSTEM_CONFERIR,
            messages=msgs,
            tools=[{"type": "web_search_20250305", "name": "web_search", "max_uses": 4}],
            timeout=90.0,
        )
    except Exception as e:
        raise HTTPException(502, f"Não consegui consultar agora: {type(e).__name__}")

    # A resposta mistura texto e blocos de busca — junta só o texto.
    texto = "\n".join(b.text for b in resp.content if getattr(b, "type", "") == "text").strip()
    buscas = [b.input.get("query", "") for b in resp.content
              if getattr(b, "type", "") == "server_tool_use" and getattr(b, "input", None)]
    if not texto:
        texto = "Não consegui formular uma resposta. Tenta reformular a pergunta."

    return {"resposta": texto, "buscas": [q for q in buscas if q]}


@app.post("/ficha-internet")
# SÍNCRONO DE PROPÓSITO (incidente 23/07). Este endpoint faz I/O BLOQUEANTE:
# SerpApi via urllib, cliente Anthropic síncrono e leitura da página do produto —
# uma busca leva de 10s a 30s. Declarado como `async def`, esse bloqueio trava o
# EVENT LOOP INTEIRO: enquanto uma busca roda, nenhuma outra requisição é atendida
# e o operador vê "sem contato com o banco" (não lista propostas, não pega número).
# Como `def` comum, o Starlette executa a função num THREADPOOL e o event loop
# segue livre para servir a tela. Não custa plano nem worker novo — e é melhor que
# aumentar workers, porque cada worker sozinho continuaria travando o próprio loop.
# NÃO reconverter para `async def` sem antes tornar as chamadas internas assíncronas.
def ficha_internet(payload: dict, usuario: str = Depends(verificar_token)):
    """Referência de preço na internet para UM item SEM match no banco.

    Chamado pelo frontend item a item, de forma assíncrona (FORA do /extrair): a
    tela sobe na hora com o matching do banco, e cada ficha da internet preenche
    sua 3ª coluna quando fica pronta. Um item que falha não derruba os outros.

    O motor faz camada 2 (cache de fichas, TTL 24h) -> camada 3 (cascata de busca
    + julgamento). A camada 1 (banco de preços) é o matching que já roda no
    /extrair — por isso este endpoint só é chamado para item sem match.

    NÃO escreve preço no banco: devolve referência. O que migra pra `produtos` é
    sempre o preço que o operador lança (fluxo normal de /salvar-proposta).
    """
    if not _MOTOR_PRECOS_OK:
        raise HTTPException(503, "Motor de preços indisponível (backend/motor_precos.py ausente)")

    item = payload.get("item") or {}
    desc = (item.get("descricao") or item.get("descricao_original") or "").strip()
    if not desc:
        raise HTTPException(400, "Item sem descrição")

    # O motor lê 'descricao'/'specs_complementares'/'quantidade'/'unidade'.
    # Normaliza descricao_original -> descricao sem mutar o payload do cliente.
    item_motor = dict(item)
    if not item_motor.get("descricao"):
        item_motor["descricao"] = desc

    # CNPJ da proposta = âncora do nó (client-first). termo_rebusca = correção do
    # operador (reescreveu o termo): o motor usa o termo dele e aprende com isso.
    cnpj = _cnpj_do_cliente(payload.get("cnpj"))
    termo_rebusca = (payload.get("termo_rebusca") or "").strip() or None

    try:
        ficha = _resolver_ficha_precos(item_motor, get_supabase(), get_claude(),
                                       usuario_email=usuario, cnpj=cnpj,
                                       termo_rebusca=termo_rebusca)
    except Exception as e:
        raise HTTPException(502, f"Não consegui buscar preço agora: {type(e).__name__}")
    return ficha


# ── PROPOSTAS ─────────────────────────────────────────────────────────────────

@app.post("/produto-alerta")
async def salvar_alerta_produto(payload: dict, usuario: str = Depends(verificar_token)):
    """Salva/atualiza alerta de produto. Busca por descricao (ilike).
    payload: {descricao, alerta: {texto, links[], thumb_b64}, alerta_imagem?}
    """
    import json as _jap
    sb = get_supabase()
    descricao = (payload.get("descricao") or "").strip()
    alerta_obj = payload.get("alerta") or {}
    alerta_imagem = payload.get("alerta_imagem")  # base64 full — None = não enviado

    if not descricao:
        raise HTTPException(status_code=400, detail="descricao obrigatória")

    alerta_json = _jap.dumps(alerta_obj, ensure_ascii=False) if alerta_obj else None
    if alerta_obj and not any([alerta_obj.get("texto"), alerta_obj.get("links"), alerta_obj.get("thumb_b64")]):
        alerta_json = None

    res = sb.table("produtos").select("id").ilike("descricao", descricao).limit(1).execute()
    if not res.data:
        return {"ok": True, "aviso": "produto não encontrado no banco"}

    update_payload = {"alerta": alerta_json}
    if alerta_imagem is not None:
        update_payload["alerta_imagem"] = alerta_imagem or None

    sb.table("produtos").update(update_payload).eq("id", res.data[0]["id"]).execute()
    return {"ok": True}


@app.get("/produto-alerta-imagem")
async def buscar_alerta_imagem(descricao: str, usuario: str = Depends(verificar_token)):
    """Retorna a imagem full (base64) do alerta. Carregada sob demanda (click)."""
    sb = get_supabase()
    if not descricao.strip():
        raise HTTPException(status_code=400, detail="descricao obrigatória")
    res = sb.table("produtos").select("alerta_imagem").ilike("descricao", descricao).limit(1).execute()
    if not res.data:
        return {"alerta_imagem": None}
    return {"alerta_imagem": res.data[0].get("alerta_imagem")}


def _norm_item_payload(it):
    """Garante que um item vindo do payload seja SEMPRE um dict.

    O modelo de extração às vezes devolve item como string, e esse formato viaja
    pelo frontend de volta aos endpoints que salvam/geram (salvar-proposta,
    upsert-precos, gerar-csv). Sem isto, `item.get(...)` estoura
    'str' object has no attribute 'get' — e como o auto-save dispara logo após a
    extração, o operador via a proposta na tela e o erro ao mesmo tempo.
    String vira {descricao...}; qualquer outra coisa vira dict vazio (não derruba).
    """
    if isinstance(it, dict):
        return it
    if isinstance(it, str):
        return {"descricao_original": it, "descricao_final": it,
                "descricao": it, "quantidade": 1, "unidade": "UN"}
    return {}


def _itens_payload(payload):
    """Lista de itens do payload, cada um garantidamente dict."""
    return [_norm_item_payload(i) for i in (payload.get("itens") or [])]


@app.post("/salvar-proposta")
async def salvar_proposta(payload: dict, usuario: str = Depends(verificar_token)):
    """Upsert de proposta e itens. status: 'rascunho' | 'confirmada'.
    Identifica pelo numero_proposta — cria se não existir, atualiza se existir.
    Itens antigos são substituídos pelos novos (delete + insert).
    """
    sb = get_supabase()
    itens = _itens_payload(payload)
    valor_total = sum(
        float(i.get("preco_un") or 0) * float(i.get("quantidade") or 1)
        for i in itens
    )
    com_preco  = sum(1 for i in itens if float(i.get("preco_un") or 0) > 0)
    sem_preco  = len(itens) - com_preco
    status     = payload.get("status", "confirmada")
    numero     = str(payload.get("proposta") or payload.get("numero_proposta") or "")

    prop_data = {
        "numero_proposta":      numero,
        "cliente":              payload.get("cliente", ""),
        "cnpj":                 payload.get("cnpj") or None,
        "rc_neg":               payload.get("rc_neg") or None,
        "usuario_email":        usuario,
        "usuario_nome":         payload.get("usuario_nome", ""),
        "total_itens":          len(itens),
        "com_preco":            com_preco,
        "sem_preco":            sem_preco,
        "valor_total_estimado": valor_total,
        "frete_recebimento":    float(payload.get("frete") or payload.get("frete_recebimento") or 0),
        "prazo_entrega":        payload.get("prazo_entrega") or None,
        "status":               status,
        # Texto que a IA leu na extração. Só grava quando vier — reabrir uma proposta
        # e salvar de novo não pode apagar a fonte com string vazia.
        **({"fonte_texto": str(payload["fonte_texto"])[:60000]}
           if (payload.get("fonte_texto") or "").strip() else {}),
    }

    # Upsert: buscar pelo numero_proposta
    existing = sb.table("propostas").select("id").eq("numero_proposta", numero).limit(1).execute()
    if existing.data:
        proposta_id = existing.data[0]["id"]
        sb.table("propostas").update(prop_data).eq("id", proposta_id).execute()
        # Substituir itens
        sb.table("itens_proposta").delete().eq("proposta_id", proposta_id).execute()
    else:
        res = sb.table("propostas").insert(prop_data).execute()
        proposta_id = res.data[0]["id"]

    if itens:
        rows = [{
            "proposta_id":          proposta_id,
            "descricao_original":   i.get("descricao_original", ""),
            "descricao_final":      i.get("descricao_final", ""),
            "codigo_cliente":       (i.get("codigo_cliente") or "").strip() or None,
            "quantidade":           float(i.get("quantidade") or 1),
            "unidade":              i.get("unidade", "UN"),
            "preco_venda":          float(i.get("preco_un") or 0),
            "preco_custo":          float(i.get("preco_custo") or 0),
            "frete_vinda":          float(i.get("frete_vinda") or 0),
            "confianca_match":      i.get("confianca_match", ""),
            "specs_complementares": i.get("specs_complementares", ""),
            "fornecedor":           i.get("fornecedor", ""),
            "link_fornecedor":      i.get("link_fornecedor", ""),
            "fornecedor_canal":     i.get("fornecedor_canal", ""),
            "fornecedor_contato":   i.get("fornecedor_contato", ""),
            "sku_fornecedor":       i.get("sku_fornecedor", ""),
            "obs_interna":          i.get("obs_interna", ""),
            "datasheet_id":         i.get("datasheet_id") or None,
            "apresentacao_id":      i.get("apresentacao_id") or None,
        } for i in itens]
        sb.table("itens_proposta").insert(rows).execute()

    return {"proposta_id": proposta_id, "total_itens": len(itens), "status": status}


@app.get("/propostas/{proposta_id}/detalhe")
async def detalhe_proposta(proposta_id: int, usuario: str = Depends(verificar_token)):
    """Retorna proposta completa + itens para reabrir na tela de revisão."""
    sb = get_supabase()
    prop = sb.table("propostas").select("*").eq("id", proposta_id).limit(1).execute()
    if not prop.data:
        raise HTTPException(status_code=404, detail="Proposta não encontrada")
    itens = sb.table("itens_proposta").select("*").eq("proposta_id", proposta_id).execute()
    return {"proposta": prop.data[0], "itens": itens.data or []}


@app.get("/propostas")
async def listar_propostas(
    busca: str = None,
    numero: str = None,
    cnpj: str = None,
    usuario_email: str = None,
    data_inicio: str = None,
    data_fim: str = None,
    todos: bool = False,
    limit: int = 50,
    usuario: str = Depends(verificar_token)
):
    """Lista propostas com filtros.
    `busca` é a busca única da barra de pesquisa: casa com número da proposta,
    cliente, CNPJ OU descrição de item (ex.: 'Convergint' ou 'MC200L')."""
    sb = get_supabase()
    q = sb.table("propostas").select("*").order("data_geracao", desc=True).limit(limit)

    # Por padrão mostra só do usuário autenticado, salvo se todos=True
    if not todos:
        q = q.eq("usuario_email", usuario)

    if busca and busca.strip():
        # sanitiza p/ não quebrar a sintaxe do filtro OR do PostgREST
        termo = re.sub(r"[,()*]", " ", busca.strip()).strip()
        if termo:
            # propostas cujos ITENS casam na descrição
            ids_por_item = []
            try:
                it = (sb.table("itens_proposta")
                        .select("proposta_id")
                        .or_(f"descricao_final.ilike.*{termo}*,descricao_original.ilike.*{termo}*")
                        .limit(500).execute())
                ids_por_item = list({r["proposta_id"] for r in (it.data or []) if r.get("proposta_id") is not None})
            except Exception:
                ids_por_item = []
            or_parts = [
                f"numero_proposta.ilike.*{termo}*",
                f"cliente.ilike.*{termo}*",
                f"cnpj.ilike.*{termo}*",
            ]
            if ids_por_item:
                or_parts.append(f"id.in.({','.join(str(i) for i in ids_por_item)})")
            q = q.or_(",".join(or_parts))

    # filtros específicos (compatibilidade / uso programático)
    if numero:
        q = q.ilike("numero_proposta", f"%{numero}%")
    if cnpj:
        q = q.ilike("cnpj", f"%{cnpj}%")
    if usuario_email:
        q = q.ilike("usuario_email", f"%{usuario_email}%")
    if data_inicio:
        q = q.gte("data_geracao", data_inicio)
    if data_fim:
        q = q.lte("data_geracao", data_fim + "T23:59:59")

    res = q.execute()
    return res.data


@app.get("/propostas/{proposta_id}/itens")
async def itens_proposta(proposta_id: int, usuario: str = Depends(verificar_token)):
    """Retorna itens de uma proposta"""
    sb = get_supabase()
    res = sb.table("itens_proposta").select("*").eq("proposta_id", proposta_id).execute()
    return res.data


@app.put("/itens-proposta/{item_id}/origem")
async def atualizar_origem_item(
    item_id: int, payload: dict, usuario: str = Depends(verificar_token)
):
    """Atualiza campos internos de origem de um item"""
    sb = get_supabase()
    sb.table("itens_proposta").update({
        "fornecedor":      payload.get("fornecedor", ""),
        "link_fornecedor": payload.get("link_fornecedor", ""),
        "sku_fornecedor":  payload.get("sku_fornecedor", ""),
        "obs_interna":     payload.get("obs_interna", ""),
    }).eq("id", item_id).execute()
    return {"ok": True}


# ── ORDENS DE COMPRA ──────────────────────────────────────────────────────────

# ============================================================================
# Casar PO do cliente com proposta salva (matcher) — reusa leitura .msg/PDF + Claude
# ============================================================================
SYSTEM_PO = """Você é o assistente de compras da Kist Soluções em Telecom e Energia.
Você recebe um e-mail (assunto + corpo) e possivelmente documentos PDF anexos.

SUA TAREFA: identificar a ORDEM DE COMPRA (PO) emitida pelo cliente para a Kist e extrair os itens dela.

ANÁLISE CRÍTICA OBRIGATÓRIA:
- O e-mail pode conter: a PO do cliente, proposta da Kist, logotipos, T&Cs.
- Itens vêm SOMENTE da PO do cliente — nunca da proposta comercial da Kist.
- Se houver proposta Kist: use APENAS para identificar cliente/CNPJ.
- Quantidade: coluna QTDE/Quantidade da PO. Dimensões (ex: 430MM X 240MM) NÃO são quantidade.
- Preco_unitario: valor unitário na PO. Se não houver, use 0.
- Não invente itens. Se não identificar a PO, retorne itens vazio.

Retorne APENAS JSON puro (sem markdown, sem ``` ):
{"itens":[{"descricao":"...","quantidade":0,"preco_unitario":0}],"destino":"endereço/cidade-UF de entrega ou ''"}"""

SYSTEM_PROP_TINY = """Você recebe o texto de uma PROPOSTA COMERCIAL / orçamento (sistema Tiny).
Extraia em JSON PURO (sem markdown, sem ``` ):
{"cliente":"...","cnpj":"...","numero_proposta":"...","itens":[{"descricao":"...","quantidade":0,"preco_venda":0}]}
Números com ponto decimal. Sem dado -> "" ou 0. Não invente itens. Só o JSON."""

def _pdf_po_texto(data):
    """Extrai texto de PDF de PO.
    - Sempre inclui página 1 (CNPJ, número da PO).
    - Para PDFs com padrão 'Item:NNNNN' (Embraer/SAP): inclui APENAS as páginas
      que contêm itens, descartando as páginas de T&C.
    - Para outros PDFs: comportamento original (extract_tables + extract_text).
    """
    try:
        import pdfplumber, io as _io
        partes = []
        with pdfplumber.open(_io.BytesIO(data)) as pdf:
            textos_paginas = [(page.extract_text() or "").strip() for page in pdf.pages]

        tem_item_nnnnn = any(re.search(r'\bItem:\d{5}', t) for t in textos_paginas)

        if tem_item_nnnnn:
            for i, txt in enumerate(textos_paginas):
                if i == 0 or re.search(r'\bItem:\d{5}', txt):
                    if txt:
                        partes.append(txt)
        else:
            with pdfplumber.open(_io.BytesIO(data)) as pdf:
                for page in pdf.pages:
                    for table in (page.extract_tables() or []):
                        # Pular tabelas header-only (1 linha) — a IA usaria
                        # a tabela vazia como "fonte principal" e ignoraria
                        # o extract_text onde os itens realmente estão
                        if len(table or []) <= 1:
                            continue
                        rows = []
                        for row in (table or []):
                            cells = [str(c or "").replace("\n", " ").strip() for c in row]
                            rows.append(" | ".join(cells))
                        if rows:
                            partes.append("\n".join(rows))
                    txt = (page.extract_text() or "").strip()
                    if txt:
                        partes.append(txt)

        return "\n\n".join(partes)
    except Exception:
        return ""


# ═══════════════════════════════════════════════════════════════════════════
# EXTRAÇÃO DE PO — Arquitetura em 3 camadas:
#   1. Parser determinístico (regex com âncoras confiáveis, sem IA)
#   2. Validação cruzada independente (subtotal PDF × soma dos itens)
#   3. IA como fallback + revalidação
# ═══════════════════════════════════════════════════════════════════════════

def _parse_numero(s):
    """Converte número do PDF para float detectando o formato separador.
    - '6.00' ou '11.00'  → ponto decimal (Anglo/SAP) → float direto
    - '59,60'            → vírgula decimal (BR)
    - '1.234,56'         → ponto milhar + vírgula decimal (BR extenso)
    """
    s = re.sub(r'[^\d.,]', '', s or '')
    if not s:
        return 0.0
    if '.' in s and ',' in s:
        return float(s.replace('.', '').replace(',', '.'))
    elif ',' in s:
        return float(s.replace(',', '.'))
    else:
        return float(s)


def _extrair_validacao_po(texto):
    """Extrai dados de validação INDEPENDENTES do corpo do PDF
    (subtotal, valor total, n° de itens). Usados para checar qualquer
    método de extração sem se basear nos itens em si."""
    dados = {}
    m = re.search(r'[Ss]ubtotal\s+R\$\s*([\d.]+,\d{2})', texto)
    if m:
        dados['subtotal'] = _parse_numero(m.group(1))
    m = re.search(r'[Vv]alor\s+[Tt]otal[:\s]+R\$\s*([\d.]+,\d{2})', texto)
    if m:
        dados['valor_total'] = _parse_numero(m.group(1))
    # "Total dos itens 479,78" (formato proposta Kist/Tiny)
    m = re.search(r'[Tt]otal\s+dos\s+itens\s+([\d.,]+)', texto)
    if m and 'subtotal' not in dados:
        dados['subtotal'] = _parse_numero(m.group(1))
    # "Número de itens: N"
    m = re.search(r'[Nn][uú]mero\s+de\s+itens[:\s]+(\d+)', texto)
    if m:
        dados['n_itens'] = int(m.group(1))
    return dados


def _validar_itens_po(itens, dados_val):
    """Valida itens extraídos contra os totais independentes do PDF.
    Retorna (ok: bool, avisos: list[str]).
    - Subtotal: soma(qtde × preco) deve bater com o subtotal do PDF (tol. R$1)
    - Contagem: n° itens deve bater com 'Número de itens' do PDF (se presente)
    """
    if not itens:
        return False, ['nenhum item extraído']
    avisos = []
    sub_calc = sum(_parse_numero(str(i.get('quantidade', 0))) *
                   _parse_numero(str(i.get('preco_unitario', 0))) for i in itens)
    if 'subtotal' in dados_val and dados_val['subtotal'] > 0:
        diff = abs(sub_calc - dados_val['subtotal'])
        if diff > 1.0:
            avisos.append(
                f"subtotal diverge: calculado R${sub_calc:.2f} "
                f"≠ PDF R${dados_val['subtotal']:.2f} (Δ R${diff:.2f})"
            )
    if 'n_itens' in dados_val and len(itens) != dados_val['n_itens']:
        avisos.append(
            f"contagem diverge: extraídos {len(itens)} "
            f"≠ PDF {dados_val['n_itens']} itens"
        )
    return len(avisos) == 0, avisos


def _parsear_itens_convergint(texto):
    """Parser determinístico para POs no formato Convergint / SAP Business One.
    Âncora principal: data DD/MM/YYYY divide descrição dos dados numéricos.
    Cada linha de item tem: SEQ SKU DESCRIÇÃO DATA [UNID] QTDE R$ PREÇO R$ TOTAL

    Detectado por: presença de 'DD/MM/YYYY' + 'R$ X,XX  R$ X,XX' em linhas de item.
    Validação item-a-item: QTDE × PREÇO ≈ TOTAL (tolerância R$0,02/unidade).
    """
    if not re.search(
        r'^\s*\d+\s+\S+\s.+\d{2}/\d{2}/\d{4}.+R\$\s*[\d.]+,\d{2}\s+R\$',
        texto, re.M
    ):
        return None

    ITEM_PAT = re.compile(
        r'^\s*(\d+)\s+'          # seq
        r'(\S+)\s+'               # SKU
        r'(.+?)\s+'                # descrição (lazy, não cruza linha)
        r'\d{2}/\d{2}/\d{4}\s+' # data (âncora, consumida)
        r'(?:[A-Za-zÀ-ÿ]{1,4}\s+)?' # unidade opcional (UN/un/PC/m…)
        r'(\d+[.,]\d{1,3})\s+'   # quantidade
        r'R\$\s*([\d.]+,\d{2})\s+' # preço unitário
        r'R\$\s*([\d.]+,\d{2})', # valor total
        re.MULTILINE
    )

    # Remove NCM da descrição — capturado porque fica entre descrição e data no PDF.
    # Trata NCMs limpos (8536.90.90) e garbled (PR8A5T3A6.90.90).
    _NCM_TRAIL  = re.compile(r'\s+\S*\.\d{2}\.\d{2}\S*$')
    _PREP_TRAIL = re.compile(r'\s+(?:DE|DO|DA|DOS|DAS|EM|NO|NA|E|A|O|,)\s*$', re.I)
    def _limpar_desc(d):
        d = _NCM_TRAIL.sub('', d).strip()
        return _PREP_TRAIL.sub('', d).strip()

    itens = []
    for m in ITEM_PAT.finditer(texto):
        desc  = _limpar_desc(m.group(3).strip())
        qtd   = _parse_numero(m.group(4))
        preco = _parse_numero(m.group(5))
        total = _parse_numero(m.group(6))

        # Validação item: qtde × preço ≈ total (tol. 2 cents por unidade)
        esperado = round(qtd * preco, 2)
        preco_ok = abs(esperado - total) <= max(0.02 * max(1, qtd), 0.05)
        itens.append({
            'descricao':    desc,
            'quantidade':   qtd,
            'preco_unitario': preco if preco_ok else 0.0,
        })

    return itens if itens else None


def _parsear_itens_po_nativo(texto):
    """Parser direto para POs no formato Item:NNNNN (Embraer, SAP e similares).
    Retorna lista de itens ou None se padrão não detectado."""
    if not re.search(r'\bItem:\d{5}', texto):
        return None

    itens = []
    blocos = re.split(r'(?=\bItem:\d{5}\b)', texto)
    skip_prefixes = ("Quantidade", "Utiliza", "GPX", "Item:", "Assinado", "ValorTotal")

    for bloco in blocos:
        if not re.match(r'\s*Item:\d{5}', bloco):
            continue
        linhas = [l.strip() for l in bloco.split("\n") if l.strip()]
        if not linhas:
            continue
        linha1 = linhas[0]
        pn = ""
        pn_m = re.search(r'\bPN:(\S+)', linha1)
        if pn_m:
            pn = pn_m.group(1).rstrip("-/")
        den = ""
        den_m = re.search(r'Denomina\w+:(.+)$', linha1, re.I)
        if den_m:
            den = den_m.group(1).strip()
        qtd, preco = 1, 0.0
        for linha in linhas[1:5]:
            m = re.match(r'\d{2}\.\w{3,4}\.\d{4}\s+(\d+)\s+\w+\s+([\d.]+,\d{2})', linha)
            if m:
                qtd = int(m.group(1))
                try:
                    preco = float(m.group(2).replace(".", "").replace(",", "."))
                except Exception:
                    pass
                break
        complemento = ""
        for linha in linhas[2:7]:
            if any(linha.startswith(p) for p in skip_prefixes):
                continue
            if re.match(r'\d{2}\.\w{3,4}\.\d{4}', linha):
                continue
            if re.match(r'^[A-Z\xC0-\xFF0-9/,.+ ()*\-]+$', linha) and len(linha) > 3:
                complemento = linha
                break
        if complemento:
            descricao = f"{complemento} PN:{pn}" if pn else complemento
        elif den:
            descricao = f"{den} PN:{pn}" if pn else den
        else:
            descricao = f"PN:{pn}" if pn else ""
        if descricao:
            itens.append({"descricao": descricao, "quantidade": qtd, "preco_unitario": preco})

    return itens if itens else None

# PDFs irrelevantes em emails de PO (políticas, T&Cs, etc.)
_PDF_SKIP = re.compile(
    r'politic|policy|pagamento|payment|entrega|delivery'
    r'|termo|term|condi|cartilha|recebimento|manual|instruc'
    r'|instrucao|instrução|procedimento|cadastro|homologac'
    r'|geral|norma|regul|contrato|contract|acordo|agreement',
    re.I
)
# PDFs de lista de itens nunca passam de 1,5 MB — acima disso é manual/T&C
_PDF_MAX_BYTES = 1_500_000

# Nomes de PDF claramente irrelevantes para PO (política, cartilha, manual)
_PO_SKIP = re.compile(r'politic|cartilha|manual|instrucao|procedimento', re.I)
# Limite de tamanho para PDF como documento nativo (10 MB — Anthropic aceita até 32 MB)
_DOC_MAX_BYTES = 10_000_000

async def _ler_po(arquivo):
    """Lê arquivo de PO e retorna (texto, pdf_docs).
    - texto: body do email + texto extraído de PDFs legíveis
    - pdf_docs: [(nome, bytes)] de PDFs baseados em imagem (para Claude Vision/Document)
    Usada tanto para o arquivo principal quanto para proposta_tiny.
    """
    nome = (arquivo.filename or "").lower()
    dados = await arquivo.read()

    if nome.endswith(".msg"):
        with open("/tmp/po_upload.msg", "wb") as f:
            f.write(dados)
        msg = extract_msg.openMsg("/tmp/po_upload.msg")

        partes_texto = []
        pdf_docs = []

        corpo = f"Assunto: {msg.subject}\n\n{(msg.body or '').strip()}"
        if corpo.strip():
            partes_texto.append(corpo)

        for att in msg.attachments:
            fn = att.longFilename or att.shortFilename or ""
            flo = fn.lower()
            if not att.data or not flo.endswith(".pdf"):
                continue
            if _PO_SKIP.search(flo):
                continue  # ignorar T&C, cartilhas, manuais
            if len(att.data) > _DOC_MAX_BYTES:
                continue  # muito grande para documento nativo

            t = _pdf_po_texto(att.data)
            if t.strip():
                partes_texto.append(f"[PDF: {fn}]\n{t}")
            else:
                # PDF baseado em imagem (escaneado) → documento nativo para Vision
                pdf_docs.append((fn, att.data))

        texto = "\n\n[---]\n\n".join(p for p in partes_texto if p.strip())
        return texto, pdf_docs

    if nome.endswith(".pdf"):
        t = _pdf_po_texto(dados)
        if t.strip():
            return t, []
        else:
            return "", [(arquivo.filename or "po.pdf", dados)]

    try:
        return dados.decode("utf-8", "ignore"), []
    except Exception:
        return "", []

def _digitos(s):
    return re.sub(r"\D", "", s or "")

def _toks(s):
    return set(re.findall(r"[a-z0-9]+", (s or "").lower()))

# ══════════════════════════════════════════════════════════════════════════════
# CERTEZA DE "MESMO ITEM" — PO do cliente x proposta casada (v3.23)
#
# A PO não é um documento independente: é o cliente confirmando QUAIS LINHAS da
# NOSSA proposta ele está comprando. O preço que está na PO saiu do CSV que a
# Kist mandou pro Tiny. Logo o preço é chave compartilhada, não "bônus".
#
# O que quebrou na PO-0000090887 (Igreja Universal, 29/07):
#   1. a comparação olhava só `descricao_final`. Desde o motor de internet, o
#      "usar esta" reescreve a final com o título do anúncio — a palavra do
#      cliente sobrevive só na `descricao_original`, e o matcher não a via.
#      O FUNDO PREPARADOR era texto IDÊNTICO à original e foi recusado.
#   2. exigia descrição parecida (>=0.6) mesmo com preço batendo ao centavo.
#      O cliente emite a PO com a descrição do catálogo INTERNO dele
#      ("MINI DISJUNTOR BIPOLAR 16A TIPO C"), que nunca vai parecer com o
#      título do anúncio. Similaridade media: 0.18 a 0.36. Recusava tudo.
# Resultado: 4 dos 6 itens chegaram na OC sem custo e sem fornecedor, com a
# origem preenchida e disponível no banco. Isso é falha de regra de negócio.
# ══════════════════════════════════════════════════════════════════════════════

def _norm_preco(v):
    try:
        return round(float(v or 0), 2)
    except (TypeError, ValueError):
        return 0.0

def _sim_desc(dtok, prop_item):
    """Similaridade da descrição da PO contra AS DUAS descrições da proposta.
    Vale a melhor: a `original` é a palavra do cliente, a `final` é o texto
    comercial. Esconder qualquer uma das duas do matcher perde herança."""
    melhor, igual = 0.0, False
    for campo in ("descricao_original", "descricao_final"):
        itoks = _toks(prop_item.get(campo))
        if not itoks or not dtok:
            continue
        if dtok == itoks:
            igual = True
        uni = len(dtok | itoks); inter = len(dtok & itoks)
        if uni:
            melhor = max(melhor, inter / uni)
    return melhor, igual

def _contagem_precos(itens, campo):
    """Quantas linhas têm cada preço. Preço que aparece 1x só é chave única."""
    c = {}
    for it in (itens or []):
        p = _norm_preco(it.get(campo))
        if p > 0:
            c[p] = c.get(p, 0) + 1
    return c

def _item_certo(dtok, preco_po, prop_item, preco_ancora=False):
    """'Mesmo item' com ALTA certeza. Três caminhos, todos conservadores:
       (a) descrição idêntica (contra original OU final);
       (b) preço exato + ÂNCORA — aquele preço aparece uma única vez na PO E uma
           única vez na proposta, ou seja o pareamento é 1-para-1 e não há em
           quem errar. Só o chamador ancorado (a proposta casada) passa True;
       (c) preço exato + descrição bem parecida (regra histórica, >=0.6).
    Fora disso NÃO carrega — na dúvida, em branco continua melhor que dado do
    item errado."""
    if not dtok:
        return False
    sim, igual = _sim_desc(dtok, prop_item)
    if igual:
        return True
    pv = _norm_preco(prop_item.get("preco_venda"))
    preco_bate = preco_po > 0 and pv > 0 and abs(pv - preco_po) < 0.01
    if not preco_bate:
        return False
    return bool(preco_ancora) or sim >= 0.6

# campos de compra que a proposta empresta pra OC. `fornecedor_canal` e
# `fornecedor_contato` (v3.20) FALTAVAM aqui — a OC nascia sem o WhatsApp/e-mail
# do fornecedor, e item cuja origem é contato (sem link) chegava na OC sem lastro
# e nunca alimentava o banco de preços.
_CAMPOS_ORIGEM = ("fornecedor", "fornecedor_canal", "fornecedor_contato",
                  "link_fornecedor", "sku_fornecedor")

def _copiar_origem(destino, prop_item):
    destino.update({
        "preco_custo": float(prop_item.get("preco_custo") or 0),
        "frete_vinda": float(prop_item.get("frete_vinda") or 0),
    })
    for c in _CAMPOS_ORIGEM:
        destino[c] = prop_item.get(c) or ""
    return destino

def _casar_propostas(cnpjs_dig, itens_match, numeros_prop=None):
    """Casa propostas: CNPJ COMPLETO é o principal; a RAIZ (8 díg.) é filtro extra
    de confiança.

    `numeros_prop`: números de proposta lidos do arquivo da PROPOSTA COMERCIAL que
    o operador anexou junto com a PO. Quando ele anexa a proposta, não há o que
    adivinhar — aquela é A proposta. Ela entra ANCORADA, no topo, mesmo que o
    CNPJ da PO venha de filial diferente ou o texto da PO não pareça com nada."""
    sb = get_supabase()
    full = {c for c in cnpjs_dig if len(c) == 14}
    roots = {c[:8] for c in full}
    ancoras = {re.sub(r"\D", "", str(n or "")) for n in (numeros_prop or [])}
    ancoras = {a for a in ancoras if len(a) >= 4}
    if not roots and not ancoras:
        return []
    try:
        props = (sb.table("propostas")
                   .select("id,numero_proposta,cliente,cnpj,data_geracao")
                   .order("data_geracao", desc=True).limit(3000).execute().data) or []
    except Exception:
        props = []
    def _e_ancora(p):
        return bool(ancoras) and re.sub(r"\D", "", str(p.get("numero_proposta") or "")) in ancoras

    casadas = [p for p in props
               if _digitos(p.get("cnpj"))[:8] in roots or _e_ancora(p)]
    if not casadas:
        return []
    ids = [p["id"] for p in casadas]
    try:
        itens = (sb.table("itens_proposta").select("*")
                   .in_("proposta_id", ids).limit(4000).execute().data) or []
    except Exception:
        itens = []
    por_prop = {}
    for it in itens:
        por_prop.setdefault(it["proposta_id"], []).append(it)
    po_toks = [(_toks(i.get("descricao")), float(i.get("preco_unitario") or 0)) for i in (itens_match or [])]
    cands = []
    for p in casadas:
        lista = por_prop.get(p["id"], [])
        score = 0.0
        for pt, ppreco in po_toks:
            melhor = 0.0
            for it in lista:
                itoks = _toks(it.get("descricao_final") or it.get("descricao_original"))
                if not pt or not itoks:
                    continue
                inter = len(pt & itoks); uni = len(pt | itoks)
                sim = (inter / uni) if uni else 0.0
                pv = float(it.get("preco_venda") or 0)
                if ppreco and pv and abs(pv - ppreco) < 0.01:
                    sim += 0.5
                melhor = max(melhor, sim)
            score += melhor
        # CNPJ COMPLETO igual = principal (sobe muito); raiz só (filial difer.) fica abaixo
        cnpj_exato = _digitos(p.get("cnpj")) in full
        if cnpj_exato:
            score += 5.0
        # o operador anexou ESTA proposta: não há o que pontuar, ela é a certa
        ancorada = _e_ancora(p)
        if ancorada:
            score += 1000.0
        cands.append({"proposta": p, "score": round(score, 2), "cnpj_exato": cnpj_exato,
                      "ancorada": ancorada, "itens": lista})
    cands.sort(key=lambda c: (c["score"], c["proposta"].get("data_geracao") or ""), reverse=True)
    return cands[:8]

def _enriquecer_itens_po(itens_po):
    """Pra cada item da PO, busca no histórico (itens_proposta) e puxa custo/link/fornecedor/frete/PN."""
    sb = get_supabase()
    out = []
    for i in (itens_po or []):
        desc = i.get("descricao") or ""
        preco_po = float(i.get("preco_unitario") or 0)
        enr = {"descricao": desc, "quantidade": i.get("quantidade") or 1, "preco_unitario": preco_po,
               "preco_venda": preco_po, "preco_custo": 0.0, "frete_vinda": 0.0,
               "fornecedor": "", "fornecedor_canal": "", "fornecedor_contato": "",
               "link_fornecedor": "", "sku_fornecedor": "", "match_banco": False}
        dtok = _toks(desc)
        chave = max(dtok, key=len, default="")
        chave = re.sub(r"[,()*]", "", chave)
        if len(chave) >= 3:
            try:
                r = (sb.table("itens_proposta")
                       .select("id,descricao_final,descricao_original,preco_venda,preco_custo,frete_vinda,"
                               "fornecedor,fornecedor_canal,fornecedor_contato,link_fornecedor,sku_fornecedor")
                       .or_(f"descricao_final.ilike.*{chave}*,descricao_original.ilike.*{chave}*")
                       .limit(25).execute().data) or []
            except Exception:
                r = []
            # sem proposta casada não há âncora de preço: aqui a busca é no
            # histórico inteiro, então segue a regra conservadora (idêntica ou
            # preço + descrição parecida).
            best, bsim, bigual = None, -1.0, False
            for it in r:
                if not _item_certo(dtok, preco_po, it):
                    continue
                sim, igual = _sim_desc(dtok, it)
                if (igual, sim) > (bigual, bsim):
                    bsim, bigual, best = sim, igual, it
            if best:
                _copiar_origem(enr, best)   # preco_venda fica o da PO (preservado)
                enr["match_banco"] = True
        out.append(enr)
    return out

def _montar_itens_oc(itens_po, prop_itens):
    """OC = itens da PO (descrição/qtd/preço PRESERVADOS 100%). A proposta só
    empresta o dado de COMPRA: custo, frete de vinda, fornecedor, canal, contato,
    link e SKU.

    Pareamento é GLOBAL e 1-para-1, não linha a linha na ordem em que vieram:
    monta todos os pares possíveis, ordena por força da evidência e casa os mais
    fortes primeiro. Assim uma linha fraca não rouba o item de uma linha forte só
    porque apareceu antes na PO — e nenhum item da proposta é herdado duas vezes.

    Quantidade NÃO entra na regra: compra parcial (vendi 10, ele comprou 5) é
    normal e não pode quebrar a herança."""
    itens_po = list(itens_po or [])
    prop_itens = list(prop_itens or [])
    cont_po = _contagem_precos(itens_po, "preco_unitario")
    cont_prop = _contagem_precos(prop_itens, "preco_venda")

    base = []
    for i in itens_po:
        desc = i.get("descricao") or ""
        preco_po = _norm_preco(i.get("preco_unitario"))
        base.append({
            "descricao": desc, "quantidade": i.get("quantidade") or 1,
            "preco_venda": float(i.get("preco_unitario") or 0),   # PREÇO DA PO
            "preco_custo": 0.0, "frete_vinda": 0.0,
            "fornecedor": "", "fornecedor_canal": "", "fornecedor_contato": "",
            "link_fornecedor": "", "sku_fornecedor": "",
            "item_proposta_id": None, "match_proposta": False,
            "certeza": "", "aviso": "",
            "_dtok": _toks(desc), "_preco": preco_po,
        })

    # 1) todos os pares plausíveis, com o peso da evidência
    pares = []
    for pi, oc in enumerate(base):
        preco_po = oc["_preco"]
        # âncora: esse preço identifica UMA linha da PO e UMA linha da proposta
        ancora = (preco_po > 0
                  and cont_po.get(preco_po, 0) == 1
                  and cont_prop.get(preco_po, 0) == 1)
        for ii, it in enumerate(prop_itens):
            if not _item_certo(oc["_dtok"], preco_po, it, preco_ancora=ancora):
                continue
            sim, igual = _sim_desc(oc["_dtok"], it)
            if igual:
                peso, certeza = 3.0, "descricao_identica"
            elif ancora:
                peso, certeza = 2.0, "preco_ancora"
            else:
                peso, certeza = 1.0, "preco_e_descricao"
            pares.append((peso + sim, peso, sim, pi, ii, certeza))

    pares.sort(key=lambda p: (p[0], -p[3], -p[4]), reverse=True)

    # 2) casamento guloso 1-para-1
    po_usada, prop_usado = set(), set()
    for _forca, _peso, _sim, pi, ii, certeza in pares:
        if pi in po_usada or ii in prop_usado:
            continue
        it = prop_itens[ii]
        oc = base[pi]
        _copiar_origem(oc, it)
        oc["item_proposta_id"] = it.get("id")
        oc["match_proposta"] = True
        oc["certeza"] = certeza
        po_usada.add(pi); prop_usado.add(ii)

    # 3) quem não casou: se o preço bate mas é ambíguo, avisa em vez de chutar
    for pi, oc in enumerate(base):
        if pi in po_usada:
            continue
        p = oc["_preco"]
        if p > 0 and cont_prop.get(p, 0) > 1:
            oc["aviso"] = ("preço bate com mais de um item da proposta — "
                           "confirme a origem na mão")
        elif p > 0 and cont_prop.get(p, 0) == 0 and prop_itens:
            oc["aviso"] = "esse preço não existe na proposta casada"

    for oc in base:
        oc.pop("_dtok", None); oc.pop("_preco", None)
    return base

@app.post("/casar-po")
async def casar_po(
    arquivo: UploadFile = File(None),
    proposta_tiny: list[UploadFile] = File(default=[]),
    texto: str = Form(None),
    usuario: str = Depends(verificar_token)
):
    """Recebe a PO do cliente (.msg/.pdf/texto), extrai e casa com propostas salvas.
    proposta_tiny aceita múltiplos arquivos (cliente comprou de 2+ cotações distintas)."""
    import json as _json_po
    import base64 as _b64_po

    conteudo, pdf_docs = "", []
    if arquivo and arquivo.filename:
        conteudo, pdf_docs = await _ler_po(arquivo)
    elif texto:
        conteudo = texto

    if not conteudo.strip() and not pdf_docs:
        return {"erro": "Não consegui ler o conteúdo. Tente o PDF da PO ou cole o texto."}

    cnpjs = re.findall(r"\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}", conteudo)
    cnpjs_dig = list(dict.fromkeys([_digitos(c) for c in cnpjs if len(_digitos(c)) == 14]))
    pos = re.findall(r"PO[-\s]?\d{5,}", conteudo, re.I)
    po_num = pos[0].strip() if pos else ""

    itens_po, destino = [], ""
    avisos_extracao: list[str] = []

    # ── Extração definitiva via Claude Sonnet ─────────────────────────────────
    # Sonnet recebe: PDFs como documentos nativos (escaneados ou textuais)
    # + texto do email/PDFs legíveis. Lê tudo, identifica a PO e extrai itens.
    try:
        msg_content = []

        # PDFs baseados em imagem → documento nativo (Claude lê diretamente)
        for fn, pdf_bytes in pdf_docs:
            msg_content.append({
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/pdf",
                    "data": _b64_po.standard_b64encode(pdf_bytes).decode(),
                },
            })

        # Texto extraído (body email + PDFs legíveis)
        if conteudo.strip():
            msg_content.append({"type": "text", "text": conteudo[:15000]})

        if not msg_content:
            raise ValueError("sem conteúdo")

        claude = get_claude()
        r = claude.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=3000,
            system=SYSTEM_PO,
            messages=[{"role": "user", "content": msg_content}],
            timeout=90.0,
        )
        raw = "".join(b.text for b in r.content if getattr(b, "type", "") == "text").strip()
        # Extração robusta: primeiro { até último }
        raw = re.sub(r'^```(?:json)?\s*', '', raw.strip())
        raw = re.sub(r'\s*```$', '', raw.strip())
        s = raw.find('{'); e = raw.rfind('}')
        if s != -1 and e != -1:
            raw = raw[s:e+1]
        _ia = _json_po.loads(raw)
        itens_po = _ia.get("itens", []) or []
        destino  = _ia.get("destino", "") or ""

        # Extrair CNPJs do texto para enriquecer a busca de candidatas
        for it in itens_po:
            if not isinstance(it.get("quantidade"), (int, float)):
                try: it["quantidade"] = float(str(it["quantidade"]).replace(",","."))
                except: it["quantidade"] = 1
            if not isinstance(it.get("preco_unitario"), (int, float)):
                try: it["preco_unitario"] = float(str(it["preco_unitario"]).replace(",","."))
                except: it["preco_unitario"] = 0

        # Extrair CNPJs também do raw (Sonnet pode incluir na resposta ou no conteúdo lido)
        for cnpj_raw in re.findall(r"\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}", raw):
            d = _digitos(cnpj_raw)
            if len(d) == 14 and d not in cnpjs_dig:
                cnpjs_dig.append(d)
        # Número PO do raw se não encontrado antes
        if not po_num:
            pos2 = re.findall(r"PO[-\s]?\d{5,}", raw, re.I)
            if pos2: po_num = pos2[0].strip()

    except Exception as _ex:
        avisos_extracao.append(f"Extração Sonnet: {_ex}")

    # Propostas do Tiny (opcional, múltiplas) — reforço da busca
    itens_match = list(itens_po)
    numeros_prop: list[str] = []
    for pt_file in (proposta_tiny or []):
        if not (pt_file and pt_file.filename):
            continue
        try:
            import json as _json
            ptxt, _ = await _ler_po(pt_file)
            if not ptxt.strip():
                continue
            claude = get_claude()
            rp = claude.messages.create(
                model="claude-haiku-4-5-20251001", max_tokens=2500,
                system=SYSTEM_PROP_TINY,
                messages=[{"role": "user", "content": ptxt[:8000]}],
            )
            tp = "".join(b.text for b in rp.content if getattr(b, "type", "") == "text").strip().strip("`")
            if tp.lower().startswith("json"):
                tp = tp[4:]
            prop_tiny = _json.loads(tp)
            for it in (prop_tiny.get("itens") or []):
                itens_match.append({"descricao": it.get("descricao"), "preco_unitario": it.get("preco_venda") or 0})
            cnpj_prop = _digitos(prop_tiny.get("cnpj") or "")
            if len(cnpj_prop) == 14 and cnpj_prop not in cnpjs_dig:
                cnpjs_dig.append(cnpj_prop)
            # o número da proposta anexada é a âncora mais forte que existe:
            # o operador está dizendo QUAL proposta essa PO confirma.
            n_prop = re.sub(r"\D", "", str(prop_tiny.get("numero_proposta") or ""))
            if len(n_prop) >= 4 and n_prop not in numeros_prop:
                numeros_prop.append(n_prop)
        except Exception:
            continue

    candidatas = _casar_propostas(cnpjs_dig, itens_match, numeros_prop)
    for c in candidatas:
        c["itens_oc"] = _montar_itens_oc(itens_po, c.get("itens"))

    KIST_CNPJ = "10573732000396"
    cnpj_cliente = next((c for c in cnpjs_dig if len(c) == 14 and c != KIST_CNPJ), "")
    cnpjs_uniq = list(dict.fromkeys(cnpjs))

    return {
        "po_numero": po_num,
        "cnpjs": cnpjs_uniq,
        "cnpj_cliente": cnpj_cliente,
        "destino": destino,
        "itens_po": _enriquecer_itens_po(itens_po),
        "candidatas": candidatas,
        "avisos": avisos_extracao,  # lista de divergências de validação (vazia = tudo OK)
    }

@app.post("/ordens-compra")
async def criar_oc(payload: dict, usuario: str = Depends(verificar_token)):
    """Cria uma nova ordem de compra"""
    sb = get_supabase()
    # imposto padrão (config que o operador pode sobrescrever depois)
    imposto_def = 12.0
    try:
        c = sb.table("config_kist").select("valor").eq("chave", "imposto_percent_default").limit(1).execute()
        if c.data:
            imposto_def = float(c.data[0]["valor"])
    except Exception:
        pass
    # identidade do cliente: CNPJ + razão social + UF (puxa da Receita se houver CNPJ)
    cnpj_in = re.sub(r"\D", "", payload.get("cnpj") or "")
    cnpj_oc = payload.get("cnpj") or None
    cliente_oc = payload.get("cliente") or None
    uf_oc = payload.get("uf") or None
    if _cnpj_valido(cnpj_in):
        cnpj_oc = _cnpj_formatado(cnpj_in)
        try:
            rec = _consulta_receita(cnpj_in)
            if rec.get("razao_social"):
                cliente_oc = rec["razao_social"]
            if not uf_oc and rec.get("uf"):
                uf_oc = rec["uf"]
        except Exception:
            pass
    res = sb.table("ordens_compra").insert({
        "titulo":        payload.get("titulo", ""),
        "numero_po":     (payload.get("numero_po") or None),   # PO do cliente (pode ser nula/pendente)
        "usuario_email": usuario,
        "usuario_nome":  payload.get("usuario_nome", ""),
        "status":        "rascunho",
        "obs":           payload.get("obs", ""),
        "imposto_percent": imposto_def,
        "cnpj":          cnpj_oc,
        "cliente":       cliente_oc,
        "uf":            uf_oc,
    }).execute()
    oc_id = res.data[0]["id"]

    # Adicionar itens se fornecidos
    itens = payload.get("itens", [])
    if itens:
        # ── REDE DE SEGURANÇA DA ORIGEM (v3.23) ───────────────────────────────
        # Quando o item veio de uma proposta casada (`item_proposta_id`), a origem
        # é lida AQUI, do banco, e não depende do frontend ter transportado os
        # campos no payload. Fonte única da verdade: se o dado existe na proposta,
        # ele chega na OC. O que o operador mandou preenchido sempre vence — ele
        # é a hierarquia superior, inclusive contra esta rede.
        origem_por_item = {}
        ids_prop = [i.get("item_proposta_id") for i in itens if i.get("item_proposta_id")]
        if ids_prop:
            try:
                res_ip = (sb.table("itens_proposta")
                            .select("id,preco_custo,frete_vinda,fornecedor,fornecedor_canal,"
                                    "fornecedor_contato,link_fornecedor,sku_fornecedor")
                            .in_("id", ids_prop).execute().data) or []
                origem_por_item = {r["id"]: r for r in res_ip}
            except Exception:
                origem_por_item = {}

        rows = []
        for i in itens:
            base = dict(origem_por_item.get(i.get("item_proposta_id")) or {})

            def _campo(nome, *alternativos):
                """payload primeiro (operador manda), banco como rede."""
                for k in (nome,) + alternativos:
                    v = i.get(k)
                    if v not in (None, "", 0, 0.0):
                        return v
                return base.get(nome) or ""

            custo = float(i.get("preco_custo") or 0) or float(base.get("preco_custo") or 0)
            frete = float(i.get("frete_vinda") or 0) or float(base.get("frete_vinda") or 0)
            rows.append({
                "oc_id":               oc_id,
                "item_proposta_id":    i.get("item_proposta_id"),
                "descricao":           i.get("descricao", ""),
                "quantidade_proposta": float(i.get("quantidade_proposta") or 1),
                "quantidade_comprar":  float(i.get("quantidade_comprar") or i.get("quantidade_proposta") or 1),
                "unidade":             i.get("unidade", "UN"),
                "preco_venda":         float(i.get("preco_venda") or 0),
                "preco_custo":         custo,
                "frete_vinda":         frete,
                # origem do preço herdada da proposta (aceita as duas nomenclaturas):
                "nome_fornecedor":     _campo("fornecedor", "nome_fornecedor"),
                "fornecedor_canal":    _campo("fornecedor_canal"),
                "fornecedor_contato":  _campo("fornecedor_contato"),
                "link_fornecedor":     _campo("link_fornecedor"),
                "sku_fornecedor":      _campo("sku_fornecedor"),
            })
        sb.table("oc_itens").insert(rows).execute()

    return {"oc_id": oc_id}


@app.get("/ordens-compra")
async def listar_ocs(
    status: str = None,
    todos: bool = False,
    limit: int = 100,
    usuario: str = Depends(verificar_token)
):
    """Lista ordens de compra"""
    sb = get_supabase()
    q = sb.table("ordens_compra").select("*").order("criado_em", desc=True).limit(limit)
    if not todos:
        q = q.eq("usuario_email", usuario)
    if status:
        q = q.eq("status", status)
    # Excluir arquivadas por padrão
    q = q.neq("status", "arquivada")
    res = q.execute()
    ocs = res.data or []

    # Totais por OC (venda, custo, fretes, imposto, lucros) para os cards e dashboard
    if ocs:
        ids = [o["id"] for o in ocs]
        itens = sb.table("oc_itens").select(
            "oc_id,preco_venda,preco_custo,frete_vinda,quantidade_comprar,quantidade_proposta"
        ).in_("oc_id", ids).execute().data or []
        tot = {}
        for r in itens:
            qd = r.get("quantidade_comprar")
            if qd is None:
                qd = r.get("quantidade_proposta") or 0
            t = tot.setdefault(r["oc_id"], {"valor_venda": 0.0, "valor_custo": 0.0, "soma_frete_vinda_itens": 0.0})
            t["valor_venda"] += float(r.get("preco_venda") or 0) * float(qd or 0)
            t["valor_custo"] += float(r.get("preco_custo") or 0) * float(qd or 0)
            t["soma_frete_vinda_itens"] += float(r.get("frete_vinda") or 0)
        for o in ocs:
            base = tot.get(o["id"], {"valor_venda": 0.0, "valor_custo": 0.0, "soma_frete_vinda_itens": 0.0})
            o.update({"valor_venda": base["valor_venda"], "valor_custo": base["valor_custo"]})
            # frete de vinda efetivo: soma dos itens se houver, senão o global
            soma_itens = base["soma_frete_vinda_itens"]
            frete_vinda = soma_itens if soma_itens > 0 else float(o.get("frete_vinda_global") or 0)
            frete_ida = float(o.get("frete_ida") or 0)
            cobrado = bool(o.get("frete_ida_cobrado"))
            imposto_pct = float(o.get("imposto_percent") if o.get("imposto_percent") is not None else 12)
            custo_total = base["valor_custo"] + frete_vinda + frete_ida
            nota = base["valor_venda"] + (frete_ida if cobrado else 0.0)
            lucro_bruto = nota - custo_total
            imposto = nota * imposto_pct / 100.0
            lucro_liquido = lucro_bruto - imposto
            o["frete_vinda_efetivo"] = frete_vinda
            o["custo_total"] = custo_total
            o["nota"] = nota
            o["imposto_valor"] = imposto
            o["valor_lucro"] = lucro_bruto          # lucro bruto (R$) — compat
            o["lucro_bruto"] = lucro_bruto
            o["lucro_liquido"] = lucro_liquido
    return ocs


# ══════════════════════════════════════════════════════════════════════════════
# A OC ALIMENTA O BANCO DE PREÇOS (v3.21)
#
# Por que isto existe: o banco só sabia CUSTO quando o operador digitava na tela de
# proposta — e ele digita em 37,7% dos itens. A OC sabe em 89,7%, porque ali ele
# está comprando de verdade: preço, fornecedor, link. Era a melhor fonte de custo
# do sistema, e estava desconectada.
#
# Sem isto, o checkbox de rastreabilidade do Fábio bloqueia 54,4% dos matches e a
# única saída é ele redigitar tudo à mão, cotação após cotação.
#
# GATILHO (definido pelo Leonardo): é por ITEM, não por status da OC.
#   item rastreável + com preço + a OC saiu de rascunho -> alimenta.
# Assim 'confirmada', 'parcialmente_comprada', 'comprada' e 'disponivel' entram
# todas pelo mesmo critério, e item sem lastro se exclui sozinho — inclusive os
# 15 itens com custo e fornecedor nenhum, que são estimativa, não compra.
# ══════════════════════════════════════════════════════════════════════════════

def _rastreavel_oc(item: dict) -> bool:
    """Mesma regra da proposta. Atenção: em oc_itens o campo chama `nome_fornecedor`,
    em produtos chama `fornecedor` — o mesmo dado com dois nomes."""
    if not item:
        return False
    link = (item.get("link_fornecedor") or "").strip()
    nome = (item.get("nome_fornecedor") or "").strip()
    cont = (item.get("fornecedor_contato") or "").strip()
    try:
        custo = float(item.get("preco_custo") or 0) > 0
    except (TypeError, ValueError):
        custo = False
    try:
        venda = float(item.get("preco_venda") or 0) > 0
    except (TypeError, ValueError):
        venda = False
    return bool(link or (nome and cont)) and custo and venda


def _oc_alimenta_banco(sb, oc_id: int, usuario: str = "") -> dict:
    """Leva custo e origem dos itens de uma OC pro banco de preços.

    O QUE GRAVA, e o que deliberadamente NÃO grava:
      produto JÁ EXISTE  -> só custo + origem (fornecedor/canal/contato/link/SKU).
                            NÃO toca em preco_un nem data_ref: a venda já entrou pelo
                            CSV da proposta, e dois escritores no mesmo campo criam
                            ambiguidade sobre quem vence.
      produto NÃO EXISTE -> insere completo. É item que veio da PO e nunca casou com
                            o banco; sem isso ele nunca entraria.
    """
    try:
        oc = sb.table("ordens_compra").select("id,status,cliente,cnpj")\
               .eq("id", oc_id).limit(1).execute()
    except Exception:
        return {"erro": "não consegui ler a OC"}
    if not oc.data:
        return {"erro": "OC não encontrada"}
    if (oc.data[0].get("status") or "") == "rascunho":
        return {"pulou": "OC em rascunho — ainda não é compra"}

    cliente = oc.data[0].get("cliente") or ""
    cnpj    = _cnpj_do_cliente(oc.data[0].get("cnpj"))
    quem    = {"usuario_email": usuario or "",
               "usuario_nome": APELIDOS.get(usuario, "") or (usuario or "").split("@")[0]}
    hoje    = date.today().isoformat()

    try:
        itens = sb.table("oc_itens").select("*").eq("oc_id", oc_id).execute().data or []
    except Exception:
        return {"erro": "não consegui ler os itens"}

    atualizados, inseridos, sem_lastro = 0, 0, 0
    for it in itens:
        desc = (it.get("descricao") or "").strip()
        if not desc:
            continue
        if not _rastreavel_oc(it):
            sem_lastro += 1
            continue

        origem = {"preco_custo": float(it.get("preco_custo") or 0)}
        for campo_oc, campo_prod in (("nome_fornecedor", "fornecedor"),
                                     ("link_fornecedor", "link_fornecedor"),
                                     ("fornecedor_canal", "fornecedor_canal"),
                                     ("fornecedor_contato", "fornecedor_contato"),
                                     ("sku_fornecedor", "sku_fornecedor")):
            if (it.get(campo_oc) or "").strip():
                origem[campo_prod] = it[campo_oc].strip()

        try:
            res = sb.table("produtos").select("id")\
                    .ilike("descricao", _ilike_literal(desc)).limit(1).execute()
            if res and res.data:
                sb.table("produtos").update({**origem, **quem})\
                  .eq("id", res.data[0]["id"]).execute()
                atualizados += 1
            else:
                sb.table("produtos").insert({
                    "descricao": desc, "un": it.get("unidade") or "UN",
                    "preco_un": float(it.get("preco_venda") or 0), "data_ref": hoje,
                    "cliente": cliente, "obs": "inserido via ordem de compra",
                    **origem, **({"cnpj": cnpj} if cnpj else {}), **quem,
                }).execute()
                inseridos += 1
        except Exception as e:
            # Corrida com o /upsert-precos: alguém inseriu a mesma descrição no meio.
            if "23505" in str(e) or "duplicate key" in str(e).lower():
                try:
                    r2 = sb.table("produtos").select("id")\
                           .ilike("descricao", _ilike_literal(desc)).limit(1).execute()
                    if r2.data:
                        sb.table("produtos").update({**origem, **quem})\
                          .eq("id", r2.data[0]["id"]).execute()
                        atualizados += 1
                except Exception:
                    pass

    return {"atualizados": atualizados, "inseridos": inseridos,
            "sem_lastro": sem_lastro, "total": len(itens)}


@app.post("/ordens-compra/{oc_id}/alimentar-banco")
async def oc_alimentar_banco(oc_id: int, usuario: str = Depends(verificar_token)):
    """Dispara à mão. O caminho normal é automático (ao sair de rascunho ou ao
    editar item de OC que já saiu)."""
    return _oc_alimenta_banco(get_supabase(), oc_id, usuario)


@app.post("/ordens-compra/alimentar-banco-lote")
async def oc_alimentar_banco_lote(usuario: str = Depends(verificar_token)):
    """Backfill: as OCs que já existiam antes desta feature nunca alimentaram nada.
    São a maior fonte de custo real do sistema. Admin roda uma vez."""
    if usuario not in ADMIN_EMAILS:
        raise HTTPException(status_code=403, detail="Só o admin.")
    sb = get_supabase()
    ocs = sb.table("ordens_compra").select("id").neq("status", "rascunho").execute().data or []
    tot = {"ocs": 0, "atualizados": 0, "inseridos": 0, "sem_lastro": 0}
    for o in ocs:
        r = _oc_alimenta_banco(sb, o["id"], usuario)
        if "erro" in r or "pulou" in r:
            continue
        tot["ocs"] += 1
        for k in ("atualizados", "inseridos", "sem_lastro"):
            tot[k] += r.get(k, 0)
    return tot


@app.get("/ordens-compra/{oc_id}/itens")
async def itens_oc(oc_id: int, usuario: str = Depends(verificar_token)):
    """Retorna itens de uma OC"""
    sb = get_supabase()
    res = sb.table("oc_itens").select("*").eq("oc_id", oc_id).execute()
    return res.data


@app.put("/ordens-compra/{oc_id}")
async def atualizar_oc(
    oc_id: int, payload: dict, usuario: str = Depends(verificar_token)
):
    """Atualiza status e campos de uma OC"""
    sb = get_supabase()
    campos = {}
    for f in ["titulo","numero_po","status","frete_estimado","frete_real","obs",
              "frete_vinda_global","frete_ida","frete_ida_cobrado","imposto_percent",
              "cnpj","cliente","uf"]:
        if f in payload:
            campos[f] = payload[f]
    # CNPJ editado na mão: grava formatado se for válido (senão, o que veio)
    if "cnpj" in campos and campos["cnpj"]:
        dig = re.sub(r"\D", "", campos["cnpj"])
        if _cnpj_valido(dig):
            campos["cnpj"] = _cnpj_formatado(dig)
    if "uf" in campos and campos["uf"]:
        campos["uf"] = str(campos["uf"]).strip().upper()[:2]

    # Saiu de rascunho? Então virou compra, e o que ela sabe de custo/fornecedor
    # tem que chegar no banco. Lê o status ANTES pra saber se é transição.
    _era_rascunho = False
    if "status" in campos and campos["status"] != "rascunho":
        try:
            _ant = sb.table("ordens_compra").select("status").eq("id", oc_id).limit(1).execute()
            _era_rascunho = bool(_ant.data) and (_ant.data[0].get("status") or "") == "rascunho"
        except Exception:
            _era_rascunho = False

    sb.table("ordens_compra").update(campos).eq("id", oc_id).execute()

    _banco = None
    if _era_rascunho:
        # Falhar aqui não pode derrubar a atualização da OC — o operador mudou o
        # status e isso já é verdade. Alimentar o banco é consequência, não requisito.
        try:
            _banco = _oc_alimenta_banco(sb, oc_id, usuario)
        except Exception:
            _banco = {"erro": "não consegui alimentar o banco agora"}
    # o imposto que o operador definir vira o novo padrão (sobrescreve)
    if "imposto_percent" in payload and payload["imposto_percent"] is not None:
        try:
            sb.table("config_kist").upsert(
                {"chave": "imposto_percent_default", "valor": str(float(payload["imposto_percent"]))},
                on_conflict="chave"
            ).execute()
        except Exception:
            pass
    return {"ok": True, **({"banco": _banco} if _banco else {})}


@app.delete("/ordens-compra/{oc_id}")
async def excluir_oc(oc_id: int, usuario: str = Depends(verificar_token)):
    """Exclui uma OC e todos os seus itens (limpeza de testes / OC errada).
    Apaga os oc_itens primeiro caso a FK não seja ON DELETE CASCADE."""
    sb = get_supabase()
    sb.table("oc_itens").delete().eq("oc_id", oc_id).execute()
    sb.table("ordens_compra").delete().eq("id", oc_id).execute()
    return {"ok": True, "excluida": oc_id}


@app.delete("/propostas/{proposta_id}")
async def excluir_proposta(proposta_id: int, usuario: str = Depends(verificar_token)):
    """Exclui uma proposta e seus itens (limpeza de testes / proposta errada).
    Apaga os itens_proposta primeiro caso a FK não seja ON DELETE CASCADE."""
    sb = get_supabase()
    sb.table("itens_proposta").delete().eq("proposta_id", proposta_id).execute()
    sb.table("propostas").delete().eq("id", proposta_id).execute()
    return {"ok": True, "excluida": proposta_id}


@app.put("/oc-itens/{item_id}")
async def atualizar_item_oc(
    item_id: int, payload: dict, usuario: str = Depends(verificar_token)
):
    """Atualiza campos de um item de OC"""
    sb = get_supabase()
    campos = {}
    for f in ["descricao","unidade","quantidade_comprar","preco_venda","preco_custo",
              "frete_vinda","nome_fornecedor","link_fornecedor","sku_fornecedor",
              "forma_pagamento","numero_parcelas","data_vencimento","final_cartao",
              "status_pagamento","status_item","numero_pedido_fornecedor",
              "prazo_entrega","rastreio","obs"]:
        if f in payload:
            campos[f] = payload[f]
    if not campos:
        return {"ok": True}
    sb.table("oc_itens").update(campos).eq("id", item_id).execute()

    # O operador preenche o custo item a item, DEPOIS da OC sair de rascunho —
    # é nessa hora que ele está comprando. Se esperássemos só a transição de
    # status, o custo real nunca chegaria no banco.
    _banco = None
    if any(c in campos for c in ("preco_custo", "preco_venda", "nome_fornecedor",
                                 "link_fornecedor", "fornecedor_canal", "fornecedor_contato")):
        try:
            r = sb.table("oc_itens").select("oc_id").eq("id", item_id).limit(1).execute()
            if r.data:
                _banco = _oc_alimenta_banco(sb, r.data[0]["oc_id"], usuario)
        except Exception:
            _banco = None
    return {"ok": True, **({"banco": _banco} if _banco else {})}


@app.post("/oc-itens")
async def adicionar_item_oc(payload: dict, usuario: str = Depends(verificar_token)):
    """Adiciona um novo item a uma OC existente"""
    sb = get_supabase()
    oc_id = payload.get("oc_id")
    if not oc_id:
        raise HTTPException(status_code=400, detail="oc_id obrigatorio")
    row = {
        "oc_id": oc_id,
        "descricao": payload.get("descricao") or "",
        "unidade": payload.get("unidade") or "UN",
        "quantidade_proposta": float(payload.get("quantidade") or 1),
        "quantidade_comprar": float(payload.get("quantidade") or 1),
        "preco_venda": float(payload.get("preco_venda") or 0),
        "preco_custo": float(payload.get("preco_custo") or 0),
        "frete_vinda": 0.0,
        "status_item": "pendente",
    }
    res = sb.table("oc_itens").insert(row).execute()
    novo = res.data[0] if res.data else row
    return novo


@app.get("/cartoes")
async def listar_cartoes(usuario: str = Depends(verificar_token)):
    """Cadastro leve de cartões. Preenche-se sozinho conforme as compras."""
    sb = get_supabase()
    res = sb.table("cartoes").select("*").order("final_cartao").execute()
    return res.data or []


@app.post("/cartoes")
async def upsert_cartao(payload: dict, usuario: str = Depends(verificar_token)):
    """Aprende/atualiza um cartão pelo final (4 díg.) + dia de vencimento (1-31).
    Editar o dia atualiza todas as compras daquele cartão, pois o vencimento
    é derivado daqui (não congelado no item)."""
    sb = get_supabase()
    final = (str(payload.get("final_cartao") or "")).strip()[-4:]
    if not final:
        raise HTTPException(status_code=400, detail="final_cartao obrigatório")
    registro = {"final_cartao": final}
    if payload.get("dia_vencimento") is not None:
        try:
            dia = int(payload["dia_vencimento"])
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="dia_vencimento inválido")
        registro["dia_vencimento"] = max(1, min(31, dia))
    if payload.get("apelido") is not None:
        registro["apelido"] = payload["apelido"]
    # upsert pelo final (final_cartao é único)
    sb.table("cartoes").upsert(registro, on_conflict="final_cartao").execute()
    return {"ok": True, **registro}


@app.delete("/oc-itens/{item_id}")
async def remover_item_oc(item_id: int, usuario: str = Depends(verificar_token)):
    """Remove item de uma OC"""
    sb = get_supabase()
    sb.table("oc_itens").delete().eq("id", item_id).execute()
    return {"ok": True}


# Status de OC que ainda representam compra em aberto. 'enviada' NÃO entra:
# quando a OC vai pro cliente a compra está encerrada (decisão do Leonardo, 29/07).
STATUS_OC_CONSOLIDA = ["rascunho", "confirmada", "parcialmente_comprada", "comprada"]


@app.get("/ordens-compra/itens-consolidados")
async def itens_consolidados(
    todos: bool = False,
    ocs: str = "",
    usuario: str = Depends(verificar_token)
):
    """Visão consolidada de itens A COMPRAR, agrupados por descrição.

    `ocs` = ids de OC separados por vírgula (as POs que o operador marcou na tela).
    Vazio = todas as OCs ativas dele (comportamento histórico, nada muda).
    A marcação é sempre uma INTERSEÇÃO com o que ele já veria: marcar OC nunca
    amplia visibilidade — quem manda em 'de quem eu vejo' continua sendo `todos`.
    """
    sb = get_supabase()

    # Buscar OCs ativas (não arquivadas, não disponíveis, não enviadas)
    q = sb.table("ordens_compra").select(
        "id,titulo,numero_po,cliente,status,usuario_email,usuario_nome")
    if not todos:
        q = q.eq("usuario_email", usuario)
    ocs_ativas = q.in_("status", STATUS_OC_CONSOLIDA).execute()

    if not ocs_ativas.data:
        return []

    # Filtro de marcação. Comparação por string: o id é bigint, mas vem da URL
    # como texto — não quero depender de conversão pra não sumir item calado.
    marcadas = {p.strip() for p in (ocs or "").split(",") if p.strip()}
    permitidas = ocs_ativas.data
    if marcadas:
        permitidas = [o for o in permitidas if str(o["id"]) in marcadas]
    if not permitidas:
        return []

    oc_ids = [oc["id"] for oc in permitidas]
    oc_map = {oc["id"]: oc for oc in permitidas}

    # Só o que ainda está PENDENTE — a lista é do que falta comprar.
    itens_res = sb.table("oc_itens").select("*")\
        .in_("oc_id", oc_ids)\
        .in_("status_item", ["pendente"])\
        .execute()

    # Agrupar por descrição
    from collections import defaultdict
    grupos = defaultdict(list)
    for item in (itens_res.data or []):
        key = (item.get("descricao") or "").strip().upper() or "(SEM DESCRIÇÃO)"
        oc = oc_map.get(item["oc_id"], {})
        grupos[key].append({
            **item,
            "oc_titulo": oc.get("titulo", ""),
            "oc_numero_po": oc.get("numero_po", ""),
            "oc_cliente": oc.get("cliente", ""),
            "oc_usuario": oc.get("usuario_nome", ""),
        })

    resultado = []
    for desc, itens in grupos.items():
        total_qty = sum(float(i.get("quantidade_comprar") or 0) for i in itens)
        unidade = itens[0].get("unidade", "UN")

        # ORIGENS distintas do grupo. Mesma regra de rastro do resto do sistema:
        # quem (nome) · por onde (canal) · contato · link · SKU. Nada é inventado —
        # item sem origem aparece como sem origem, não herda a do irmão.
        origens, sem_origem = {}, 0
        for i in itens:
            nome    = (i.get("nome_fornecedor") or "").strip()
            link    = (i.get("link_fornecedor") or "").strip()
            canal   = (i.get("fornecedor_canal") or "").strip()
            contato = (i.get("fornecedor_contato") or "").strip()
            sku     = (i.get("sku_fornecedor") or "").strip()
            if not (nome or link or contato):
                sem_origem += 1
                continue
            chave = (nome.upper(), link, contato)
            o = origens.setdefault(chave, {
                "fornecedor": nome, "canal": canal, "contato": contato,
                "link": link, "sku": sku, "quantidade": 0.0,
            })
            o["quantidade"] += float(i.get("quantidade_comprar") or 0)
            if not o["sku"] and sku:
                o["sku"] = sku
            if not o["canal"] and canal:
                o["canal"] = canal

        # Quebra por PO — de onde vem cada pedaço da quantidade.
        por_oc = {}
        for i in itens:
            k = i["oc_id"]
            p = por_oc.setdefault(k, {
                "oc_id": k, "titulo": i.get("oc_titulo", ""),
                "numero_po": i.get("oc_numero_po", ""),
                "cliente": i.get("oc_cliente", ""),
                "usuario": i.get("oc_usuario", ""), "quantidade": 0.0,
            })
            p["quantidade"] += float(i.get("quantidade_comprar") or 0)

        resultado.append({
            "descricao": desc,
            "unidade": unidade,
            "total_quantidade": total_qty,
            "total_ocs": len(por_oc),      # POs distintas
            "total_itens": len(itens),     # linhas de OC somadas
            "origens": list(origens.values()),
            "sem_origem": sem_origem,
            "ocs": sorted(por_oc.values(), key=lambda x: str(x["numero_po"] or "")),
            "itens": itens,
        })

    resultado.sort(key=lambda x: x["descricao"])
    return resultado


@app.post("/ordens-compra/arquivar-antigas")
async def arquivar_antigas(usuario: str = Depends(verificar_token)):
    """Arquiva OCs com status 'disponivel' há mais de 30 dias"""
    sb = get_supabase()
    from datetime import datetime, timedelta
    limite = (datetime.now() - timedelta(days=30)).isoformat()
    res = sb.table("ordens_compra")\
        .update({"status": "arquivada"})\
        .eq("status", "disponivel")\
        .lt("atualizado_em", limite)\
        .execute()
    return {"arquivadas": len(res.data) if res.data else 0}


# ══════════════════════════════════════════════════════════════════════════════
# ANALISTA DE NEGÓCIOS INTERNO — chamados / sugestões (v3.10)
# ══════════════════════════════════════════════════════════════════════════════
import json as _json_an

ADMIN_EMAILS = set(os.environ.get("ADMIN_EMAILS", "leonardobarrey@gmail.com").split(","))

# Apelido usado pelo Analista na conversa (só no papo — o registro guarda nome/e-mail real).
APELIDOS = {
    "leonardobarrey@gmail.com": "Leonardo",
    "fabiokist@gmail.com": "Cavalo",
    "thiagokist@gmail.com": "José",
}


def _now_iso():
    from datetime import datetime, timezone
    return datetime.now(timezone.utc).isoformat()


FERRAMENTA_FICHA = {
    "name": "propor_chamado",
    "description": (
        "Chame ESTA ferramenta quando tiver informação suficiente para propor o chamado. "
        "Enquanto faltar informação, NÃO chame — apenas responda em texto com a próxima pergunta. "
        "Chamar a ferramenta não registra nada: o chamado só nasce depois que o operador "
        "confirmar na tela."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "tipo":  {"type": "string", "enum": ["bug", "melhoria", "duvida"],
                      "description": "bug = deveria funcionar e não funciona; melhoria = algo novo; duvida = não sabe como fazer"},
            "titulo": {"type": "string", "description": "título curto e claro"},
            "solicitacao": {"type": "string", "description": "o que está sendo pedido, em 1-2 frases"},
            "dor": {"type": "string", "description": "a dor/problema que isso resolve"},
            "comportamento_esperado": {"type": "string", "description": "como o sistema deve se comportar depois"},
            "area": {"type": "string",
                     "enum": ["proposta", "banco", "oc", "receber_po", "extracao", "login", "outro"]},
            "ja_suportado": {"type": "boolean",
                             "description": "true se o sistema JÁ faz isso — explique onde no parecer"},
            "parecer": {"type": "string", "description": "seu parecer técnico curto"},
            "prioridade": {"type": "string", "enum": ["baixa", "media", "alta"]},
        },
        "required": ["tipo", "titulo", "solicitacao", "dor", "comportamento_esperado"],
    },
}


SYSTEM_ANALISTA = """Você é o Analista de Negócios interno do sistema "Kist Cabine de Compras",
usado pelos 3 operadores da Kist. Seu papel: receber sugestões de melhoria, bugs e dúvidas,
entender a fundo o que resolvem, e registrar chamados bem estruturados para o Leonardo (dono/dev) priorizar.

IDIOMA: português do Brasil. Tom: direto, prático, cordial, sem enrolação.

Abaixo deste texto vem o CONHECIMENTO ATUAL do sistema: o que ele já faz (núcleo), as entregas
recentes já no ar, e os chamados já abertos. Use isso como verdade.

SEU MÉTODO, nesta ordem:
1. CLASSIFIQUE: é bug, melhoria ou dúvida? (bug = deveria funcionar e não funciona; melhoria = algo
   novo/diferente; dúvida = a pessoa não sabe como fazer).
2. CHEQUE SE JÁ EXISTE: compare o pedido com o conhecimento do sistema. Se o sistema JÁ FAZ aquilo,
   explique onde e como se faz e trate como "já suportado". Se não tiver certeza, PERGUNTE em vez de
   afirmar. NUNCA invente uma capacidade que não está no conhecimento.
3. CHEQUE DUPLICATA: se já existe chamado aberto igual/parecido, aponte o número (#NN) e avise que já
   está na fila.
4. COLETE O QUE FALTA: no MÁXIMO uma pergunta por vez. Bug: passos pra reproduzir, qual proposta/OC/tela,
   o que esperava vs. o que aconteceu, print se tiver. Melhoria: a dor real, com que frequência acontece,
   como faz hoje (contorno), quem se beneficia.
5. ENTENDA A DOR, não só o pedido. Às vezes o que a pessoa pede não é o que melhor resolve — aponte com cuidado.

QUANDO TIVER INFORMAÇÃO SUFICIENTE, chame a ferramenta `propor_chamado` para montar a FICHA.
NÃO abra sozinho: o chamado só é registrado depois que o operador confirmar na tela.

FORMATO DE RESPOSTA:
- Responda SEMPRE em texto normal, em PT-BR. Escreva à vontade: parágrafos, listas, o que precisar.
  Não existe limite de formato e você NÃO deve responder em JSON.
- Enquanto faltar informação: só texto, com a próxima pergunta. Não chame a ferramenta.
- Quando tiver o suficiente: escreva um resumo curto pedindo pra conferir E chame a ferramenta
  `propor_chamado` na mesma resposta.

REGRAS:
- Se for "já suportado", ainda assim chame a ferramenta com ja_suportado=true e explique no parecer
  onde/como já se faz — registrar isso ajuda a mapear o que está difícil de achar no sistema.
- Nunca prometa prazo nem diga que "vai fazer". Você só registra; quem prioriza e resolve é o Leonardo.
"""


def _estado_real(sb) -> str:
    """Os fatos, medidos AGORA. Não vem do núcleo escrito à mão — vem do banco.

    Por que isto existe: o núcleo é narrativa (o que o sistema SABE FAZER) e
    envelhece só quando há deploy. Já o estado dos dados muda a cada proposta, e
    era escrito à mão dentro do núcleo — ou seja, nascia velho. O erro que isso
    causa é o mais caro do agente: responder "já suportado, o campo existe"
    quando o campo está vazio em 95% dos casos.

    Uma query, ~10ms, custo zero. Nunca mais desatualiza.
    """
    try:
        r = sb.rpc("estado_real_kist", {}).execute()
        linhas = r.data or []
    except Exception:
        return ("\n\n=== ESTADO REAL DOS DADOS ===\n"
                "(não consegui medir agora — NÃO afirme nada sobre volume ou preenchimento "
                "de campos nesta conversa; pergunte ao operador.)")
    if not linhas:
        return ""
    out = ["\n\n=== ESTADO REAL DOS DADOS (medido agora, não é estimativa) ===",
           "Capacidade é o que o sistema SABE FAZER. Isto é o que ele TEM. São coisas",
           "diferentes: o campo pode existir e estar vazio em 95% dos casos.", ""]
    for l in linhas:
        pct = f" ({l['pct']})" if l.get("pct") else ""
        out.append(f"- {l['item']}: {l['valor']}{pct}")
    return "\n".join(out)


def _inventario_tecnico() -> str:
    """Endpoints e tabelas lidos do PRÓPRIO CÓDIGO, por AST. Sem IA, sem alucinação.

    O agente costumava saber isso por uma lista escrita à mão, que ninguém atualizava.
    Aqui o backend lê a si mesmo: o que está no arquivo é o que ele reporta.
    """
    import ast as _ast
    try:
        with open(__file__, encoding="utf-8") as f:
            txt = f.read()
        eps = []
        for no in _ast.walk(_ast.parse(txt)):
            if isinstance(no, (_ast.FunctionDef, _ast.AsyncFunctionDef)):
                for d in no.decorator_list:
                    if (isinstance(d, _ast.Call) and isinstance(d.func, _ast.Attribute)
                            and isinstance(d.func.value, _ast.Name) and d.func.value.id == "app"
                            and d.args and isinstance(d.args[0], _ast.Constant)):
                        eps.append(f"{d.func.attr.upper()} {d.args[0].value}")
        if not eps:
            return ""
        return ("\n\n=== ENDPOINTS QUE EXISTEM DE VERDADE (lidos do código agora) ===\n"
                + "\n".join("- " + e for e in sorted(set(eps), key=lambda x: x.split(" ", 1)[1])))
    except Exception:
        return ""


def _conhecimento_agente(sb) -> str:
    """Monta o conhecimento do agente em 3 camadas: núcleo + entregas + abertos."""
    nucleo = ""
    try:
        r = sb.table("config_kist").select("valor").eq("chave", "capacidades_nucleo").limit(1).execute()
        if r.data:
            nucleo = r.data[0].get("valor") or ""
    except Exception:
        nucleo = ""

    entregue = []
    try:
        r = sb.table("chamados").select("numero,titulo,tipo,resolucao")\
            .in_("status", ["em_producao", "finalizado"]).order("numero", desc=True).limit(200).execute()
        for c in (r.data or []):
            res = (c.get("resolucao") or "").strip()
            if res:
                entregue.append(f"- #{c.get('numero')} [{c.get('tipo','')}] {c.get('titulo','')}: {res}")
    except Exception:
        pass

    abertos = []
    try:
        r = sb.table("chamados").select("numero,titulo,status,solicitacao")\
            .in_("status", ["aberto", "em_desenvolvimento"]).order("numero", desc=True).limit(200).execute()
        for c in (r.data or []):
            abertos.append(f"- #{c.get('numero')} [{c.get('status','')}] {c.get('titulo','')}: {(c.get('solicitacao') or '')[:120]}")
    except Exception:
        pass

    # A narrativa descreve uma versão. Se o código já andou, o agente precisa saber
    # que pode estar falando de um sistema que não existe mais.
    aviso_versao = ""
    m_v = re.search(r"Vers[ãa]o do n[úu]cleo:\s*v?([\d.]+)", nucleo or "")
    if m_v and m_v.group(1) != VERSAO_BACKEND:
        aviso_versao = (
            f"\n\n!!! ATENÇÃO — SEU CONHECIMENTO PODE ESTAR ATRASADO !!!\n"
            f"A narrativa abaixo descreve a v{m_v.group(1)}. O código em produção está na "
            f"v{VERSAO_BACKEND}. Houve mudanças que ninguém escreveu aqui.\n"
            f"REGRA: não afirme que algo NÃO existe. Se o operador descrever um comportamento "
            f"que a narrativa não menciona, ACREDITE NELE e pergunte — não corrija. Confira a "
            f"lista de endpoints reais abaixo antes de dizer que uma capacidade não existe.")

    partes = ["=== O QUE O SISTEMA JÁ FAZ (núcleo) ===", nucleo or "(indisponível)", aviso_versao,
              _inventario_tecnico(), _estado_real(sb)]
    if entregue:
        partes += ["", "=== ENTREGAS RECENTES (já no ar — considere suportado) ===", "\n".join(entregue)]
    if abertos:
        partes += ["", "=== CHAMADOS JÁ ABERTOS (para detectar duplicata) ===", "\n".join(abertos)]
    return "\n".join(partes)


class AnalistaChatIn(BaseModel):
    mensagens: list
    operador_nome: Optional[str] = ""
    sessao_id: Optional[str] = ""


@app.post("/analista/chat")
async def analista_chat(payload: AnalistaChatIn, usuario: str = Depends(verificar_token)):
    sb = get_supabase()
    claude = get_claude()
    apelido = APELIDOS.get(usuario, "")
    saudacao = (f"\n\nVocê está conversando agora com {apelido}. Trate essa pessoa por esse nome/apelido, "
                "de forma cordial e natural — sem exagerar, sem repetir a cada frase.") if apelido else ""
    # Anexos: a leitura já foi feita no upload; aqui entra só o texto (barato).
    lim = _anexo_limites(sb)
    system = (SYSTEM_ANALISTA + saudacao + "\n\n" + _regras_anexos(sb)
              + "\n\n" + _conhecimento_agente(sb)
              + _anexos_contexto(sb, payload.sessao_id or "", usuario, lim))

    msgs = []
    for m in (payload.mensagens or [])[-40:]:
        role = m.get("role")
        content = m.get("content", "")
        if role in ("user", "assistant") and isinstance(content, str) and content.strip():
            msgs.append({"role": role, "content": content})
    if not msgs or msgs[0]["role"] != "user":
        return {"reply": "Oi! Sou o analista do sistema. Me conta a melhoria que você quer sugerir ou o problema que encontrou.", "ficha": None}

    try:
        resp = claude.messages.create(
            model="claude-sonnet-4-6", max_tokens=4000,
            system=system, messages=msgs, temperature=0.2, timeout=90.0,
            tools=[FERRAMENTA_FICHA],
        )
        # A prosa vem em bloco(s) de texto; a ficha vem estruturada pela ferramenta.
        # Nada de json.loads em cima da fala do modelo — era daí que vinha o
        # "Tive um problema pra processar": a resposta longa estourava max_tokens,
        # o JSON truncava no meio e o parse quebrava JUSTO quando a ficha ficava
        # pronta (a hora em que a resposta é maior).
        reply = "".join(b.text for b in resp.content
                        if getattr(b, "type", "") == "text").strip()
        ficha = None
        for b in resp.content:
            if getattr(b, "type", "") == "tool_use" and getattr(b, "name", "") == "propor_chamado":
                ficha = dict(b.input or {})
        if ficha and not reply:
            reply = "Montei a ficha com o que você me passou. Confere aí e me confirma."
        return {"reply": reply or "…", "ficha": ficha}
    except Exception as e:
        # Falha do sistema: o operador não pode achar que foi ele que escreveu errado.
        num = _abrir_chamado_automatico(
            sb, usuario, f"analista:chat:{type(e).__name__}",
            titulo="Analista de Negócios falhou ao responder",
            solicitacao=("O operador estava conversando com o analista e a resposta não veio. "
                         "A conversa fica travada: ele repete a mensagem e toma o mesmo erro."),
            dor=("O operador não consegue registrar o chamado. Ele reescreve tudo, toma o erro de "
                 "novo e desiste — e a demanda se perde."),
            esperado="O analista responde. Se falhar, o operador vê o motivo real, não uma desculpa.",
            detalhe=f"{type(e).__name__}: {e}"[:400],
            area="outro",
        )
        aviso = (f" Registrei o chamado #{str(num).zfill(4)} e o Leonardo foi avisado."
                 if num else " Não consegui nem registrar o chamado — avise o Leonardo.")
        return {
            "reply": "Falhei aqui do meu lado — não foi nada que você escreveu, e o texto não se "
                     "perdeu." + aviso + " Pode tentar enviar de novo.",
            "ficha": None,
            "erro": f"{type(e).__name__}: {e}"[:400],
            "chamado": num,
        }


class ChamadoIn(BaseModel):
    tipo: Optional[str] = "melhoria"
    titulo: Optional[str] = ""
    descricao_operador: Optional[str] = ""
    solicitacao: Optional[str] = ""
    dor: Optional[str] = ""
    comportamento_esperado: Optional[str] = ""
    resumo_analista: Optional[str] = ""
    parecer_analista: Optional[str] = ""
    area: Optional[str] = ""
    ja_suportado: Optional[bool] = False
    prioridade: Optional[str] = "media"
    transcript: Optional[list] = None
    operador_nome: Optional[str] = ""
    sessao_id: Optional[str] = ""


@app.post("/chamados")
async def criar_chamado(payload: ChamadoIn, usuario: str = Depends(verificar_token)):
    sb = get_supabase()
    status = "ja_suportada" if payload.ja_suportado else "aberto"
    row = {
        "operador_email": usuario,
        "operador_nome": payload.operador_nome or "",
        "tipo": payload.tipo or "melhoria",
        "titulo": payload.titulo or "",
        "descricao_operador": payload.descricao_operador or "",
        "solicitacao": payload.solicitacao or "",
        "dor": payload.dor or "",
        "comportamento_esperado": payload.comportamento_esperado or "",
        "resumo_analista": payload.resumo_analista or "",
        "parecer_analista": payload.parecer_analista or "",
        "area": payload.area or "",
        "ja_suportado": bool(payload.ja_suportado),
        "prioridade": payload.prioridade or "media",
        "status": status,
        "transcript": payload.transcript or [],
    }
    if payload.ja_suportado:
        row["resolvido_em"] = _now_iso()
    try:
        r = sb.table("chamados").insert(row).execute()
        novo = (r.data or [{}])[0]
        # Amarra os anexos da conversa ao chamado recém-criado.
        # (Nasceram na sessao_id, antes do chamado existir.)
        anexos = 0
        if (payload.sessao_id or "").strip() and novo.get("id"):
            try:
                ra = sb.table("chamado_anexos").update({"chamado_id": novo["id"]})\
                    .eq("sessao_id", payload.sessao_id).eq("operador_email", usuario)\
                    .is_("chamado_id", "null").execute()
                anexos = len(ra.data or [])
            except Exception:
                anexos = 0
        return {"ok": True, "numero": novo.get("numero"), "id": novo.get("id"),
                "status": status, "anexos": anexos}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao registrar chamado: {str(e)}")


@app.get("/chamados")
async def listar_chamados(status: Optional[str] = None, arquivados: bool = False,
                          usuario: str = Depends(verificar_token)):
    sb = get_supabase()
    is_admin = usuario in ADMIN_EMAILS
    q = sb.table("chamados").select("*")
    if is_admin:
        if not arquivados:
            q = q.eq("arquivado", False)
        if status:
            q = q.eq("status", status)
    else:
        q = q.eq("operador_email", usuario)
    r = q.order("numero", desc=True).execute()
    return {"chamados": r.data or [], "is_admin": is_admin}


class ChamadoUpdate(BaseModel):
    status: Optional[str] = None
    resolucao: Optional[str] = None
    prioridade: Optional[str] = None
    area: Optional[str] = None
    duplicada_de: Optional[int] = None
    arquivado: Optional[bool] = None
    titulo: Optional[str] = None


@app.put("/chamados/{chamado_id}")
async def atualizar_chamado(chamado_id: int, payload: ChamadoUpdate,
                            usuario: str = Depends(verificar_token)):
    if usuario not in ADMIN_EMAILS:
        raise HTTPException(status_code=403, detail="Só o admin pode gerenciar chamados.")
    sb = get_supabase()
    upd = {"atualizado_em": _now_iso()}
    for campo in ("status", "resolucao", "prioridade", "area", "duplicada_de", "arquivado", "titulo"):
        v = getattr(payload, campo)
        if v is not None:
            upd[campo] = v
    if payload.status in ("em_producao", "finalizado"):
        upd["avisar_operador"] = True
        upd["resolvido_em"] = _now_iso()
    try:
        r = sb.table("chamados").update(upd).eq("id", chamado_id).execute()
        return {"ok": True, "chamado": (r.data or [{}])[0]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/chamados/{chamado_id}/visto")
async def marcar_chamado_visto(chamado_id: int, usuario: str = Depends(verificar_token)):
    """O operador zera o aviso 'no ar' dos próprios chamados quando os vê."""
    sb = get_supabase()
    sb.table("chamados").update({"avisar_operador": False})\
        .eq("id", chamado_id).eq("operador_email", usuario).execute()
    return {"ok": True}


@app.post("/chamados/arquivar-concluidos")
async def arquivar_concluidos(usuario: str = Depends(verificar_token)):
    """Limpa do kanban os chamados em produção/finalizados (ficam na base para consulta)."""
    if usuario not in ADMIN_EMAILS:
        raise HTTPException(status_code=403, detail="Só o admin.")
    sb = get_supabase()
    r = sb.table("chamados").update({"arquivado": True})\
        .in_("status", ["em_producao", "finalizado"]).eq("arquivado", False).execute()
    return {"ok": True, "arquivados": len(r.data or [])}


# ═══════════════════════════════════════════════════════════════════════════
# ANEXOS DE CHAMADOS (v3.12)
#
# Desenho (importante pra quem for mexer depois):
#   - O arquivo NÃO trafega no /analista/chat. O front sobe o arquivo pro
#     Supabase Storage (bucket privado 'chamados'), o backend lê UMA vez e
#     guarda o resumo da leitura em chamado_anexos.leitura.
#   - Esse resumo (texto barato) é injetado no contexto do agente a cada turno.
#     Sem isso, cada mensagem custaria uma chamada de visão.
#   - O anexo nasce numa sessao_id (antes do chamado existir). Ao confirmar a
#     ficha, /chamados amarra sessao_id -> chamado_id.
#   - Guardar é barato e não tem teto de quantidade. LER é o que custa: os
#     tetos de leitura estão em config_kist['anexos_limites'] (runtime).
# ═══════════════════════════════════════════════════════════════════════════
import uuid as _uuid_anx

ANEXOS_BUCKET = os.environ.get("ANEXOS_BUCKET", "chamados")

_ANEXO_LIMITES_PADRAO = {
    "max_mb": 50,             # teto duro de upload (bate com o bucket)
    "leitura_max_mb": 25,     # acima disso: guarda, não lê
    "pdf_paginas": 40,        # páginas lidas por PDF
    "imagens_sessao": 10,     # imagens lidas por conversa
    "chars_por_anexo": 12000,
    "chars_total": 60000,     # teto do bloco de anexos no contexto do agente
    "zip_max_entradas": 60,
    "zip_expansao_max_mb": 500,
}

_EXT_IMG = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
_EXT_TXT = {".txt", ".md", ".log", ".json", ".csv", ".tsv"}
_EXT_XLS = {".xlsx", ".xls", ".xlsm"}
_ANEXO_EXT_OK = _EXT_IMG | _EXT_TXT | _EXT_XLS | {".pdf", ".msg", ".eml", ".docx", ".zip"}


def _anexo_limites(sb) -> dict:
    lim = dict(_ANEXO_LIMITES_PADRAO)
    try:
        r = sb.table("config_kist").select("valor").eq("chave", "anexos_limites").limit(1).execute()
        if r.data:
            v = _json_an.loads(r.data[0].get("valor") or "{}")
            if isinstance(v, dict):
                for k in lim:
                    if k in v:
                        lim[k] = int(v[k])
    except Exception:
        pass
    return lim


def _ext_de(nome: str) -> str:
    n = (nome or "").lower().strip()
    p = n.rfind(".")
    return n[p:] if p > 0 else ""


def _nome_seguro(nome: str) -> str:
    """Nunca confiar no nome que vem do navegador nem do zip: tira caminho,
    tira '..', deixa só caractere previsível."""
    base = re.split(r"[\\/]", (nome or "").strip())[-1]
    base = base.replace("..", "_")
    base = re.sub(r"[^A-Za-z0-9._\- ]+", "_", base).strip() or "arquivo"
    return base[:120]


# ── Storage (defensivo: nomes de chave mudam entre versões do storage3) ────
def _anx_storage():
    return get_supabase().storage.from_(ANEXOS_BUCKET)


def _anx_url_absoluta(u: str) -> str:
    if not u:
        return ""
    if u.startswith("http"):
        return u
    base = SUPABASE_URL.rstrip("/")
    if u.startswith("/storage/v1"):
        return base + u
    return base + "/storage/v1" + (u if u.startswith("/") else "/" + u)


def _anx_signed_upload(path: str):
    try:
        r = _anx_storage().create_signed_upload_url(path) or {}
        if not isinstance(r, dict):
            r = getattr(r, "__dict__", {}) or {}
        u = r.get("signed_url") or r.get("signedURL") or r.get("signedUrl") or ""
        return _anx_url_absoluta(u) or None
    except Exception:
        return None


def _anx_signed_download(path: str, expira: int = 3600):
    try:
        r = _anx_storage().create_signed_url(path, expira) or {}
        if not isinstance(r, dict):
            r = getattr(r, "__dict__", {}) or {}
        u = r.get("signedURL") or r.get("signedUrl") or r.get("signed_url") or ""
        return _anx_url_absoluta(u) or None
    except Exception:
        return None


def _anx_download(path: str) -> bytes:
    return _anx_storage().download(path)


def _anx_remove(path: str):
    try:
        _anx_storage().remove([path])
    except Exception:
        pass


# ── Leitura ───────────────────────────────────────────────────────────────
SYSTEM_LEITURA_ANEXO = """Você lê um anexo de um chamado de suporte interno (sistema "Kist Cabine de Compras").
Descreva OBJETIVAMENTE, em PT-BR, o que o arquivo mostra — para que um analista entenda o problema sem abrir o arquivo.

- Se for print de erro: TRANSCREVA a mensagem de erro literalmente e diga em que tela/contexto aparece.
- Se for print de tela: diga qual tela é, o que está preenchido e o que parece errado.
- Se for documento/planilha: diga o que é, de quem, e liste os dados que importam (itens, quantidades, valores, CNPJ, números).
- Se estiver ilegível, cortado ou vazio: DIGA ISSO claramente e diga o que faltou aparecer.
- Não invente. Não opine sobre a solução. Não use markdown. Máximo ~12 linhas."""


def _leitura_ia_imagem(claude, data: bytes, nome: str) -> str:
    resp = claude.messages.create(
        model="claude-sonnet-4-6", max_tokens=700, system=SYSTEM_LEITURA_ANEXO,
        temperature=0, timeout=45.0,
        messages=[{"role": "user", "content": [
            {"type": "text", "text": f"Arquivo: {nome}"},
            {"type": "image", "source": {"type": "base64",
                "media_type": _media_type_img(data),
                "data": _b64.standard_b64encode(data).decode()}},
        ]}],
    )
    return "".join(b.text for b in resp.content if getattr(b, "type", "") == "text").strip()


def _leitura_ia_pdf(claude, data: bytes, nome: str) -> str:
    """Só para PDF sem texto extraível (digitalizado)."""
    resp = claude.messages.create(
        model="claude-sonnet-4-6", max_tokens=900, system=SYSTEM_LEITURA_ANEXO,
        temperature=0, timeout=90.0,
        messages=[{"role": "user", "content": [
            {"type": "document", "source": {"type": "base64",
                "media_type": "application/pdf",
                "data": _b64.standard_b64encode(data).decode()}},
            {"type": "text", "text": f"Arquivo: {nome}"},
        ]}],
    )
    return "".join(b.text for b in resp.content if getattr(b, "type", "") == "text").strip()


def _anexo_pdf_texto(data: bytes, max_pag: int):
    """Extrai texto do PDF parando em max_pag. Devolve (texto, lidas, total)."""
    try:
        import pdfplumber, io as _io
        partes, lidas, total = [], 0, 0
        with pdfplumber.open(_io.BytesIO(data)) as pdf:
            total = len(pdf.pages)
            for page in pdf.pages[: max(0, int(max_pag))]:
                lidas += 1
                t = (page.extract_text() or "").strip()
                if t:
                    partes.append(t)
        return "\n\n".join(partes), lidas, total
    except Exception:
        return "", 0, 0


def _anexo_excel_texto(data: bytes) -> str:
    """Leitor de Excel próprio do bloco de anexos.
    (O /extrair tem o dele, como closure — não dá pra reusar e não vamos tocar lá.)"""
    try:
        import openpyxl, io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
        partes = []
        for sname in wb.sheetnames[:6]:
            ws = wb[sname]
            linhas = []
            for row in ws.iter_rows(values_only=True):
                vals = [str(c).strip() if c is not None else "" for c in row]
                if any(v and v != "None" for v in vals):
                    linhas.append(" | ".join(vals))
                if len(linhas) >= 200:
                    break
            if linhas:
                partes.append(f"[ABA: {sname}]\n" + "\n".join(linhas))
        return "\n\n".join(partes)
    except Exception:
        return ""


def _anexo_docx_texto(data: bytes) -> str:
    """docx é um zip com word/document.xml — dá pra ler sem dependência nova."""
    try:
        import zipfile, io as _io
        with zipfile.ZipFile(_io.BytesIO(data)) as z:
            xml = z.read("word/document.xml").decode("utf-8", "ignore")
        xml = re.sub(r"</w:p>", "\n", xml)
        xml = re.sub(r"<[^>]+>", "", xml)
        return re.sub(r"\n{3,}", "\n\n", xml).strip()
    except Exception:
        return ""


def _anexo_eml_texto(data: bytes) -> str:
    try:
        import email
        from email import policy
        m = email.message_from_bytes(data, policy=policy.default)
        corpo = ""
        try:
            p = m.get_body(preferencelist=("plain", "html"))
            corpo = p.get_content() if p else ""
        except Exception:
            corpo = ""
        if "<" in corpo and ">" in corpo:
            corpo = re.sub(r"<[^>]+>", " ", corpo)
        anexos = [a.get_filename() for a in m.iter_attachments() if a.get_filename()]
        cab = f"De: {m.get('From','')}\nPara: {m.get('To','')}\nAssunto: {m.get('Subject','')}\n"
        if anexos:
            cab += f"Anexos do e-mail: {', '.join(anexos)}\n"
        return (cab + "\nCorpo:\n" + (corpo or "").strip()).strip()
    except Exception:
        return ""


def _ler_anexo(claude, data: bytes, nome: str, orc: dict, lim: dict, prof: int = 0) -> dict:
    """Lê UM anexo. Devolve {'texto','status','paginas','img_lidas'}.

    status: lido | parcial | nao_lido | erro
    prof=0 → pode abrir container (.zip/.msg). prof>=1 → não abre (guarda profundidade 1).
    orc é mutável: {'img': n} — o teto de imagens vale pra conversa inteira.
    """
    ext = _ext_de(nome)
    teto_char = int(lim["chars_por_anexo"])

    def _corta(t, status="lido"):
        t = (t or "").strip()
        if not t:
            return {"texto": "", "status": "nao_lido", "paginas": 0, "img_lidas": 0}
        if len(t) > teto_char:
            return {"texto": t[:teto_char] + "\n[...conteúdo cortado no teto de leitura...]",
                    "status": "parcial", "paginas": 0, "img_lidas": 0}
        return {"texto": t, "status": status, "paginas": 0, "img_lidas": 0}

    try:
        if len(data) > int(lim["leitura_max_mb"]) * 1024 * 1024:
            mb = len(data) // (1024 * 1024)
            return {"texto": f"[arquivo de ~{mb} MB — guardado e disponível para download, "
                             f"mas grande demais para leitura automática]",
                    "status": "nao_lido", "paginas": 0, "img_lidas": 0}

        # ── Imagem ────────────────────────────────────────────────────────
        if ext in _EXT_IMG:
            if orc.get("img", 0) <= 0:
                return {"texto": "[imagem guardada — teto de imagens lidas nesta conversa já foi atingido]",
                        "status": "parcial", "paginas": 0, "img_lidas": 0}
            orc["img"] -= 1
            t = _leitura_ia_imagem(claude, data, nome)
            r = _corta(t)
            r["img_lidas"] = 1
            return r

        # ── PDF ───────────────────────────────────────────────────────────
        if ext == ".pdf":
            if data[:5] != b"%PDF-":
                return {"texto": "[arquivo com extensão .pdf mas conteúdo inválido — guardado, não lido]",
                        "status": "erro", "paginas": 0, "img_lidas": 0}
            max_pag = int(lim["pdf_paginas"])
            txt, lidas, total = _anexo_pdf_texto(data, max_pag)
            if len((txt or "").strip()) >= 40:
                r = _corta(txt)
                r["paginas"] = lidas
                if total > lidas:
                    r["texto"] += f"\n[PDF tem {total} páginas; li as {lidas} primeiras.]"
                    r["status"] = "parcial"
                return r
            # Sem texto = digitalizado. Manda nativo pro Sonnet se couber.
            # total==0 significa que o pdfplumber NÃO abriu o arquivo: é PDF
            # quebrado, não digitalizado. Mandar pro Sonnet daria 400 e custo à toa.
            if total == 0:
                return {"texto": "[PDF ilegível — não foi possível abrir o arquivo. Guardado para download]",
                        "status": "erro", "paginas": 0, "img_lidas": 0}
            if len(data) <= 10 * 1024 * 1024 and total <= 100:
                t = _leitura_ia_pdf(claude, data, nome)
                r = _corta(t)
                r["paginas"] = total
                return r
            return {"texto": f"[PDF digitalizado com {total} páginas — guardado, mas grande demais "
                             f"para leitura automática de imagem]",
                    "status": "nao_lido", "paginas": 0, "img_lidas": 0}

        # ── Excel / texto / docx / eml ─────────────────────────────────────
        if ext in _EXT_XLS:
            return _corta(_anexo_excel_texto(data))
        if ext in _EXT_TXT:
            return _corta(data.decode("utf-8", "ignore"))
        if ext == ".docx":
            return _corta(_anexo_docx_texto(data))
        if ext == ".eml":
            return _corta(_anexo_eml_texto(data))

        # ── .msg (e-mail do Outlook) ──────────────────────────────────────
        if ext == ".msg":
            if prof >= 1:
                return {"texto": "[e-mail dentro de container — não aberto]", "status": "nao_lido",
                        "paginas": 0, "img_lidas": 0}
            # INVARIANTE: copiar os bytes pra /tmp antes do extract_msg
            # (o mount read-only dá I/O error em acesso aleatório).
            tmp = f"/tmp/anx_{_uuid_anx.uuid4().hex[:8]}.msg"
            with open(tmp, "wb") as _f:
                _f.write(data)
            partes, pag, imgs = [], 0, 0
            try:
                m = extract_msg.openMsg(tmp)
                partes.append(f"Assunto: {m.subject}\nDe: {m.sender}\n\nCorpo:\n{(m.body or '').strip()}")
                for att in (m.attachments or []):
                    afn = _nome_seguro(att.longFilename or att.shortFilename or "anexo")
                    if not att.data:
                        continue
                    sub = _ler_anexo(claude, att.data, afn, orc, lim, prof=1)
                    pag += sub.get("paginas", 0)
                    imgs += sub.get("img_lidas", 0)
                    partes.append(f"--- anexo do e-mail: {afn} ({sub['status']}) ---\n{sub['texto']}")
            finally:
                try:
                    os.remove(tmp)
                except Exception:
                    pass
            r = _corta("\n\n".join(partes))
            r["paginas"] = pag
            r["img_lidas"] = imgs
            return r

        # ── .zip ──────────────────────────────────────────────────────────
        if ext == ".zip":
            if prof >= 1:
                return {"texto": "[zip dentro de zip — não aberto por segurança]",
                        "status": "parcial", "paginas": 0, "img_lidas": 0}
            import zipfile, io as _io
            teto_exp = int(lim["zip_expansao_max_mb"]) * 1024 * 1024
            max_ent = int(lim["zip_max_entradas"])
            partes, pag, imgs, expandido, avisos = [], 0, 0, 0, []
            with zipfile.ZipFile(_io.BytesIO(data)) as z:
                infos = [i for i in z.infolist() if not i.is_dir()]
                # Zip bomb: aborta antes de descompactar
                declarado = sum(i.file_size for i in infos)
                if declarado > teto_exp or (len(data) and declarado > len(data) * 10 and declarado > 50 * 1024 * 1024):
                    return {"texto": f"[zip recusado na leitura: descompactado daria ~{declarado // (1024*1024)} MB. "
                                     f"O arquivo está guardado e pode ser baixado.]",
                            "status": "nao_lido", "paginas": 0, "img_lidas": 0}
                lista = [_nome_seguro(i.filename) for i in infos]
                partes.append(f"[ZIP com {len(infos)} arquivo(s): {', '.join(lista[:40])}"
                              + (" …" if len(lista) > 40 else "") + "]")
                for i in infos[:max_ent]:
                    fn = _nome_seguro(i.filename)
                    if _ext_de(fn) not in _ANEXO_EXT_OK:
                        avisos.append(f"{fn} (tipo não lido)")
                        continue
                    if expandido + i.file_size > teto_exp:
                        avisos.append(f"{fn} (teto de expansão)")
                        continue
                    try:
                        sub_bytes = z.read(i)
                    except Exception:
                        avisos.append(f"{fn} (falha ao extrair)")
                        continue
                    expandido += len(sub_bytes)
                    sub = _ler_anexo(claude, sub_bytes, fn, orc, lim, prof=1)
                    pag += sub.get("paginas", 0)
                    imgs += sub.get("img_lidas", 0)
                    partes.append(f"--- dentro do zip: {fn} ({sub['status']}) ---\n{sub['texto']}")
                if len(infos) > max_ent:
                    avisos.append(f"+{len(infos) - max_ent} arquivo(s) além do teto de entradas")
            if avisos:
                partes.append("[não lidos: " + "; ".join(avisos) + "]")
            r = _corta("\n\n".join(partes))
            r["paginas"] = pag
            r["img_lidas"] = imgs
            if avisos and r["status"] == "lido":
                r["status"] = "parcial"
            return r

        return {"texto": "", "status": "nao_lido", "paginas": 0, "img_lidas": 0}

    except Exception as e:
        return {"texto": f"[falha ao ler: {type(e).__name__}: {e}]", "status": "erro",
                "paginas": 0, "img_lidas": 0}


# ── Contexto do agente ────────────────────────────────────────────────────
_REGRAS_ANEXOS_PADRAO = """REGRAS SOBRE ANEXOS:
- Os anexos desta conversa já foram lidos pelo sistema; o bloco ANEXOS DESTA CONVERSA é a leitura real do conteúdo. Trate como verdade e cite o que viu.
- AVALIE SUFICIÊNCIA: se a leitura não fecha a história (print cortado, PDF sem a página que importa, planilha sem os itens, anexo nao_lido/parcial/erro), diga o que faltou e peça o arquivo/trecho específico. Enquanto faltar, "ficha" continua null.
- Não peça anexo quando o texto já basta. Anexo é evidência, não burocracia.
- Nunca invente conteúdo de anexo que não está na leitura abaixo."""


def _regras_anexos(sb) -> str:
    try:
        r = sb.table("config_kist").select("valor").eq("chave", "analista_regras_anexos").limit(1).execute()
        if r.data and (r.data[0].get("valor") or "").strip():
            return r.data[0]["valor"]
    except Exception:
        pass
    return _REGRAS_ANEXOS_PADRAO


def _anexos_contexto(sb, sessao_id: str, usuario: str, lim: dict) -> str:
    """Bloco de texto com a leitura dos anexos da sessão. Barato: já está lido."""
    if not sessao_id:
        return ""
    try:
        r = sb.table("chamado_anexos").select("nome,status,leitura")\
            .eq("sessao_id", sessao_id).eq("operador_email", usuario)\
            .order("id").execute()
    except Exception:
        return ""
    linhas, usado, teto = [], 0, int(lim["chars_total"])
    for a in (r.data or []):
        if a.get("status") == "enviando":
            continue
        txt = (a.get("leitura") or "").strip() or "(sem conteúdo legível)"
        bloco = f"[anexo: {a.get('nome')} · status={a.get('status')}]\n{txt}"
        if usado + len(bloco) > teto:
            linhas.append(f"[anexo: {a.get('nome')} · status={a.get('status')} · leitura omitida: "
                          f"teto de contexto atingido]")
            usado = teto
            continue
        linhas.append(bloco)
        usado += len(bloco)
    if not linhas:
        return ""
    return "\n\n=== ANEXOS DESTA CONVERSA (leitura real do conteúdo) ===\n" + "\n\n".join(linhas)


def _orcamento_sessao(sb, sessao_id: str, lim: dict) -> dict:
    """Quanto do teto de leitura a conversa já consumiu."""
    usadas = 0
    try:
        r = sb.table("chamado_anexos").select("img_lidas").eq("sessao_id", sessao_id).execute()
        usadas = sum(int(a.get("img_lidas") or 0) for a in (r.data or []))
    except Exception:
        usadas = 0
    return {"img": max(0, int(lim["imagens_sessao"]) - usadas)}


# ── Endpoints ─────────────────────────────────────────────────────────────
class AnexoAssinarIn(BaseModel):
    sessao_id: str
    nome: str
    tamanho: Optional[int] = 0
    mime: Optional[str] = ""


@app.post("/chamados/anexos/assinar")
async def anexo_assinar(payload: AnexoAssinarIn, usuario: str = Depends(verificar_token)):
    """Reserva a vaga do anexo e devolve como subir.
    modo='signed' → o navegador sobe DIRETO pro Storage (o backend não segura os
    bytes — o Render free não aguenta 50 MB atravessando a instância).
    modo='proxy'  → fallback: sobe via backend."""
    sb = get_supabase()
    lim = _anexo_limites(sb)
    nome = _nome_seguro(payload.nome)
    ext = _ext_de(nome)
    if ext not in _ANEXO_EXT_OK:
        raise HTTPException(status_code=400, detail=f"Tipo de arquivo não aceito ({ext or 'sem extensão'}).")
    tam = int(payload.tamanho or 0)
    if tam > int(lim["max_mb"]) * 1024 * 1024:
        raise HTTPException(status_code=400, detail=f"Arquivo maior que {lim['max_mb']} MB.")
    if not (payload.sessao_id or "").strip():
        raise HTTPException(status_code=400, detail="sessao_id ausente.")

    path = f"sessao/{payload.sessao_id}/{_uuid_anx.uuid4().hex[:10]}-{nome}"
    try:
        r = sb.table("chamado_anexos").insert({
            "sessao_id": payload.sessao_id, "operador_email": usuario, "nome": nome,
            "path": path, "mime": payload.mime or "", "tamanho": tam, "status": "enviando",
        }).execute()
        anexo_id = (r.data or [{}])[0].get("id")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao registrar anexo: {e}")

    url = _anx_signed_upload(path)
    return {"ok": True, "anexo_id": anexo_id, "path": path,
            "modo": "signed" if url else "proxy", "url": url or ""}


@app.post("/chamados/anexos/{anexo_id}/upload")
async def anexo_upload_proxy(anexo_id: int, arquivo: UploadFile = File(...),
                             usuario: str = Depends(verificar_token)):
    """Fallback: sobe via backend quando a signed URL não rolou."""
    sb = get_supabase()
    r = sb.table("chamado_anexos").select("*").eq("id", anexo_id).eq("operador_email", usuario).limit(1).execute()
    if not r.data:
        raise HTTPException(status_code=404, detail="Anexo não encontrado.")
    row = r.data[0]
    data = await arquivo.read()
    lim = _anexo_limites(sb)
    if len(data) > int(lim["max_mb"]) * 1024 * 1024:
        raise HTTPException(status_code=400, detail=f"Arquivo maior que {lim['max_mb']} MB.")
    try:
        _anx_storage().upload(row["path"], data,
                              {"content-type": row.get("mime") or "application/octet-stream",
                               "upsert": "true"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao subir arquivo: {e}")
    sb.table("chamado_anexos").update({"tamanho": len(data)}).eq("id", anexo_id).execute()
    return {"ok": True}


@app.post("/chamados/anexos/{anexo_id}/ler")
async def anexo_ler(anexo_id: int, usuario: str = Depends(verificar_token)):
    """Lê o anexo UMA vez e guarda o resumo. Chamado logo depois do upload."""
    sb = get_supabase()
    r = sb.table("chamado_anexos").select("*").eq("id", anexo_id).eq("operador_email", usuario).limit(1).execute()
    if not r.data:
        raise HTTPException(status_code=404, detail="Anexo não encontrado.")
    row = r.data[0]
    lim = _anexo_limites(sb)
    orc = _orcamento_sessao(sb, row["sessao_id"], lim)

    try:
        data = _anx_download(row["path"])
    except Exception as e:
        sb.table("chamado_anexos").update({
            "status": "erro", "leitura": f"[arquivo não chegou no Storage: {e}]"}).eq("id", anexo_id).execute()
        raise HTTPException(status_code=400, detail="O arquivo não chegou no Storage. Tenta subir de novo.")

    res = _ler_anexo(get_claude(), data, row["nome"], orc, lim)
    upd = {"status": res["status"], "leitura": res["texto"],
           "paginas": res.get("paginas", 0), "img_lidas": res.get("img_lidas", 0),
           "tamanho": len(data)}
    sb.table("chamado_anexos").update(upd).eq("id", anexo_id).execute()
    resumo = (res["texto"] or "").strip().split("\n")[0][:180]
    return {"ok": True, "status": res["status"], "resumo": resumo, "leitura": res["texto"]}


@app.get("/chamados/anexos")
async def listar_anexos_sessao(sessao_id: str, usuario: str = Depends(verificar_token)):
    sb = get_supabase()
    r = sb.table("chamado_anexos").select("id,nome,status,tamanho,leitura")\
        .eq("sessao_id", sessao_id).eq("operador_email", usuario).order("id").execute()
    out = []
    for a in (r.data or []):
        out.append({"id": a["id"], "nome": a["nome"], "status": a["status"], "tamanho": a.get("tamanho") or 0,
                    "resumo": (a.get("leitura") or "").strip().split("\n")[0][:180]})
    return {"anexos": out}


@app.delete("/chamados/anexos/{anexo_id}")
async def remover_anexo(anexo_id: int, usuario: str = Depends(verificar_token)):
    """O dono remove enquanto o chamado ainda não foi aberto."""
    sb = get_supabase()
    r = sb.table("chamado_anexos").select("*").eq("id", anexo_id).eq("operador_email", usuario).limit(1).execute()
    if not r.data:
        raise HTTPException(status_code=404, detail="Anexo não encontrado.")
    row = r.data[0]
    if row.get("chamado_id"):
        raise HTTPException(status_code=400, detail="Esse anexo já faz parte de um chamado aberto.")
    _anx_remove(row["path"])
    sb.table("chamado_anexos").delete().eq("id", anexo_id).execute()
    return {"ok": True}


@app.get("/chamados/{chamado_id}/anexos")
async def anexos_do_chamado(chamado_id: int, usuario: str = Depends(verificar_token)):
    """Admin vê de todos; operador só dos próprios chamados. URL de download
    assinada, válida por 1h (o bucket é privado)."""
    sb = get_supabase()
    c = sb.table("chamados").select("operador_email").eq("id", chamado_id).limit(1).execute()
    if not c.data:
        raise HTTPException(status_code=404, detail="Chamado não encontrado.")
    if usuario not in ADMIN_EMAILS and c.data[0].get("operador_email") != usuario:
        raise HTTPException(status_code=403, detail="Sem acesso a esse chamado.")
    r = sb.table("chamado_anexos").select("id,nome,status,tamanho,leitura,path")\
        .eq("chamado_id", chamado_id).order("id").execute()
    out = []
    for a in (r.data or []):
        out.append({"id": a["id"], "nome": a["nome"], "status": a["status"],
                    "tamanho": a.get("tamanho") or 0,
                    "leitura": a.get("leitura") or "",
                    "url": _anx_signed_download(a["path"], 3600) or ""})
    return {"anexos": out}


@app.post("/chamados/anexos/limpar-orfaos")
async def limpar_anexos_orfaos(dias: int = 7, usuario: str = Depends(verificar_token)):
    """Conversa abandonada deixa anexo sem chamado. Limpeza manual (admin)."""
    if usuario not in ADMIN_EMAILS:
        raise HTTPException(status_code=403, detail="Só o admin.")
    from datetime import datetime, timezone, timedelta
    corte = (datetime.now(timezone.utc) - timedelta(days=max(1, int(dias)))).isoformat()
    sb = get_supabase()
    r = sb.table("chamado_anexos").select("id,path").is_("chamado_id", "null").lt("criado_em", corte).execute()
    for a in (r.data or []):
        _anx_remove(a["path"])
        sb.table("chamado_anexos").delete().eq("id", a["id"]).execute()
    return {"ok": True, "removidos": len(r.data or [])}


# ══════════════════════════════════════════════════════════════════════════
# DATASHEETS (chamado #6)
#
# O cliente pede a ficha técnica dos itens ofertados. Hoje o operador copia
# link por link no ChatGPT, espera e baixa cada PDF na mão.
#
# Import protegido, mesmo padrão do motor de preços: se `datasheet.py` ou
# `reportlab` faltarem no deploy, o app sobe normal e só estas rotas avisam.
# ══════════════════════════════════════════════════════════════════════════
try:
    import datasheet as _ds_mod
    _DATASHEET_OK = True
    _DATASHEET_ERRO = ""
except Exception as _e_ds:                                   # pragma: no cover
    _ds_mod = None
    _DATASHEET_OK = False
    _DATASHEET_ERRO = f"{type(_e_ds).__name__}: {_e_ds}"

DATASHEETS_BUCKET = os.environ.get("DATASHEETS_BUCKET", "datasheets")
_DS_TIMEOUT_HTTP = 20
_DS_UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
          "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")


def _ds_storage():
    return get_supabase().storage.from_(DATASHEETS_BUCKET)


def _ds_upload(path: str, dados: bytes, mime: str = "application/pdf"):
    _ds_storage().upload(path, dados, {"content-type": mime, "upsert": "true"})


def _ds_signed(path: str, expira: int = 3600) -> str:
    if not path:
        return ""
    try:
        r = _ds_storage().create_signed_url(path, expira) or {}
        return r.get("signedURL") or r.get("signedUrl") or ""
    except Exception:
        return ""


def _ds_remove(path: str):
    if not path:
        return
    try:
        _ds_storage().remove([path])
    except Exception:
        pass


# ── Logo ──────────────────────────────────────────────────────────────────
# Vive no config_kist em base64, não no repositório: trocar a marca não exige
# redeploy, e não há binário para subir no GitHub. Cache em memória porque são
# 22 KB que não mudam.
_DS_LOGO_CACHE = {"b64": None, "bytes": None}


def _ds_logo(sb) -> Optional[bytes]:
    """Logo da KIST, em duas fontes — nesta ordem:

      1. config_kist['datasheet_logo_b64'] (base64) — permite trocar a marca em
         runtime, sem redeploy;
      2. backend/assets/logo_kist.png — o arquivo no repositório.

    A fonte 2 existe porque transferir 12 KB de base64 por texto e' fragil: um
    unico caractere errado no meio do blob invalida o PNG inteiro, e o erro so'
    aparece na hora de desenhar. Arrastar o PNG no GitHub nao transcreve nada.
    Cache em memoria: sao 22 KB que nao mudam.
    """
    if _DS_LOGO_CACHE["bytes"] is not None:
        return _DS_LOGO_CACHE["bytes"]

    dados = None
    try:
        r = sb.table("config_kist").select("valor").eq("chave", "datasheet_logo_b64")\
            .limit(1).execute()
        b64 = (r.data or [{}])[0].get("valor") or ""
    except Exception:
        b64 = ""
    if b64.strip():
        try:
            dados = _b64.b64decode(b64.strip(), validate=True)
        except Exception:
            dados = None          # blob corrompido: cai pro arquivo em vez de quebrar

    if dados is None:
        caminho = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               "assets", "logo_kist.png")
        try:
            with open(caminho, "rb") as f:
                dados = f.read()
        except Exception:
            dados = None

    # Um PNG/JPEG de verdade, nao 3 bytes de sobra de um base64 truncado.
    if not dados or len(dados) < 512:
        return None
    _DS_LOGO_CACHE["bytes"] = dados
    return dados


# ── Rede ──────────────────────────────────────────────────────────────────
# Funções pequenas e injetadas no módulo do datasheet, para ele continuar
# testável sem rede. Toda falha vira exceção que o `datasheet.py` já trata.
# Cabecalho de navegador de verdade. Loja grande (Mercado Livre, Amazon,
# distribuidor com WAF) devolve 403 para requisicao "pelada" com so' o
# User-Agent. Sem isto, a pagina volta vazia, nao ha' imagem para extrair, e o
# item sai sem foto — foi o que aconteceu com a mazer e com o ML.
_DS_CABECALHOS = {
    "User-Agent": _DS_UA,
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,"
              "image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
}


def _ds_pagina(url: str) -> str:
    r = requests.get(url, timeout=_DS_TIMEOUT_HTTP, headers=_DS_CABECALHOS,
                     allow_redirects=True)
    r.raise_for_status()
    if not r.encoding or r.encoding.lower() == "iso-8859-1":
        r.encoding = r.apparent_encoding or "utf-8"
    texto = r.text[:400000]
    # O content-type NAO decide sozinho. Servidor mal configurado manda
    # "application/octet-stream" ou "text/plain" servindo HTML — e a pagina era
    # descartada sem ninguem olhar. Se o CORPO parece HTML, e' HTML.
    ct = (r.headers.get("content-type") or "").lower()
    parece_html = bool(re.search(r"<html|<meta|<body|<!doctype", texto[:4000], re.I))
    if not parece_html and ("html" not in ct and "xml" not in ct and "json" not in ct):
        return ""
    return texto


# ══════════════════════════════════════════════════════════════════════════
# MERCADO LIVRE — token de aplicacao
#
# O ML passou a exigir access token em TODA chamada, inclusive nos recursos
# publicos, e o token vale ~6h. Sem ele a API responde 401 e a pagina HTML cai
# no antibot ("Seguridad - Mercado Libre") — foi o que derrubou 36% da base.
#
# Credenciais por env var no Render (nenhuma no codigo):
#   ML_CLIENT_ID / ML_CLIENT_SECRET   -> obrigatorias
#   ML_REFRESH_TOKEN                  -> opcional; se houver, usa refresh_token
#                                        (mais duravel). Sem ela, tenta
#                                        client_credentials.
#
# Sem credencial configurada o motor NAO tenta a API e NAO tenta o HTML do ML:
# avisa que falta credencial. Bater na porta trancada e' so' gastar tempo.
# ══════════════════════════════════════════════════════════════════════════
ML_CLIENT_ID = os.environ.get("ML_CLIENT_ID", "").strip()
ML_CLIENT_SECRET = os.environ.get("ML_CLIENT_SECRET", "").strip()
ML_REFRESH_TOKEN = os.environ.get("ML_REFRESH_TOKEN", "").strip()

_ML_TOKEN = {"valor": "", "expira": 0.0, "erro": ""}


def ml_configurado() -> bool:
    return bool(ML_CLIENT_ID and ML_CLIENT_SECRET)


def _ml_token(forcar: bool = False) -> str:
    """Token do ML, em cache ate' 5 min antes de expirar."""
    if not ml_configurado():
        _ML_TOKEN["erro"] = ("credenciais do Mercado Livre não configuradas "
                             "(ML_CLIENT_ID / ML_CLIENT_SECRET no Render)")
        return ""
    if not forcar and _ML_TOKEN["valor"] and time.time() < _ML_TOKEN["expira"]:
        return _ML_TOKEN["valor"]

    corpo = {"client_id": ML_CLIENT_ID, "client_secret": ML_CLIENT_SECRET}
    if ML_REFRESH_TOKEN:
        corpo.update({"grant_type": "refresh_token", "refresh_token": ML_REFRESH_TOKEN})
    else:
        corpo["grant_type"] = "client_credentials"
    try:
        r = requests.post("https://api.mercadolibre.com/oauth/token", data=corpo,
                          timeout=_DS_TIMEOUT_HTTP,
                          headers={"Accept": "application/json",
                                   "Content-Type": "application/x-www-form-urlencoded"})
        d = r.json() if r.content else {}
        if r.status_code >= 400 or not d.get("access_token"):
            _ML_TOKEN["erro"] = (f"o Mercado Livre recusou as credenciais "
                                 f"(HTTP {r.status_code}: "
                                 f"{str(d.get('message') or d.get('error') or '')[:90]})")
            _ML_TOKEN["valor"] = ""
            return ""
        _ML_TOKEN["valor"] = d["access_token"]
        _ML_TOKEN["expira"] = time.time() + max(60, int(d.get("expires_in") or 21600) - 300)
        _ML_TOKEN["erro"] = ""
        return _ML_TOKEN["valor"]
    except Exception as e:
        _ML_TOKEN["erro"] = f"não consegui falar com o Mercado Livre ({type(e).__name__})"
        _ML_TOKEN["valor"] = ""
        return ""


def _ds_json(url: str):
    cab = {"User-Agent": _DS_UA, "Accept": "application/json"}
    e_ml = "api.mercadolibre.com" in (url or "")

    if e_ml:
        tok = _ml_token()
        if not tok:
            raise RuntimeError(_ML_TOKEN["erro"] or "sem token do Mercado Livre")
        cab["Authorization"] = f"Bearer {tok}"

    r = requests.get(url, timeout=_DS_TIMEOUT_HTTP, headers=cab)
    # 401 = token venceu antes da hora (troca de senha, secret girado). Renova
    # UMA vez e repete. Se falhar de novo, e' credencial, nao expiracao.
    if e_ml and r.status_code == 401:
        tok = _ml_token(forcar=True)
        if tok:
            cab["Authorization"] = f"Bearer {tok}"
            r = requests.get(url, timeout=_DS_TIMEOUT_HTTP, headers=cab)
    r.raise_for_status()
    return r.json()


def _ds_baixar(url: str, referer: str = "") -> bytes:
    # CDN que barra hotlink olha o Referer e espera a PAGINA DO PRODUTO, nao o
    # dominio da propria imagem. Eu mandava o dominio do CDN — inutil.
    _h = {"User-Agent": _DS_UA, "Accept": "image/avif,image/webp,image/*,*/*;q=0.8",
          "Accept-Language": "pt-BR,pt;q=0.9"}
    _base = referer or url or ""
    _m = re.match(r"(https?://[^/]+)", _base)
    if _m:
        _h["Referer"] = (referer or (_m.group(1) + "/"))
    r = requests.get(url, timeout=_DS_TIMEOUT_HTTP, stream=True, headers=_h)
    r.raise_for_status()
    dados, teto = b"", 8 * 1024 * 1024
    for pedaco in r.iter_content(64 * 1024):
        dados += pedaco
        if len(dados) > teto:
            break
    return dados


# ── Serialização ──────────────────────────────────────────────────────────
def _ds_para_front(row: dict, com_url: bool = True) -> dict:
    payload = row.get("payload") or {}
    return {
        "id": row.get("id"),
        "produto_id": row.get("produto_id"),
        "nome_produto": row.get("nome_produto") or "",
        "fabricante": row.get("fabricante") or "",
        "modelo": row.get("modelo") or "",
        "status": row.get("status") or "rascunho",
        "versao": row.get("versao") or 1,
        "critica": row.get("critica") or "",
        "imagem_origem": row.get("imagem_origem") or "",
        "tem_foto": bool(row.get("imagem_path")),
        "conteudo": payload.get("conteudo") or {},
        "identificacao": payload.get("identificacao") or {},
        "foto": payload.get("foto") or {},
        "avisos": payload.get("avisos") or [],
        "nome_arquivo": payload.get("nome_arquivo") or "",
        "modo": row.get("modo") or payload.get("modo") or "tecnico",
        "aprovado_por": row.get("aprovado_por") or "",
        "aprovado_em": str(row.get("aprovado_em") or ""),
        "pdf_url": _ds_signed(row.get("pdf_path") or "") if com_url else "",
    }


# Datasheet tecnico e apresentacao comercial sao DOCUMENTOS IRMAOS do mesmo
# item — nao dois modos do mesmo documento. Cada um tem a sua linha em
# `datasheets`, o seu vinculo no item e no produto, e o seu proprio cache.
# Pedir apresentacao nunca pode devolver o datasheet, e vice-versa.
DS_CAMPO = {"tecnico": "datasheet_id", "comercial": "apresentacao_id"}


def _ds_campo(modo: str) -> str:
    return DS_CAMPO.get((modo or "tecnico").strip().lower(), "datasheet_id")


# Link que NAO identifica um produto: e' resultado de busca, listagem ou
# categoria. Serve para o operador achar o item, nao para identificar o item.
_RX_LINK_BUSCA = re.compile(
    r"/sch/|/search|/busca|/procura|/s\?|/b\?|[?&](_nkw|q|query|busca|search|k)=",
    re.I)


def _ds_link_e_busca(link: str) -> bool:
    return bool(_RX_LINK_BUSCA.search(link or ""))


def _ds_chave_link(link: str) -> str:
    """Normaliza o link para servir de chave de cache.

    Tira querystring de rastreamento e barra final — o mesmo anuncio chega com
    `?utm_source=` diferente toda vez e viraria um datasheet novo a cada busca.

    LINK DE BUSCA NAO VIRA CHAVE. Achado no backtest de 29/07: tres itens
    diferentes com link `ebay.com/sch/i.html?_nkw=SIGA-CC1`, `...SIGA-CT1` e
    `...CABO HDMI` colapsavam todos para `ebay.com/sch/i.html` depois de tirar a
    querystring — a MESMA chave. Aprovar o documento de um faria os outros
    puxarem ele do cache: documento errado na frente do cliente, em silencio.
    Sem chave, cada um gera o seu; a busca por descricao continua valendo.
    """
    l = (link or "").strip().lower()
    if not l or _ds_link_e_busca(l):
        return ""
    l = re.sub(r"[?#].*$", "", l)
    l = re.sub(r"/+$", "", l)
    return l[:400]


# ── Endpoints ─────────────────────────────────────────────────────────────
class DatasheetGerarIn(BaseModel):
    item: dict
    item_id: Optional[int] = None
    produto_id: Optional[int] = None
    datasheet_id: Optional[int] = None      # regeração
    pistas: Optional[str] = ""              # correção do operador na identificação
    critica: Optional[str] = ""             # o que ele reprovou
    fonte_texto: Optional[str] = ""
    imagem_b64: Optional[str] = ""          # foto que o operador subiu
    imagem_url: Optional[str] = ""          # ou o link dela
    contato_rodape: Optional[str] = ""
    # "tecnico"   = identifica + busca web + valida (o de hoje)
    # "comercial" = so' o prompt do Fabio, cru, sem consultar fonte externa
    modo: Optional[str] = "tecnico"


@app.post("/datasheets/gerar")
async def datasheet_gerar(payload: DatasheetGerarIn, usuario: str = Depends(verificar_token)):
    """Gera (ou regera) o datasheet de UM item.

    Síncrono e lento de propósito — o operador pediu e está esperando, igual ao
    /conferir. O 'gerar todos' é uma fila no FRONT, um item por vez: com um
    worker só no Render, disparar 20 buscas em paralelo derruba o backend (foi
    exatamente o incidente de 20/07).
    """
    if not _DATASHEET_OK:
        raise HTTPException(503, f"Módulo de datasheet indisponível ({_DATASHEET_ERRO})")
    sb = get_supabase()
    logo = _ds_logo(sb)
    if not logo:
        raise HTTPException(503, "Logo da KIST não encontrado. Coloque o arquivo em "
                                 "backend/assets/logo_kist.png OU rode sql/logo_kist.sql "
                                 "(chave config_kist['datasheet_logo_b64']).")

    item = payload.item or {}
    _modo = (payload.modo or "tecnico").strip().lower()
    if _modo not in ("tecnico", "comercial"):
        _modo = "tecnico"
    anterior, ident_ant, pistas_ant, ds_row = None, None, "", None
    if payload.datasheet_id:
        r = sb.table("datasheets").select("*").eq("id", payload.datasheet_id).limit(1).execute()
        if r.data:
            ds_row = r.data[0]
            anterior = (ds_row.get("payload") or {}).get("conteudo") or None
            # A identificação anterior viaja junto: quando o operador só troca a
            # FOTO, o módulo reaproveita as duas e não chama o modelo — sem ela,
            # o fabricante/modelo já confirmados seriam recalculados do zero.
            ident_ant = (ds_row.get("payload") or {}).get("identificacao") or None
            # As pistas da versão anterior. O front manda `pistas` em toda
            # chamada (é estado de tela e nunca é limpo), então só a COMPARAÇÃO
            # com estas revela se o operador pediu algo novo agora.
            pistas_ant = (ds_row.get("payload") or {}).get("pistas") or ""

    # Foto que o operador mandou vence a busca — ele é a hierarquia superior.
    img_op = None
    if (payload.imagem_b64 or "").strip():
        try:
            img_op = _b64.b64decode(re.sub(r"^data:[^,]+,", "", payload.imagem_b64.strip()))
        except Exception:
            raise HTTPException(400, "Não consegui ler a imagem enviada.")
    elif (payload.imagem_url or "").strip():
        try:
            img_op = _ds_baixar(payload.imagem_url.strip())
        except Exception as e:
            raise HTTPException(400, f"Não consegui baixar a imagem desse link ({type(e).__name__}).")

    # Foto que o operador JÁ tinha enviado numa versão anterior. Ele não vai
    # reenviar o arquivo a cada crítica — sem isto, corrigir o texto derruba a
    # imagem e ele fica num vaivém que nunca fecha. Falhar aqui não impede a
    # geração: o pior caso é a busca de foto normal.
    img_preservada = None
    if (not img_op and ds_row
            and (ds_row.get("imagem_origem") or "") == "operador"
            and ds_row.get("imagem_path")):
        try:
            img_preservada = _ds_storage().download(ds_row["imagem_path"])
        except Exception:
            img_preservada = None

    try:
        r = _ds_mod.gerar(
            get_claude(), item, logo,
            baixar=_ds_baixar, buscar_pagina=_ds_pagina, buscar_json=_ds_json,
            fonte_texto=payload.fonte_texto or "",
            pistas=payload.pistas or "", critica=payload.critica or "",
            anterior=anterior, ident_anterior=ident_ant,
            pistas_anterior=pistas_ant,
            imagem_operador=img_op, imagem_preservada=img_preservada,
            contato_rodape=payload.contato_rodape or "",
            modo=_modo,
            system_comercial=(_ds_mod.prompt_comercial(sb) if _modo == "comercial" else ""),
        )
    except Exception as e:
        raise HTTPException(502, f"Falhou ao gerar o datasheet: {type(e).__name__}: {e}")

    # Identificação ambígua: não grava nada, devolve as perguntas.
    if r.get("precisa_operador"):
        return {"precisa_operador": True, "etapa": "identificacao",
                "identificacao": r.get("identificacao") or {},
                "avisos": r.get("avisos") or []}

    conteudo = r["conteudo"]
    ident = r["identificacao"]
    versao = int((ds_row or {}).get("versao") or 0) + 1
    base = f"ds/{payload.produto_id or 'novo'}/{int(time.time())}_{versao}"

    pdf_path = f"{base}.pdf"
    _ds_upload(pdf_path, r["pdf_bytes"], "application/pdf")
    img_path = ""
    if r.get("imagem_bytes"):
        ext = {"image/jpeg": "jpg", "image/png": "png",
               "image/webp": "webp", "image/gif": "gif"}.get(r["foto"].get("mime"), "png")
        img_path = f"{base}_foto.{ext}"
        _ds_upload(img_path, r["imagem_bytes"], r["foto"].get("mime") or "image/png")

    # Histórico: a versão reprovada + a crítica ficam guardadas. É o que permite
    # a regeração corrigir em vez de sortear de novo.
    historico = list((ds_row or {}).get("historico") or [])
    if ds_row:
        historico.append({
            "versao": ds_row.get("versao"),
            "critica": payload.critica or "",
            "conteudo": (ds_row.get("payload") or {}).get("conteudo") or {},
            # O diagnostico da foto TEM que viajar no historico. Sem ele, quando
            # o operador regera informando a imagem na mao, o motivo da falha
            # original some — e some justamente o dado que conserta a busca.
            "foto": (ds_row.get("payload") or {}).get("foto") or {},
            "origem": (ds_row.get("payload") or {}).get("origem") or {},
            "pdf_path": ds_row.get("pdf_path") or "",
        })
        historico = historico[-10:]

    linha = {
        "produto_id": payload.produto_id,
        "chave_link": _ds_chave_link(item.get("link_fornecedor") or ""),
        "chave_desc": _norm_entrada(item.get("descricao_final")
                                    or item.get("descricao_original") or ""),
        "nome_produto": conteudo.get("nome_produto") or ident.get("nome_produto") or "Item",
        "fabricante": ident.get("fabricante") or "",
        "modelo": conteudo.get("modelo") or ident.get("modelo") or "",
        "payload": {"conteudo": conteudo, "identificacao": ident,
                    "foto": r.get("foto") or {}, "origem": r.get("origem") or {},
                    "avisos": r.get("avisos") or [],
                    "nome_arquivo": r.get("nome_arquivo") or "",
                    "pistas": r.get("pistas") or "",
                    "modo": _modo},
        "imagem_path": img_path or None,
        "imagem_origem": (r.get("foto") or {}).get("origem") or "ausente",
        "pdf_path": pdf_path,
        "modo": _modo,
        "status": "rascunho",
        "versao": versao,
        "critica": payload.critica or "",
        "historico": historico,
        "criado_por": usuario,
    }

    if ds_row:
        _ds_remove(ds_row.get("pdf_path") or "")
        sb.table("datasheets").update(linha).eq("id", ds_row["id"]).execute()
        ds_id = ds_row["id"]
    else:
        ins = sb.table("datasheets").insert(linha).execute()
        ds_id = (ins.data or [{}])[0].get("id")

    novo = dict(linha, id=ds_id)
    saida = _ds_para_front(novo)
    saida["precisa_operador"] = False
    saida["etapa"] = "revisao"
    saida["tem_foto"] = bool(img_path)
    return saida


class DatasheetAprovarIn(BaseModel):
    item_id: Optional[int] = None
    produto_id: Optional[int] = None
    descricao: Optional[str] = ""      # p/ achar a linha do banco quando nao houve match
    proposta_id: Optional[int] = None
    indice: Optional[int] = None       # posicao do item na proposta salva


@app.post("/datasheets/{ds_id}/aprovar")
async def datasheet_aprovar(ds_id: int, payload: DatasheetAprovarIn,
                            usuario: str = Depends(verificar_token)):
    """Aprovado = o datasheet vira DADO do produto e não precisa mais revisão.

    A partir daqui, a próxima proposta com o mesmo produto puxa do cache.
    """
    from datetime import datetime, timezone
    sb = get_supabase()
    r = sb.table("datasheets").select("*").eq("id", ds_id).limit(1).execute()
    if not r.data:
        raise HTTPException(404, "Datasheet não encontrado.")

    # PRODUTO: se o front nao mandou (item que nao casou com o banco), procura a
    # linha pela descricao EXATA. So' procura — nao cria. Rascunho nao alimenta o
    # banco de precos; quem cria a linha e' o /upsert-precos, no CSV. Quando ela
    # nascer la', o datasheet_id vai junto (viaja no `origem`).
    _pid = payload.produto_id
    if not _pid and (payload.descricao or "").strip():
        try:
            _r = sb.table("produtos").select("id")\
                .ilike("descricao", _ilike_literal(payload.descricao.strip()))\
                .limit(1).execute()
            if _r.data:
                _pid = _r.data[0]["id"]
        except Exception:
            _pid = None

    upd = {"status": "aprovado", "aprovado_por": usuario,
           "aprovado_em": datetime.now(timezone.utc).isoformat(),
           "atualizado_em": datetime.now(timezone.utc).isoformat()}
    if _pid:
        upd["produto_id"] = _pid
    sb.table("datasheets").update(upd).eq("id", ds_id).execute()

    # Vínculo nos dois lados. Falhar aqui não derruba a aprovação — o PDF já
    # existe e o operador já pode baixar.
    # O campo depende do MODO da linha aprovada: datasheet e apresentacao sao
    # irmaos e nunca ocupam a mesma coluna.
    _campo = _ds_campo(r.data[0].get("modo") or "tecnico")
    vinculos = {"produto": False, "item": False, "campo": _campo}
    if _pid:
        try:
            sb.table("produtos").update({_campo: ds_id}).eq("id", _pid).execute()
            vinculos["produto"] = True
        except Exception:
            pass

    # ITEM: por id quando existe; senao pela (proposta, indice), que e' como o
    # item se identifica ANTES de virar CSV. Sem isto, aprovar durante a
    # conferencia nao persistia nada e o operador regerava o mesmo documento.
    if payload.item_id:
        try:
            sb.table("itens_proposta").update({_campo: ds_id})\
              .eq("id", payload.item_id).execute()
            vinculos["item"] = True
        except Exception:
            pass
    elif payload.proposta_id is not None and payload.indice is not None:
        try:
            _it = sb.table("itens_proposta").select("id")\
                .eq("proposta_id", payload.proposta_id)\
                .order("id").execute()
            _linhas = _it.data or []
            if 0 <= payload.indice < len(_linhas):
                sb.table("itens_proposta")\
                  .update({_campo: ds_id})\
                  .eq("id", _linhas[payload.indice]["id"]).execute()
                vinculos["item"] = True
        except Exception:
            pass

    row = dict(r.data[0], **upd)
    saida = _ds_para_front(row)
    saida["vinculos"] = vinculos
    saida["produto_id"] = _pid
    return saida


@app.post("/datasheets/{ds_id}/reprovar")
async def datasheet_reprovar(ds_id: int, payload: dict,
                             usuario: str = Depends(verificar_token)):
    """Marca a reprovação e guarda a crítica.

    Não regera aqui: quem regera é o /datasheets/gerar com `datasheet_id` +
    `critica`, para o modelo receber a versão anterior junto do que deu errado.
    """
    sb = get_supabase()
    critica = (payload.get("critica") or "").strip()
    if not critica:
        raise HTTPException(400, "Escreva o que precisa ser ajustado — sem isso a "
                                 "regeração é sorteio, não correção.")
    r = sb.table("datasheets").select("*").eq("id", ds_id).limit(1).execute()
    if not r.data:
        raise HTTPException(404, "Datasheet não encontrado.")
    sb.table("datasheets").update({"status": "reprovado", "critica": critica})\
      .eq("id", ds_id).execute()
    return {"ok": True, "id": ds_id, "critica": critica}


@app.get("/datasheets/{ds_id}")
async def datasheet_ver(ds_id: int, usuario: str = Depends(verificar_token)):
    sb = get_supabase()
    r = sb.table("datasheets").select("*").eq("id", ds_id).limit(1).execute()
    if not r.data:
        raise HTTPException(404, "Datasheet não encontrado.")
    return _ds_para_front(r.data[0])


@app.get("/datasheets")
async def datasheet_buscar(produto_id: Optional[int] = None,
                           link: Optional[str] = None,
                           descricao: Optional[str] = None,
                           modo: Optional[str] = "tecnico",
                           usuario: str = Depends(verificar_token)):
    """Cache: existe documento APROVADO DESTE MODO para este item?

    O `modo` e' filtro obrigatorio. Sem ele, pedir apresentacao devolvia o
    datasheet tecnico que ja' existia — foi exatamente o que o Leonardo viu.

    Ordem das chaves: produto do banco → link normalizado → descricao
    normalizada. E' o que evita regerar (e repagar) o mesmo documento.
    """
    sb = get_supabase()
    _m = (modo or "tecnico").strip().lower()
    if _m not in DS_CAMPO:
        _m = "tecnico"

    def _q():
        return sb.table("datasheets").select("*")\
                 .eq("status", "aprovado").eq("modo", _m)

    if produto_id:
        r = _q().eq("produto_id", produto_id).order("id", desc=True).limit(1).execute()
        if r.data:
            return {"achou": True, "por": "produto", "modo": _m,
                    "datasheet": _ds_para_front(r.data[0])}
    if link:
        r = _q().eq("chave_link", _ds_chave_link(link))\
                .order("id", desc=True).limit(1).execute()
        if r.data:
            return {"achou": True, "por": "link", "modo": _m,
                    "datasheet": _ds_para_front(r.data[0])}
    if descricao:
        r = _q().eq("chave_desc", _norm_entrada(descricao))\
                .order("id", desc=True).limit(1).execute()
        if r.data:
            return {"achou": True, "por": "descricao", "modo": _m,
                    "datasheet": _ds_para_front(r.data[0])}
    return {"achou": False, "modo": _m, "datasheet": None}


@app.post("/ml/notificacoes")
@app.get("/ml/notificacoes")
async def ml_notificacoes(request: Request):
    """Webhook do Mercado Livre. Recebe e devolve 200. So' isso, de proposito.

    O cadastro da aplicacao no ML EXIGE uma URL de notificacao, mas a Kist nao
    vende nem publica por la' — a integracao so' LE ficha de produto. Nao ha'
    nada a processar.

    Precisa existir mesmo assim: sem endpoint no ar, o ML tenta entregar,
    acumula falha e acaba desativando as notificacoes (e reclamando da
    aplicacao). Um 200 rapido encerra o assunto.

    SEM autenticacao de propriedade: quem chama e' o ML, nao um operador, e
    portanto nao tem Bearer. Como nada e' lido nem gravado, nao ha' o que
    proteger — no maximo alguem gasta um 200 a toa.
    """
    try:
        corpo = await request.json()
    except Exception:
        corpo = None
    # Log leve, so' para saber que chegou. Nao guarda em banco.
    if isinstance(corpo, dict) and corpo.get("topic"):
        print(f"[ml] notificacao ignorada · topic={corpo.get('topic')} "
              f"resource={str(corpo.get('resource'))[:80]}")
    return {"ok": True}


class DatasheetZipIn(BaseModel):
    ids: List[int]
    nome: Optional[str] = ""


@app.post("/datasheets/zip")
async def datasheet_zip(payload: DatasheetZipIn, usuario: str = Depends(verificar_token)):
    """Baixa varios documentos de uma vez, num ZIP.

    Um botao por tipo (datasheets / apresentacoes). Sem isto o operador abriria
    uma aba por item e o navegador bloquearia as janelas depois da segunda.
    """
    import zipfile
    from fastapi.responses import StreamingResponse

    ids = [int(i) for i in (payload.ids or []) if i]
    if not ids:
        raise HTTPException(400, "Nenhum documento para baixar.")

    sb = get_supabase()
    r = sb.table("datasheets").select("id,nome_produto,modelo,pdf_path,payload,modo")\
        .in_("id", ids).execute()
    linhas = r.data or []
    if not linhas:
        raise HTTPException(404, "Documentos não encontrados.")

    buf = io.BytesIO()
    dentro, faltaram = 0, []
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        usados = set()
        for linha in linhas:
            caminho = linha.get("pdf_path") or ""
            if not caminho:
                faltaram.append(linha.get("nome_produto") or linha.get("id"))
                continue
            try:
                dados = _ds_storage().download(caminho)
            except Exception:
                faltaram.append(linha.get("nome_produto") or linha.get("id"))
                continue
            nome = ((linha.get("payload") or {}).get("nome_arquivo")
                    or f"datasheet_{linha.get('id')}.pdf")
            # Dois itens com o mesmo nome de arquivo se sobrescreveriam no ZIP.
            base = nome
            n = 2
            while nome in usados:
                nome = base.replace(".pdf", f"_{n}.pdf")
                n += 1
            usados.add(nome)
            z.writestr(nome, dados)
            dentro += 1
        if faltaram:
            z.writestr("_NAO_ENTRARAM.txt",
                       "Estes documentos não puderam ser incluídos:\n" +
                       "\n".join(f"- {x}" for x in faltaram))

    if dentro == 0:
        raise HTTPException(502, "Não consegui baixar nenhum dos PDFs.")

    buf.seek(0)
    apelido = re.sub(r"[^A-Za-z0-9_.-]+", "_", (payload.nome or "documentos_kist"))[:60]
    return StreamingResponse(
        buf, media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{apelido}.zip"',
                 "X-Documentos": str(dentro), "X-Faltaram": str(len(faltaram))})


@app.delete("/datasheets/{ds_id}")
async def datasheet_excluir(ds_id: int, usuario: str = Depends(verificar_token)):
    sb = get_supabase()
    r = sb.table("datasheets").select("*").eq("id", ds_id).limit(1).execute()
    if not r.data:
        raise HTTPException(404, "Datasheet não encontrado.")
    row = r.data[0]
    _ds_remove(row.get("pdf_path") or "")
    _ds_remove(row.get("imagem_path") or "")
    for h in (row.get("historico") or []):
        _ds_remove(h.get("pdf_path") or "")
    sb.table("datasheets").delete().eq("id", ds_id).execute()
    return {"ok": True}


def _ds_marcar_aprovados(sb, itens):
    """Segundo caminho: datasheet APROVADO que casa por descricao ou por link.

    POR QUE ISTO EXISTE:
        O primeiro caminho depende de `produtos.datasheet_id`, que so' existe
        depois que o produto entrou no banco — e o banco so' e' alimentado
        quando a proposta vira CSV. Numa proposta em RASCUNHO, um datasheet ja'
        aprovado ficaria invisivel e o operador geraria tudo de novo.

        Datasheet aprovado e' trabalho conferido pelo operador. Ele nao pode
        sumir porque a proposta ainda nao virou CSV.

    Nao sobrescreve o que o primeiro caminho ja' resolveu.
    """
    alvos = [it for it in (itens or [])
             if not (it.get("datasheet_id") and it.get("apresentacao_id"))]
    if not alvos:
        return
    descs, links = [], []
    for it in alvos:
        d = _norm_entrada(it.get("descricao_final") or it.get("descricao_original") or "")
        if d and d not in descs:
            descs.append(d)
        l = _ds_chave_link(it.get("link_fornecedor") or "")
        if l and l not in links:
            links.append(l)
    if not descs and not links:
        return

    # Indexado por (modo, chave): datasheet e apresentacao do MESMO item tem a
    # mesma chave_desc e a mesma chave_link — so' o modo os separa.
    por_desc, por_link = {}, {}
    try:
        if descs:
            r = sb.table("datasheets").select("id,chave_desc,modo")\
                .eq("status", "aprovado").in_("chave_desc", descs[:200]).execute()
            for x in (r.data or []):
                por_desc.setdefault((x.get("modo") or "tecnico", x.get("chave_desc")), x["id"])
        if links:
            r = sb.table("datasheets").select("id,chave_link,modo")\
                .eq("status", "aprovado").in_("chave_link", links[:200]).execute()
            for x in (r.data or []):
                por_link.setdefault((x.get("modo") or "tecnico", x.get("chave_link")), x["id"])
    except Exception:
        return                  # selo e' conforto, nao derruba a extracao

    for it in alvos:
        chave_l = _ds_chave_link(it.get("link_fornecedor") or "")
        chave_d = _norm_entrada(it.get("descricao_final")
                                or it.get("descricao_original") or "")
        for modo, campo in DS_CAMPO.items():
            if it.get(campo):
                continue
            # Link primeiro: e' identidade mais forte que texto.
            ds_id = por_link.get((modo, chave_l)) or por_desc.get((modo, chave_d))
            if ds_id:
                it[campo] = ds_id


def _ds_marcar_itens(sb, itens):
    """Acende o selo de datasheet nos itens que acabaram de casar com o banco.

    POR QUE UMA CONSULTA EXTRA E NAO UMA COLUNA NA RPC:
        `candidatos_trgm_lote` devolve um TABLE com tipo fixo. Adicionar coluna
        exige DROP + CREATE da funcao que e' o coracao do matching. Uma consulta
        indexada nos poucos ids que casaram custa quase nada e nao arrisca nada.

    REGRA DE VINCULO (Leonardo, jul/2026):
      • veredito "mesmo" (semanticamente identico) -> vincula sozinho;
      • qualquer outro caso -> o datasheet fica DISPONIVEL na ficha e o front
        pergunta se vincula, na hora do "usar esta".
    Regua diferente da do dinheiro de proposito: custo precisa de exatidao
    porque variante muda preco; datasheet precisa de identidade tecnica, e dois
    textos diferentes do mesmo produto compartilham o documento legitimamente.
    """
    ids = []
    for it in itens or []:
        pid = ((it.get("banco") or {}) or {}).get("produto_id")
        if pid and pid not in ids:
            ids.append(pid)
    if not ids:
        return
    try:
        r = sb.table("produtos").select("id,datasheet_id,apresentacao_id")\
              .in_("id", ids).execute()
    except Exception:
        return                      # selo e' conforto, nao pode derrubar extracao
    mapa = {x["id"]: x for x in (r.data or [])}
    for it in itens or []:
        ficha = it.get("banco") or {}
        linha = mapa.get(ficha.get("produto_id")) or {}
        confirmado = (ficha.get("veredito") or "") == "mesmo"
        for campo in ("datasheet_id", "apresentacao_id"):
            ds_id = linha.get(campo)
            if not ds_id:
                continue
            ficha[campo] = ds_id
            if confirmado:
                it[campo] = ds_id                       # identidade tecnica confirmada
            else:
                it[f"{campo}_disponivel"] = ds_id       # existe, mas o operador decide
